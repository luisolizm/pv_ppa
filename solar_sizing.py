import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import math
import copy
import requests
from io import BytesIO
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak,
)
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors as rl_color
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

st.set_page_config(page_title="Sizing Tool",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={}  # opcional
)

# ═════════════════════════════════════════════════════════════════════════════
# AUTENTICACIÓN — password session-based
# La contraseña se lee de st.secrets["APP_PASSWORD"] (archivo .streamlit/secrets.toml).
# Fallback a variable de entorno APP_PASSWORD para entornos sin secrets.toml.
# NUNCA hardcodear la contraseña en el código fuente.
# ═════════════════════════════════════════════════════════════════════════════
import os as _os
try:
    _APP_PASSWORD = st.secrets["APP_PASSWORD"]
except Exception:
    _APP_PASSWORD = _os.environ.get("APP_PASSWORD", "soliz")

def _check_auth():
    """Muestra pantalla de login si no hay sesión activa. Detiene el render."""
    if st.session_state.get("authenticated"):
        return  # ya autenticado — continuar normalmente

    # Pantalla de login centrada
    st.markdown("""
<style>
  .login-wrap {
    max-width: 380px; margin: 8vh auto 0; padding: 36px 40px;
    background: #111318; border: 1px solid #1e2230; border-radius: 16px;
  }
  .login-title { font-size: 22px; font-weight: 700; color: #f1f5f9;
                 margin-bottom: 4px; text-align: center; }
  .login-sub   { font-size: 13px; color: #475569; text-align: center;
                 margin-bottom: 28px; }
</style>
<div class="login-wrap">
  <div class="login-title">⚡ Sizing Tool</div>
  <div class="login-sub">Ingresa la contraseña para continuar</div>
</div>
""", unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 1.2, 1])
    with col_c:
        pwd = st.text_input("Contraseña", type="password",
                            placeholder="••••••••", label_visibility="collapsed",
                            key="_login_pwd")
        if st.button("Entrar", use_container_width=True, type="primary"):
            if pwd == _APP_PASSWORD:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
    st.stop()   # detiene el resto del render hasta que haya sesión válida

_check_auth()


# Forzar dark mode
st.markdown("""
<style>
  /* ── Fuentes ── */
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

  html, body, [class*="css"], button, label, p, span, div,
  .stMarkdown, .stTextInput, .stNumberInput, .stSlider, .stRadio, .stCheckbox {
    font-family: 'Inter', sans-serif !important;
  }

  /* ── Fondo global oscuro ── */
  [data-testid="stAppViewContainer"] { background: #0a0c10; }
  .main { background: #0a0c10; }
  [data-testid="stSidebar"] {
    background: #0e1117 !important;
    border-right: 1px solid #1e2230 !important;
  }
  [data-testid="stSidebar"] * { color: #e2e8f0 !important; }

  /* ── Ocultar iconos de colapso sidebar ── */
  [data-testid="collapsedControl"] span,
  [data-testid="stSidebarCollapseButton"] span,
  span.material-symbols-rounded,
  span.material-symbols-outlined {
    font-size:0!important; visibility:hidden!important;
    width:0!important; height:0!important;
  }

  /* ── Tipografia monoespaciada ── */
  .mono { font-family: 'JetBrains Mono', monospace !important; }

  /* ── Cabecera ── */
  .app-title { font-size:28px; font-weight:700; color:#f1f5f9; letter-spacing:-0.6px; }
  .app-sub   { font-size:13px; color:#64748b; margin-top:-4px; margin-bottom:1.5rem; }

  /* ── Section header ── */
  .section-header {
    font-size:11px; font-weight:600; color:#475569;
    text-transform:uppercase; letter-spacing:0.12em;
    margin:1.6rem 0 0.8rem; padding-bottom:6px;
    border-bottom:1px solid #1e2230;
  }

  /* ── Tabs ── */
  .stTabs [data-baseweb="tab-list"] { background:#111318; border-radius:10px; padding:4px; gap:4px; }
  .stTabs [data-baseweb="tab"]      { background:transparent; border-radius:8px; color:#64748b; font-weight:500; font-size:14px; }
  .stTabs [aria-selected="true"]    { background:#f59e0b !important; color:#0a0c10 !important; font-weight:600 !important; }

  /* ── Cajas informativas ── */
  .info-box { background:#111318; border-left:3px solid #f59e0b; border-radius:0 8px 8px 0; padding:10px 14px; font-size:13px; color:#94a3b8; margin-bottom:1rem; }
  .nasa-box { background:#0b1623; border-left:3px solid #3b82f6; border-radius:0 8px 8px 0; padding:10px 14px; font-size:13px; color:#93c5fd; margin-bottom:1rem; }
  .warn-box { background:#111318; border-left:3px solid #f43f5e; border-radius:0 8px 8px 0; padding:10px 14px; font-size:13px; color:#94a3b8; margin-bottom:1rem; }

  /* ── TOR Hero ── */
  .tor-hero {
    background: linear-gradient(135deg, #111318 0%, #0d0f15 100%);
    border:1px solid #1e2230; border-radius:14px;
    padding:20px 24px; margin-bottom:1.4rem;
  }
  .tor-hero .th-project { font-size:10px; color:#475569; text-transform:uppercase; letter-spacing:0.10em; margin-bottom:4px; }
  .tor-hero .th-meta    { font-size:12px; color:#64748b; margin-bottom:16px; }
  .tor-hero .th-grid    { display:grid; grid-template-columns:repeat(4,1fr); gap:14px 20px; }
  .tor-hero .th-item    { display:flex; flex-direction:column; }
  .tor-hero .th-label   { font-size:10px; color:#475569; text-transform:uppercase; letter-spacing:0.06em; margin-bottom:3px; }
  .tor-hero .th-val     { font-size:20px; font-weight:700; color:#f59e0b; font-family:'JetBrains Mono',monospace; line-height:1.1; word-break:break-word; }
  .tor-hero .th-unit    { font-size:11px; color:#64748b; margin-top:3px; }

  /* ── Badges PR ── */
  .pr-badge  { display:inline-flex; align-items:center; gap:6px; padding:4px 10px; border-radius:20px; font-size:12px; font-weight:500; }
  .pr-green  { background:#052e16; color:#4ade80; border:1px solid #166534; }
  .pr-yellow { background:#1c1a04; color:#facc15; border:1px solid #713f12; }
  .pr-red    { background:#1f0a0a; color:#f87171; border:1px solid #7f1d1d; }

  /* ── Panel card ── */
  .panel-card           { background:#111318; border:1px solid #1e2230; border-radius:12px; padding:14px 18px; margin-bottom:1rem; }
  .panel-card .pc-title { font-size:10px; color:#475569; text-transform:uppercase; letter-spacing:0.08em; margin-bottom:10px; }
  .panel-card .pc-grid  { display:grid; grid-template-columns:1fr 1fr; gap:6px 16px; }
  .panel-card .pc-item  { display:flex; flex-direction:column; }
  .panel-card .pc-label { font-size:10px; color:#475569; }
  .panel-card .pc-val   { font-size:14px; font-weight:600; color:#f59e0b; font-family:'JetBrains Mono',monospace; }

  /* ── Snap cards KPI ── */
  .snap-card {
    background:#111318; border:1px solid #1e2230; border-radius:12px;
    padding:16px 12px; text-align:center;
    display:flex; flex-direction:column; align-items:center; justify-content:center;
    min-height:110px;
  }
  .snap-card .sc-label { font-size:10px; color:#475569; text-transform:uppercase; letter-spacing:0.08em; margin-bottom:6px; line-height:1.3; }
  .snap-card .sc-val   { font-size:clamp(14px,1.4vw,22px); font-weight:700; font-family:'JetBrains Mono',monospace; word-break:break-word; overflow-wrap:anywhere; line-height:1.2; max-width:100%; color:#e2e8f0; }
  .snap-card .sc-sub   { font-size:10px; color:#475569; margin-top:4px; line-height:1.3; }

  /* ── st.metric ── */
  [data-testid="stMetric"] { background:#111318; border:1px solid #1e2230; border-radius:12px; padding:14px 12px !important; text-align:center; }
  [data-testid="stMetricValue"] { font-family:'JetBrains Mono',monospace !important; font-size:clamp(13px,1.3vw,20px) !important; font-weight:700 !important; word-break:break-word !important; overflow-wrap:anywhere !important; white-space:normal !important; line-height:1.2 !important; color:#e2e8f0 !important; }
  [data-testid="stMetricLabel"] { font-family:'Inter',sans-serif !important; font-size:11px !important; color:#475569 !important; text-transform:uppercase; letter-spacing:0.06em; }

  /* ── Sidebar inputs ── */
  [data-testid="stSidebar"] input[type="number"],
  [data-testid="stSidebar"] input[type="text"],
  [data-testid="stSidebar"] .stTextInput input,
  [data-testid="stSidebar"] .stNumberInput input {
    background-color:#0a0c10 !important; color:#e2e8f0 !important;
    border:1px solid #2d3748 !important; border-radius:6px !important;
  }
  [data-testid="stSidebar"] input:focus { border-color:#f59e0b !important; box-shadow:0 0 0 2px rgba(245,158,11,0.2) !important; outline:none !important; }
  [data-testid="stSidebar"] [data-baseweb="input"],
  [data-testid="stSidebar"] [data-baseweb="base-input"] { background-color:#0a0c10 !important; }
  [data-testid="stSidebar"] [data-baseweb="input"] input,
  [data-testid="stSidebar"] [data-baseweb="base-input"] input { color:#e2e8f0 !important; background-color:#0a0c10 !important; caret-color:#f59e0b !important; }
  [data-testid="stSidebar"] button[data-testid="stNumberInputStepDown"],
  [data-testid="stSidebar"] button[data-testid="stNumberInputStepUp"] { background-color:#111318 !important; color:#e2e8f0 !important; border-color:#2d3748 !important; }
  [data-testid="stSidebar"] [data-baseweb="select"] > div { background-color:#0a0c10 !important; border-color:#2d3748 !important; color:#e2e8f0 !important; }
  [data-testid="stSidebar"] label { color:#cbd5e1 !important; }

  .stDataFrame { border-radius:10px; overflow:hidden; }
</style>
""", unsafe_allow_html=True)
# ── Constants ─────────────────────────────────────────────────────────────────
MONTHS     = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
MONTH_DAYS = [31,28,31,30,31,30,31,31,30,31,30,31]
DEFAULT_IRR      = [4.8,5.2,5.9,6.1,5.8,5.4,5.3,5.2,4.7,4.5,4.6,4.5]
NASA_START, NASA_END = 2005, 2024
# Factor de emisión del Sistema Eléctrico Nacional (SEN) mexicano.
# Fuente: SEMARNAT / CRE — Aviso Factor de Emisión SEN 2024, publicado 28-Feb-2025.
# Valor oficial: 0.444 tCO₂e/MWh = 0.444 kg CO₂e/kWh.
# Actualizar anualmente conforme al aviso SEMARNAT más reciente.
CO2_FACTOR_KG_KWH = 0.444   # kg CO₂e/kWh  (SEN 2024 · SEMARNAT/CRE · 28-Feb-2025)


PLOT_LAYOUT = dict(
    paper_bgcolor="#0f1117", plot_bgcolor="#13151f",
    font=dict(family="Inter, sans-serif", color="#94a3b8", size=12),
    xaxis=dict(gridcolor="#2a2d3a", linecolor="#2a2d3a", tickcolor="#2a2d3a"),
    yaxis=dict(gridcolor="#2a2d3a", linecolor="#2a2d3a", tickcolor="#2a2d3a"),
    margin=dict(l=10, r=10, t=30, b=10),
)
AMBER="#f59e0b"; TEAL="#14b8a6"; ROSE="#f43f5e"; BLUE="#3b82f6"; VIOLET="#8b5cf6"


# ── NASA POWER ────────────────────────────────────────────────────────────────
# ── NASA POWER (2005–2024) ────────────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN DE INPUTS — deben definirse antes de cualquier tab
# ═════════════════════════════════════════════════════════════════════════════

def _bisection_irr(cash_flows: list) -> float | None:
    """
    Calcula la TIR de una serie de flujos de caja por bisección numérica.
    cash_flows[0] debe ser la inversión inicial (negativa).
    Criterio de convergencia doble: intervalo < 1e-10 O |NPV| < 1e-9.
    Retorna la TIR en porcentaje (%), o None si no converge o no hay solución.
    """
    def _npv(rr: float) -> float:
        return sum(c / (1 + rr) ** t for t, c in enumerate(cash_flows))
    try:
        lo, hi = -0.99, 5.0
        npv_lo = _npv(lo)
        npv_hi = _npv(hi)
        if npv_lo * npv_hi > 0:
            return None          # Sin cambio de signo → sin TIR en el rango
        for _ in range(200):
            mid = (lo + hi) / 2
            fm  = _npv(mid)
            if (hi - lo) / 2 < 1e-10:
                return mid * 100
            if abs(fm) < 1e-9:
                return mid * 100
            if npv_lo * fm < 0:
                hi = mid
            else:
                lo = mid
                npv_lo = fm
        return ((lo + hi) / 2) * 100
    except Exception:
        return None


def _safe(value, fallback=0.0, min_val=None, max_val=None, label="valor"):
    """Sanitiza valor numérico: reemplaza NaN/inf por fallback y clampea al rango."""
    try:
        v = float(value)
        if not math.isfinite(v): return fallback
        if min_val is not None and v < min_val: return min_val
        if max_val is not None and v > max_val: return max_val
        return v
    except (TypeError, ValueError): return fallback


def _validate_recibo_inputs(monthly_cons, monthly_tar):
    if not monthly_cons or not monthly_tar: return False, "Datos de consumo/tarifa vacíos."
    if len(monthly_cons) != 12 or len(monthly_tar) != 12: return False, "Se requieren 12 meses de datos."
    if sum(monthly_cons) <= 0: return False, "El consumo anual es cero. Ingresa al menos un mes con consumo."
    if any(c < 0 for c in monthly_cons): return False, "El consumo no puede ser negativo."
    if any(t <= 0 for t in monthly_tar): return False, "La tarifa debe ser mayor a cero en todos los meses."
    return True, ""


def _validate_bess_inputs(peak_kw, daily_kwh, useful_kwh, pcs_kw):
    if peak_kw <= 0: return False, "La demanda pico debe ser mayor a 0 kW."
    if daily_kwh <= 0: return False, "El consumo diario debe ser mayor a 0 kWh."
    if useful_kwh <= 0: return False, "La energía útil es 0. Revisa parámetros de dimensionamiento."
    if pcs_kw <= 0: return False, "La potencia del PCS es 0. Revisa los parámetros."
    return True, ""


@st.cache_data(ttl=7200, show_spinner=False)
def get_nasa_power_irradiance(lat: float, lon: float):
    """
    Devuelve una tupla (irr_media, irr_por_anio):
      irr_media   : list[12]  — promedio climatológico mensual (kWh/m²/día)
      irr_por_anio: dict[int, list[12]] — irradiancia mensual de cada año
                    {2005: [v_ene, v_feb, …, v_dic], 2006: …}
    Valores inválidos (< 0 o None) se sustituyen por el promedio del mes.
    """
    url = (
        "https://power.larc.nasa.gov/api/temporal/monthly/point"
        "?parameters=ALLSKY_SFC_SW_DWN&community=RE"
        f"&longitude={lon}&latitude={lat}&format=JSON"
        f"&start={NASA_START}&end={NASA_END}"
    )

    try:
        r = requests.get(url, timeout=25)
        r.raise_for_status()
        data = r.json()

        raw = (data.get("properties", {})
                   .get("parameter", {})
                   .get("ALLSKY_SFC_SW_DWN", {}))

        if not raw:
            raise ValueError("NASA POWER no devolvió datos.")

        # ── 1. Construir serie por año ────────────────────────────────────────
        irr_por_anio: dict[int, list] = {}   # {año: [None]*12}
        monthly_sum   = [0.0] * 12
        monthly_count = [0]   * 12

        for key, value in raw.items():
            if len(key) != 6:
                continue
            try:
                year      = int(key[:4])
                month_idx = int(key[4:6]) - 1
                if not (0 <= month_idx <= 11):
                    continue
                if value is None or value < 0:
                    continue
                irr_por_anio.setdefault(year, [None] * 12)
                irr_por_anio[year][month_idx] = float(value)
                monthly_sum[month_idx]   += value
                monthly_count[month_idx] += 1
            except (ValueError, IndexError, TypeError):
                continue

        # ── 2. Promedio climatológico mensual ────────────────────────────────
        irr_media = [
            round(monthly_sum[i] / monthly_count[i], 4) if monthly_count[i] > 0
            else DEFAULT_IRR[i]
            for i in range(12)
        ]

        # ── 3. Rellenar meses faltantes en cada año con el promedio del mes ──
        for year, meses in irr_por_anio.items():
            for i in range(12):
                if meses[i] is None:
                    meses[i] = irr_media[i]

        return irr_media, irr_por_anio

    except requests.exceptions.Timeout:
        raise Exception("⏱️ Timeout: NASA POWER tardó demasiado.")
    except requests.exceptions.ConnectionError:
        raise Exception("🌐 Sin conexión a internet.")
    except Exception as e:
        raise Exception(f"❌ Error NASA POWER: {str(e)[:100]}")


# ── P90 riguroso: percentil 10 de la distribución de generaciones anuales ─────
def compute_p90(irr_por_anio: dict, kwp: float, pr: float) -> tuple:
    """
    Simula la generación anual para cada año histórico y devuelve:
      (p50_real, p90_real, gen_por_anio)
      p50_real    : mediana de las generaciones anuales (kWh)
      p90_real    : percentil 10 de las generaciones anuales (kWh)
                    — el sistema supera este valor el 90% de los años —
      gen_por_anio: dict {año: generación_kWh}
    Si no hay datos históricos, devuelve None para ambos percentiles.
    """
    if not irr_por_anio:
        return None, None, {}

    gen_por_anio = {}
    for year, meses in sorted(irr_por_anio.items()):
        gen_anual = sum(
            kwp * meses[m] * pr * MONTH_DAYS[m]
            for m in range(12)
        )
        gen_por_anio[year] = gen_anual

    valores = sorted(gen_por_anio.values())
    p50_real = float(np.percentile(valores, 50))
    p90_real = float(np.percentile(valores, 10))   # percentil 10 = excedido 90% de años
    return p50_real, p90_real, gen_por_anio

# ── Funciones de cálculo cacheadas ───────────────────────────────────────────

@st.cache_data(show_spinner=False)
def calc_sizing_area(area_total: float, occ_factor: int,
                     panel_wp: int, panel_area: float,
                     irr_vals: tuple, effective_pr: float) -> dict:
    """Sizing por área: número de paneles, kWp y generación mensual."""
    n_panels  = int(math.floor(area_total * occ_factor / 100 / panel_area))
    kwp       = n_panels * panel_wp / 1000
    area_util = area_total * occ_factor / 100
    area_used = n_panels * panel_area
    monthly_gen = [round(kwp * irr_vals[m] * effective_pr * MONTH_DAYS[m], 1) for m in range(12)]
    return dict(n_panels=n_panels, kwp=kwp, area_util=area_util,
                area_used=area_used, monthly_gen=monthly_gen,
                annual_gen=sum(monthly_gen))

@st.cache_data(show_spinner=False)
def calc_sizing_recibo_detallado(monthly_cons: tuple, monthly_tarifas: tuple,
                                  solar_pct: int, sizing_strategy: str,
                                  panel_wp: int, panel_area: float,
                                  irr_vals: tuple, effective_pr: float,
                                  occ_factor: int) -> dict:
    """
    Sizing con consumos y tarifas mensuales reales (12 valores cada uno).
    Devuelve kWp, generación, ahorro real mes a mes y tarifa media ponderada.
    """
    target_m = [c * solar_pct / 100 for c in monthly_cons]
    kwp_por_mes = [
        target_m[m] / (irr_vals[m] * effective_pr * MONTH_DAYS[m])
        if irr_vals[m] > 0 else 0 for m in range(12)
    ]
    kwp_raw  = max(kwp_por_mes) if "Peor" in sizing_strategy else sum(kwp_por_mes) / 12
    n_panels = int(math.ceil(kwp_raw * 1000 / panel_wp))
    kwp      = n_panels * panel_wp / 1000
    area_used = n_panels * panel_area
    area_util = area_used / (occ_factor / 100) if occ_factor > 0 else area_used
    monthly_gen = [round(kwp * irr_vals[m] * effective_pr * MONTH_DAYS[m], 1) for m in range(12)]

    # Ahorro real mes a mes: energía cubierta × tarifa de ese mes
    energia_cubierta = [min(monthly_gen[m], monthly_cons[m]) for m in range(12)]
    ahorro_mensual   = [energia_cubierta[m] * monthly_tarifas[m] for m in range(12)]
    excedente        = [monthly_gen[m] - monthly_cons[m] for m in range(12)]
    cobertura_pct    = [
        min(100.0, monthly_gen[m] / monthly_cons[m] * 100) if monthly_cons[m] > 0 else 0.0
        for m in range(12)
    ]

    consumo_anual = sum(monthly_cons)
    gasto_actual  = sum(monthly_cons[m] * monthly_tarifas[m] for m in range(12))
    tarifa_media_pond = gasto_actual / consumo_anual if consumo_anual > 0 else 0

    return dict(
        n_panels=n_panels, kwp=kwp, area_util=area_util, area_used=area_used,
        monthly_gen=monthly_gen, annual_gen=sum(monthly_gen),
        monthly_cons=list(monthly_cons),
        energia_cubierta=energia_cubierta,
        ahorro_mensual=ahorro_mensual,
        ahorro_anual=sum(ahorro_mensual),
        excedente=excedente,
        cobertura_pct=cobertura_pct,
        cobertura_anual=sum(energia_cubierta) / consumo_anual * 100 if consumo_anual > 0 else 0,
        gasto_actual=gasto_actual,
        tarifa_media_pond=tarifa_media_pond,
        monthly_tarifas=list(monthly_tarifas),
    )


@st.cache_data(show_spinner=False)
def calc_sizing_recibo_kwp(monthly_cons: tuple, monthly_tarifas: tuple,
                            kwp_manual: float,
                            panel_wp: int, panel_area: float,
                            irr_vals: tuple, effective_pr: float,
                            occ_factor: int = 75) -> dict:
    """
    Sizing con kWp fijado manualmente por el usuario.
    Redondea al número entero de paneles más cercano y calcula
    cobertura, generación y ahorro reales mes a mes.
    occ_factor: porcentaje de ocupación del área disponible (%), configurable.
    FIX: antes usaba factor fijo del 75%; ahora usa el valor del usuario.
    """
    n_panels  = max(1, round(kwp_manual * 1000 / panel_wp))
    kwp       = n_panels * panel_wp / 1000
    area_used = n_panels * panel_area
    # FIX: área bruta requerida usa el occ_factor configurado por el usuario,
    # no el valor hardcodeado al 75%.
    occ_ratio = max(occ_factor, 1) / 100
    area_util = area_used / occ_ratio

    monthly_gen      = [round(kwp * irr_vals[m] * effective_pr * MONTH_DAYS[m], 1) for m in range(12)]
    energia_cubierta = [min(monthly_gen[m], monthly_cons[m]) for m in range(12)]
    ahorro_mensual   = [energia_cubierta[m] * monthly_tarifas[m] for m in range(12)]
    excedente        = [monthly_gen[m] - monthly_cons[m] for m in range(12)]
    cobertura_pct    = [
        min(100.0, monthly_gen[m] / monthly_cons[m] * 100) if monthly_cons[m] > 0 else 0.0
        for m in range(12)
    ]
    consumo_anual     = sum(monthly_cons)
    gasto_actual      = sum(monthly_cons[m] * monthly_tarifas[m] for m in range(12))
    tarifa_media_pond = gasto_actual / consumo_anual if consumo_anual > 0 else 0.0

    return dict(
        n_panels=n_panels, kwp=kwp, area_util=area_util, area_used=area_used,
        monthly_gen=monthly_gen, annual_gen=sum(monthly_gen),
        monthly_cons=list(monthly_cons),
        energia_cubierta=energia_cubierta,
        ahorro_mensual=ahorro_mensual,
        ahorro_anual=sum(ahorro_mensual),
        excedente=excedente,
        cobertura_pct=cobertura_pct,
        cobertura_anual=sum(energia_cubierta) / consumo_anual * 100 if consumo_anual > 0 else 0,
        gasto_actual=gasto_actual,
        tarifa_media_pond=tarifa_media_pond,
        monthly_tarifas=list(monthly_tarifas),
    )


@st.cache_data(show_spinner=False)
def calc_financial_model(annual_gen: float, kwp: float, inversion_usd: float,
                         tarifa_efectiva: float, inflation: float,
                         discount_rate: float, panel_degradation: float,
                         vida_util: int, usd_to_mxn: float,
                         om_pct: float = 1.0) -> dict:
    """Modelo financiero completo: VPN, TIR, payback, LCOE, flujos anuales.
    om_pct: porcentaje de la inversión MXN destinado a O&M anual (default 1.0 %).
    """
    years = list(range(1, vida_util + 1))
    r     = discount_rate / 100
    inv_mxn = inversion_usd * usd_to_mxn

    gen_proj      = [annual_gen * (1 - panel_degradation / 100) ** (y - 1) for y in years]
    tarifas_y     = [tarifa_efectiva * (1 + inflation / 100) ** (y - 1) for y in years]
    flujo_nominal = [gen_proj[i] * tarifas_y[i] for i in range(len(years))]
    # FIX: O&M ahora usa om_pct en vez de valor hardcodeado al 1%
    om_anual      = [inv_mxn * (om_pct / 100) * (1 + inflation / 100) ** (y - 1) for y in years]
    flujo_neto    = [flujo_nominal[i] - om_anual[i] for i in range(len(years))]
    factor_desc   = [1 / (1 + r) ** y for y in years]
    flujo_desc    = [flujo_neto[i] * factor_desc[i] for i in range(len(years))]

    acum_nominal, acum = [], -inv_mxn
    for fn in flujo_neto:
        acum += fn; acum_nominal.append(acum)
    acum_desc, acum = [], -inv_mxn
    for fd in flujo_desc:
        acum += fd; acum_desc.append(acum)

    vpn = acum_desc[-1]

    # TIR — función compartida a nivel de módulo (_bisection_irr)
    tir = _bisection_irr([-inv_mxn] + flujo_neto)

    # FIX: payback con interpolación lineal (antes devolvía el año entero, ahora da fracción)
    pb_simple = None
    for i, v in enumerate(acum_nominal):
        if v >= 0:
            prev = acum_nominal[i - 1] if i > 0 else -inv_mxn
            pb_simple = round(years[i] - 1 + (-prev) / (v - prev), 1)
            break

    pb_disc = None
    for i, v in enumerate(acum_desc):
        if v >= 0:
            prev = acum_desc[i - 1] if i > 0 else -inv_mxn
            pb_disc = round(years[i] - 1 + (-prev) / (v - prev), 1)
            break

    total_gen_desc  = sum(gen_proj[i] * factor_desc[i] for i in range(len(years)))
    total_cost_desc = inv_mxn + sum(om_anual[i] * factor_desc[i] for i in range(len(years)))
    lcoe = total_cost_desc / total_gen_desc if total_gen_desc > 0 else 0

    return dict(
        vpn=vpn, tir=tir, pb_simple=pb_simple, pb_disc=pb_disc, lcoe=lcoe,
        years=years, gen_proj=gen_proj, tarifas_y=tarifas_y,
        flujo_nominal=flujo_nominal, om_anual=om_anual, flujo_neto=flujo_neto,
        factor_desc=factor_desc, flujo_desc=flujo_desc,
        acum_nominal=acum_nominal, acum_desc=acum_desc, inv_mxn=inv_mxn,
    )


@st.cache_data(show_spinner=False)
def calc_ppa_result(gen1: float, inv_usd: float, precio_ppa: float,
                    plazo: int, wacc_pct: float, esc_ppa: float,
                    deg: float, om_pct: float, inf_om: float,
                    seg_pct: float, usd_mx: float, equity_pct: float,
                    tasa_deuda: float, plazo_deuda: int, con_fin: bool,
                    vida_util_total: int = 25) -> dict:
    """Resultado financiero PPA para un plazo dado.
    vida_util_total: vida útil del sistema (años). Se usa para calcular el valor
    residual al final del contrato PPA si plazo < vida_util_total.
    """
    inv_mxn    = inv_usd * usd_mx
    equity_mxn = inv_mxn * (equity_pct / 100)
    deuda_mxn  = inv_mxn - equity_mxn
    if con_fin and tasa_deuda > 0 and plazo_deuda > 0 and deuda_mxn > 0:
        r_d = tasa_deuda / 100
        serv_deuda = deuda_mxn * r_d / (1 - (1 + r_d) ** (-plazo_deuda))
    else:
        serv_deuda = 0.0; deuda_mxn = 0.0; equity_mxn = inv_mxn
    r      = wacc_pct / 100
    years  = list(range(1, plazo + 1))
    gen_y  = [gen1 * (1 - deg / 100) ** i for i in range(plazo)]
    prec_y = [precio_ppa * (1 + esc_ppa / 100) ** i for i in range(plazo)]
    ing_y  = [gen_y[i] * prec_y[i] for i in range(plazo)]
    om_y   = [inv_mxn * om_pct  / 100 * (1 + inf_om / 100) ** i for i in range(plazo)]
    seg_y  = [inv_mxn * seg_pct / 100 * (1 + inf_om / 100) ** i for i in range(plazo)]
    deu_y  = [serv_deuda if y <= plazo_deuda else 0.0 for y in years]
    fn_y   = [ing_y[i] - om_y[i] - seg_y[i] - deu_y[i] for i in range(plazo)]
    fd_y   = [fn_y[i] / (1 + r) ** years[i] for i in range(plazo)]

    # Valor residual del sistema al final del contrato PPA
    # Si el contrato es más corto que la vida útil, el activo sigue generando valor.
    # Se estima como VPN de los flujos futuros post-contrato usando una anuidad con
    # crecimiento (fórmula de Gordon) que incorpora tanto el escalador PPA como la
    # degradación anual del panel — evita sobreestimar el valor al asumir flujo constante.
    anios_restantes = max(0, vida_util_total - plazo)
    if anios_restantes > 0 and r > 0:
        gen_post  = gen_y[-1] * (1 - deg / 100)       # generación año plazo+1
        prec_post = prec_y[-1] * (1 + esc_ppa / 100)  # precio PPA escalado
        om_post   = om_y[-1]  * (1 + inf_om / 100)    # O&M escalado
        seg_post  = seg_y[-1] * (1 + inf_om / 100)
        fn_post   = gen_post * prec_post - om_post - seg_post

        # Tasa de crecimiento neta del flujo post-contrato:
        # el ingreso crece con (escalador_ppa - degradación), los costos con inf_om.
        g_ingreso = (esc_ppa / 100) - (deg / 100)   # puede ser negativo si deg > esc_ppa
        g_costos  = inf_om / 100
        # NOTA: cuando g_ingreso < 0 (degradación supera el escalador PPA), los ingresos
        # post-contrato decrecen año a año. _gordon_pv lo maneja correctamente via suma
        # finita; el valor residual puede resultar negativo o bajo, lo cual es matemáticamente
        # válido y conservador. No se clampea a cero para no ocultar proyectos no viables.

        # Anuidad con crecimiento compuesto (Gordon generalizado):
        # PV = F1 * [(1 - ((1+g)/(1+r))^n) / (r - g)]   si r ≠ g
        # PV = F1 * n / (1+r)                             si r ≈ g  (límite exacto)
        # Ponderamos g entre ingresos y costos para el flujo neto.
        # Como fn_post puede ser negativo si g_costos > g_ingreso, calculamos
        # cada componente por separado para mayor precisión.
        def _gordon_pv(f1: float, g: float, n: int, discount: float) -> float:
            """VPN de una anuidad con crecimiento g durante n períodos, traída a hoy.
            Usa la suma exacta de n términos en todos los casos para evitar
            divergencia cuando g >= discount (escalador > WACC).
            La fórmula cerrada de Gordon diverge cuando g >= r, por lo que
            siempre calculamos la suma finita directamente.
            """
            if n <= 0:
                return 0.0
            # Suma finita exacta: PV = sum_{t=1}^{n} f1*(1+g)^(t-1) / (1+r)^t
            # = f1/(1+r) * sum_{t=0}^{n-1} ((1+g)/(1+r))^t
            ratio = (1 + g) / (1 + discount)
            if abs(ratio - 1.0) < 1e-9:
                return f1 * n / (1 + discount)
            return f1 / (1 + discount) * (1 - ratio ** n) / (1 - ratio)

        pv_ingresos = _gordon_pv(gen_post * prec_post, g_ingreso, anios_restantes, r)
        pv_costos   = _gordon_pv(om_post + seg_post,   g_costos,  anios_restantes, r)
        # _gordon_pv devuelve el VPN de los flujos post-contrato llevado al instante t=plazo.
        # Se descuenta un período adicional de `plazo` años para traerlo a t=0.
        valor_residual = (pv_ingresos - pv_costos) / (1 + r) ** plazo
    else:
        valor_residual = 0.0

    vpn = -equity_mxn + sum(fd_y) + valor_residual

    # TIR — función compartida a nivel de módulo (_bisection_irr)
    tir = _bisection_irr([-equity_mxn] + fn_y)

    # Payback simple — acumulado sobre flujos nominales
    pb = None
    acum_pb = -equity_mxn
    for i, fn in enumerate(fn_y):
        prev_acum = acum_pb
        acum_pb  += fn
        if acum_pb >= 0:
            pb = round(years[i] - 1 + (-prev_acum) / (acum_pb - prev_acum), 1)
            break

    # Payback descontado — acumulado sobre flujos descontados (fd_y)
    pb_disc = None
    acum_disc = -equity_mxn
    for i, fd in enumerate(fd_y):
        prev_disc = acum_disc
        acum_disc += fd
        if acum_disc >= 0:
            pb_disc = round(years[i] - 1 + (-prev_disc) / (acum_disc - prev_disc), 1)
            break

    return dict(vpn=vpn, tir=tir, pb=pb, pb_disc=pb_disc, ing_total=sum(ing_y),
                fn_y=fn_y, fd_y=fd_y, ing_y=ing_y, om_y=om_y,
                seg_y=seg_y, gen_y=gen_y, prec_y=prec_y, deu_y=deu_y,
                equity_mxn=equity_mxn, inv_mxn=inv_mxn, years=years,
                valor_residual=valor_residual)


@st.cache_data(show_spinner=False)
def calc_precio_minimo(gen1: float, inv_usd: float, plazo: int,
                       wacc_pct: float, esc_ppa: float, deg: float,
                       om_pct: float, inf_om: float, seg_pct: float,
                       usd_mx: float, equity_pct: float,
                       tasa_deuda: float, plazo_deuda: int, con_fin: bool,
                       vida_util_total: int = 25):
    """Precio mínimo PPA (VPN=0) por bisección. Cacheado."""
    lo, hi = 0.01, 20.0
    def vpn_at(p):
        return calc_ppa_result(gen1, inv_usd, p, plazo, wacc_pct, esc_ppa,
                               deg, om_pct, inf_om, seg_pct, usd_mx,
                               equity_pct, tasa_deuda, plazo_deuda, con_fin,
                               vida_util_total)["vpn"]
    if vpn_at(hi) < 0: return None
    for _ in range(80):
        mid = (lo+hi)/2
        if vpn_at(mid) >= 0: hi = mid
        else: lo = mid
    return round((lo+hi)/2, 4)

def build_tor_text(proj_name, proj_date, proj_loc, proj_notes,
                   panel_wp, panel_eff, panel_largo_mm, panel_ancho_mm, panel_peso_kg,
                   panel_area, n_panels, kwp,
                   pr_pct, irr_vals, monthly_gen, annual_gen,
                   p50, p90, co2_saved,
                   inversion, ahorro1, payback,
                   gen_por_anio=None):
    lines = [
        "═══════════════════════════════════════════════════════════",
        f"  TÉRMINOS DE REFERENCIA — PRE-SIZING FOTOVOLTAICO",
        "═══════════════════════════════════════════════════════════",
        f"  Proyecto  : {proj_name}",
        f"  Fecha     : {proj_date}",
        f"  Ubicación : {proj_loc}",
        f"  Fuente irr: NASA POWER climatología {NASA_START}–{NASA_END}",
        "",
        "",
        "─── PARÁMETROS DEL PANEL (referencia) ──────────────────────",
        f"  Potencia pico (Pmax) : {panel_wp} Wp",
        f"  Eficiencia           : {panel_eff:.1f} %",
        f"  Dimensiones          : {panel_largo_mm} × {panel_ancho_mm} mm",
        f"  Área unitaria        : {panel_area:.4f} m²",
        f"  Peso                 : {panel_peso_kg} kg",
        f"  Carga estructural    : {n_panels * panel_peso_kg * 1.35:,.0f} kg  ({n_panels * panel_peso_kg * 1.35 * 9.8 / 1000:.2f} kN)  ({n_panels * panel_peso_kg * 1.35 * 9.8 / 1000 / max(n_panels * panel_area, 0.01):.3f} kN/m² sobre área instalada)  [factor ×1.35 montura+BOS]",
        "",
        "─── RESULTADOS DEL SIZING ──────────────────────────────────",
        f"  Paneles estimados    : {n_panels} unidades",
        f"  Capacidad pico       : {kwp:.2f} kWp",
        f"  PR asumido           : {pr_pct:.1f} %",
        f"  Generación P50 (irr. media)  : {p50:,.0f} kWh/año",
        f"  Generación P90       : {f'{p90:,.0f} kWh/año  (percentil 10 · {len(gen_por_anio)} años NASA POWER {NASA_START}–{NASA_END})' if p90 else 'No disponible — cargar datos NASA POWER'}",
        f"  CO₂ evitado/año      : {co2_saved/1000:,.2f} t  (factor {CO2_FACTOR_KG_KWH} kg CO₂e/kWh · SEN 2024 · SEMARNAT/CRE 28-Feb-2025)",
        "",
        "─── IRRADIANCIA MENSUAL (kWh/m²/día) ──────────────────────",
    ]
    for i, m in enumerate(MONTHS):
        lines.append(f"  {m:>3} : {irr_vals[i]:.2f}   →   Gen P50: {monthly_gen[i]:>7,.0f} kWh")
    lines += [
        f"  {'TOTAL':>3}        →   Gen P50: {annual_gen:>7,.0f} kWh/año",
        "",
    ]
    if proj_notes.strip():
        lines += ["─── NOTAS / ALCANCE ────────────────────────────────────────",
                  f"  {proj_notes}", ""]
    lines += [
        "─── ESPECIFICACIONES TÉCNICAS REQUERIDAS EN COTIZACIÓN ─────",
        "  MÓDULOS FOTOVOLTAICOS:",
        "  • Certificación IEC 61215 (calificación de diseño) · IEC 61730 (seguridad) · UL 61730 si aplica",
        "  • Clase de aplicación A · Tensión de sistema ≥ 1000 V DC",
        "  • Eficiencia mínima ≥ 21% (tecnología TOPCon/HJT actuales)",
        "  • Degradación: ≤ 2% año 1 · ≤ 0.4%/año años 2–25 (TOPCon) / ≤ 0.5%/año (PERC)",
        "  • Resistencia mecánica: carga frontal ≥ 5400 Pa (IEC 61215)",
        "  • Garantía de producto ≥ 12 años · Potencia lineal ≥ 25 años (P90 ≥ 80% Pmax)",
        "  • IP67 mínimo en caja de conexiones",
        "  • IEC 62716 resistencia a amoniaco — exigible en zonas agrícolas / ganaderas",
        "  • IEC 61701 categoría 6 resistencia a niebla salina — exigible en zonas costeras",
        "",
        "  INVERSORES:",
        "  • Certificación IEC 62109-1/-2 (seguridad) · Anti-islanding IEEE 1547 / NOM-001-SEDE",
        "  • THD de corriente ≤ 3% a potencia nominal (IEEE 519)",
        "  • IP66 mínimo · IP67 recomendado para intemperie",
        "  • Protección contra arco eléctrico DC (AFCI) — NEC 690.11",
        "  • Rapid Shutdown conforme NEC 690.12 (aplica en techos de edificio con acceso de emergencia)",
        "  • Eficiencia ≥ 98% Euro o CEC a potencia nominal",
        "  • Garantía mínima 5 años (extendible a 10–20 años)",
        "  • Sistemas >500 kWp: justificar topología string / centralizado / optimizadores",
        "    según uniformidad de superficie, orientaciones y nivel de sombreado",
        "",
        "  ESTRUCTURA Y MONTAJE:",
        "  • Aluminio 6005-T5 o acero galvanizado ASTM A653 (ZF275 mínimo)",
        "  • Diseño estructural conforme ASCE 7 / MDOC-CFE (viento, sismo, nieve si aplica)",
        "  • Resistencia a corrosión: niebla salina ≥ 1000 h (ISO 9227)",
        "  • Certificación de anclajes por ingeniero estructural acreditado (DRO o Perito)",
        "",
        "  SISTEMA ELÉCTRICO Y PROTECCIONES:",
        "  • Cableado DC: H1Z2Z2-K o equivalente (IEC 62930) · Tensión ≥ 1500 V",
        "  • Protecciones AC conformes NOM-001-SEDE / NEC 690",
        "  • Puesta a tierra conforme NOM-022-STPS y NEC 690.47",
        "  • Seccionadores y fusibles certificados UL 4703 / IEC 60269",
        "",
        "  MONITOREO:",
        "  • Datalogger con acceso remoto SCADA o plataforma cloud · SLA ≥ 99%",
        "  • Resolución de registro ≤ 15 min · Protocolo Modbus TCP o SunSpec",
        "  • Variables mínimas: potencia AC/DC, energía acumulada, V/I por MPPT,",
        "    temperatura de módulo, irradiancia (piranómetro clase 2), alarmas",
        "  • Exportación de datos en CSV o API REST",
        "",
        "  VERIFICACIONES E INSPECCIONES REQUERIDAS:",
        "  a) Pre-Puesta en Marcha (antes de energizar):",
        "    – Polaridad, continuidad y aislamiento DC ≥ 1 MΩ a 500 V (IEC 62446-1 · IEC 61557-2)",
        "    – Inspección visual de montaje, torque y sellado de conectores MC4",
        "    – Continuidad de puesta a tierra (< 1 Ω) y revisión de protecciones AC",
        "  b) Puesta en Marcha (commissioning):",
        "    – Curvas I-V por string vs STC ±3% (IEC 62446-1)",
        "    – Termografía IR de módulos y conexiones bajo carga (IEC TS 62446-3)",
        "    – Verificación anti-islanding y protecciones de reconexión (IEEE 1547)",
        "    – Prueba de potencia y PR inicial (tolerancia ±3%)",
        "    – Validación de monitoreo con registro mínimo 48 h",
        "  c) Inspecciones Periódicas:",
        "    – Limpieza semestral o según análisis de soiling",
        "    – Termografía IR anual de módulos, inversores y cableado (IEC TS 62446-3)",
        "    – Prueba de aislamiento DC anual (IEC 62446-1)",
        "    – Análisis de rendimiento vs P90 (PR objetivo ≥ 75%) (IEC 61724-1)",
        "    – Revisión anual de firmware de inversores y datalogger",
        "",
        "  ENTREGABLES OBLIGATORIOS DEL PROVEEDOR:",
        "  • Simulación PVSyst / Helioscope (P50 y P90) con:",
        "    – Análisis de sombras 3D (horizon profile + obstáculos cercanos)",
        "    – Estudio de soiling (pérdida por suciedad según zona climática)",
        "    – Desglose de pérdidas del sistema (PR explicado por componente)",
        "  • Planos eléctricos y estructurales firmados por DRO o Perito acreditado",
        "  • Memoria de cálculo estructural",
        "  • Protocolo de commissioning firmado (curvas I-V, termografía, aislamiento DC, PR)",
        "  • Dictamen UVIE (Unidad de Verificación de Instalaciones Eléctricas) conforme",
        "    NOM-001-SEDE — obligatorio para contrato CFE en mayoría de estados.",
        "    Nota: puede gestionarlo el EPC o el cliente; tiene costo independiente.",
        "  • Inspección de obra municipal / STPS si aplica (instalaciones industriales)",
        "  • Trámite de interconexión CFE (Pequeña Escala o Generación Distribuida)",
        "  • Permiso CRE si aplica (> 0.5 MW)",
        "  • Manual de O&M con plan de inspecciones periódicas",
        "  • Garantía de ejecución EPC ≥ 2 años post-puesta en marcha",
        "    (financiamientos bancarios exigen mínimo 3 años)",
        "  • Seguro de Responsabilidad Civil vigente durante la obra",
        "  • Performance bond / fianza de cumplimiento ≥ 10% del contrato",
        "    (obligatorio para proyectos > 100 kWp o con financiamiento bancario)",
        "",
        "─── CONSIDERACIONES GENERALES ──────────────────────────────",
        "  • Los valores son estimados de pre-sizing (±15%).",
        "  • El P90 se calcula como percentil 10 de la generación simulada",
        f"    con {len(gen_por_anio) if gen_por_anio else 'N/A'} años de irradiancia real NASA POWER ({NASA_START}–{NASA_END}).",
        "  • La ingeniería detallada y la simulación definitiva son responsabilidad del proveedor.",
        "  • Verificar disponibilidad y capacidad de red CFE en el punto de interconexión.",
        "═══════════════════════════════════════════════════════════",
    ]
    return "\n".join(lines)


# ── Helpers PDF ───────────────────────────────────────────────────────────────
# Solo usamos rl_colors (reportlab.lib.colors)
_DARK   = rl_color.HexColor("#0f1117")
_PANEL  = rl_color.HexColor("#1a1d27")
_AMBER  = rl_color.HexColor("#f59e0b")
_TEAL   = rl_color.HexColor("#14b8a6")
_ROSE   = rl_color.HexColor("#f43f5e")
_GREY   = rl_color.HexColor("#6b7280")
_WHITE  = rl_color.HexColor("#f9fafb")
_LIGHT  = rl_color.HexColor("#d1d5db")
_BG2    = rl_color.HexColor("#13151f")


def _pdf_styles():
    base = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, **{"fontName": "Helvetica", "textColor": _LIGHT,
                                       "fontSize": 9, "leading": 13, **kw})
    return {
        "title":    S("title",   fontName="Helvetica-Bold", fontSize=18, textColor=_WHITE,
                      spaceAfter=2, alignment=TA_LEFT),
        "subtitle": S("sub",     fontSize=9,  textColor=_GREY,  spaceAfter=14),
        "section":  S("sec",     fontName="Helvetica-Bold", fontSize=7, textColor=_GREY,
                      spaceAfter=4, spaceBefore=14,
                      borderPadding=(0, 0, 4, 0)),
        "normal":   S("norm",    fontSize=9,  textColor=_LIGHT, leading=14),
        "kpi_val":  S("kpiv",    fontName="Helvetica-Bold", fontSize=15, textColor=_AMBER,
                      alignment=TA_CENTER, leading=18),
        "kpi_lbl":  S("kpil",    fontSize=7,  textColor=_GREY,  alignment=TA_CENTER),
        "kpi_sub":  S("kpis",    fontSize=7,  textColor=_GREY,  alignment=TA_CENTER),
        "footer":   S("foot",    fontSize=7,  textColor=_GREY,  alignment=TA_CENTER),
        "th":       S("th",      fontName="Helvetica-Bold", fontSize=8, textColor=_AMBER,
                      alignment=TA_CENTER),
        "td":       S("td",      fontSize=8,  textColor=_LIGHT, alignment=TA_CENTER),
        "td_l":     S("tdl",     fontSize=8,  textColor=_LIGHT, alignment=TA_LEFT),
        "note":     S("note",    fontSize=7.5, textColor=_GREY, leading=11),
    }


def _table_style(header_rows=1):
    return TableStyle([
        ("BACKGROUND",   (0, 0), (-1, header_rows - 1), _PANEL),
        ("BACKGROUND",   (0, header_rows), (-1, -1),    _BG2),
        ("ROWBACKGROUNDS",(0, header_rows), (-1, -1),   [_BG2, _PANEL]),
        ("TEXTCOLOR",    (0, 0), (-1, -1),              _LIGHT),
        ("FONTNAME",     (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1),              8),
        ("ALIGN",        (0, 0), (-1, -1),              "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1),              "MIDDLE"),
        ("ROWHEIGHT",    (0, 0), (-1, -1),              16),
        ("GRID",         (0, 0), (-1, -1),              0.4, rl_color.HexColor("#2a2d3a")),
        ("TOPPADDING",   (0, 0), (-1, -1),              4),
        ("BOTTOMPADDING",(0, 0), (-1, -1),              4),
        ("LEFTPADDING",  (0, 0), (-1, -1),              6),
        ("RIGHTPADDING", (0, 0), (-1, -1),              6),
    ])


def _kpi_table(items, styles, col_w=None):
    """items = [(label, value, sub), ...]  — genera fila de KPI cards."""
    n = len(items)
    W = letter[0] - 3.6 * cm
    cw = col_w or [W / n] * n
    rows = [
        [Paragraph(v,   styles["kpi_val"]) for _, v, _ in items],
        [Paragraph(l,   styles["kpi_lbl"]) for l, _, _ in items],
        [Paragraph(s,   styles["kpi_sub"]) for _, _, s in items],
    ]
    ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _PANEL),
        ("GRID",          (0, 0), (-1, -1), 0.4, rl_color.HexColor("#2a2d3a")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])
    return Table(rows, colWidths=cw, style=ts)


def _hr():
    return HRFlowable(width="100%", thickness=0.5,
                      color=rl_color.HexColor("#2a2d3a"), spaceAfter=6, spaceBefore=2)


def _on_page(canvas, doc):
    """Footer en cada página."""
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(_GREY)
    w, _ = letter
    canvas.drawCentredString(w / 2, 1.0 * cm,
        f"Sizing Tool · Documento generado automáticamente · Página {doc.page}")
    canvas.restoreState()


def build_pdf_sizing(
    proj_loc, lat, lon,
    panel_wp, panel_eff_declared, panel_largo_mm, panel_ancho_mm, panel_peso_kg, panel_area,
    n_panels, kwp, pr_pct, irr_vals, monthly_gen, annual_gen,
    p50, p90, co2_saved,
    inversion_usd, usd_to_mxn, ahorro1, payback,
    vpn, tir, lcoe, pb_disc,
    tarifa_efectiva, inflation, discount_rate, vida_util, om_pct,
    sizing_mode_label="—",
) -> bytes:
    """Genera el PDF de Pre-Sizing / TOR y devuelve bytes."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.6*cm, bottomMargin=1.8*cm,
    )
    S = _pdf_styles()
    story = []

    # ── Encabezado ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Sizing Tool", S["title"]))
    story.append(Paragraph(
        f"Pre-Sizing Fotovoltaico &nbsp;·&nbsp; {proj_loc} &nbsp;·&nbsp; ({lat:.4f}, {lon:.4f})",
        S["subtitle"]))
    story.append(_hr())

    # ── KPIs principales ────────────────────────────────────────────────────────
    story.append(Paragraph("RESULTADOS DEL SISTEMA", S["section"]))
    story.append(_kpi_table([
        ("Capacidad pico",    f"{kwp:.2f} kWp",           f"{n_panels} paneles × {panel_wp} Wp"),
        ("Generación P50",    f"{p50/1000:.1f} MWh/año",  "Mediana histórica"),
        ("Generación P90",    f"{p90/1000:.1f} MWh/año" if p90 else "—", "Percentil 10"),
        ("CO&#8322; evitado", f"{co2_saved/1000:,.2f} t/a",f"Factor {CO2_FACTOR_KG_KWH} kg CO₂e/kWh · SEN 2024"),
    ], S))
    story.append(Spacer(1, 6))

    # ── Panel técnico ──────────────────────────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("PARÁMETROS DEL PANEL", S["section"]))
    panel_rows = [
        [Paragraph("Parámetro", S["th"]), Paragraph("Valor", S["th"])],
        [Paragraph("Potencia pico (Pmax)", S["td_l"]),  Paragraph(f"{panel_wp} Wp", S["td"])],
        [Paragraph("Eficiencia declarada", S["td_l"]),  Paragraph(f"{panel_eff_declared:.1f}%", S["td"])],
        [Paragraph("Dimensiones", S["td_l"]),           Paragraph(f"{panel_largo_mm} × {panel_ancho_mm} mm", S["td"])],
        [Paragraph("Area unitaria", S["td_l"]),         Paragraph(f"{panel_area:.4f} m²", S["td"])],
        [Paragraph("Peso unitario panel", S["td_l"]),   Paragraph(f"{panel_peso_kg} kg", S["td"])],
        [Paragraph("Carga total (x1.35)", S["td_l"]),
         Paragraph(f"{n_panels * panel_peso_kg * 1.35:,.0f} kg  |  "
                   f"{n_panels * panel_peso_kg * 1.35 * 9.8 / 1000:,.2f} kN", S["td"])],
        [Paragraph("Carga por m² instalado", S["td_l"]),
         Paragraph(f"{n_panels * panel_peso_kg * 1.35 * 9.8 / 1000 / max(n_panels * panel_area, 0.01):.3f} kN/m²  "
                   f"[factor x1.35 montura+BOS]", S["td"])],
        [Paragraph("Performance Ratio (PR)", S["td_l"]),Paragraph(f"{pr_pct:.1f}%", S["td"])],
        [Paragraph("Modo de dimensionamiento", S["td_l"]), Paragraph(sizing_mode_label, S["td"])],
    ]
    W = letter[0] - 3.6*cm
    story.append(Table(panel_rows, colWidths=[W*0.6, W*0.4], style=_table_style()))
    story.append(Spacer(1, 8))

    # ── Irradiancia + Generación mensual ───────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("IRRADIANCIA Y GENERACIÓN MENSUAL", S["section"]))
    hdr = [Paragraph(m, S["th"]) for m in MONTHS]
    irr_row = [Paragraph(f"{v:.2f}", S["td"]) for v in irr_vals]
    gen_row = [Paragraph(f"{v/1000:.1f}", S["td"]) for v in monthly_gen]
    lbl_col_w = 2.2*cm
    data_w = (W - lbl_col_w) / 12
    irr_table = Table(
        [[Paragraph("", S["th"])] + hdr,
         [Paragraph("kWh/m²/día", S["td_l"])] + irr_row,
         [Paragraph("Gen. MWh/mes", S["td_l"])] + gen_row],
        colWidths=[lbl_col_w] + [data_w]*12,
        style=_table_style(),
    )
    story.append(irr_table)
    story.append(Paragraph(
        f"Total generación anual P50: {annual_gen:,.0f} kWh/año  ·  Fuente: NASA POWER {NASA_START}–{NASA_END}",
        S["note"]))
    story.append(Spacer(1, 8))

    # ── Parámetros financieros ─────────────────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("PARÁMETROS FINANCIEROS", S["section"]))
    fin_rows = [
        [Paragraph("Parámetro", S["th"]), Paragraph("Valor", S["th"]),
         Paragraph("Parámetro", S["th"]), Paragraph("Valor", S["th"])],
        [Paragraph("Tarifa efectiva", S["td_l"]),  Paragraph(f"${tarifa_efectiva:.3f}/kWh", S["td"]),
         Paragraph("Inflación tarifa", S["td_l"]), Paragraph(f"{inflation:.1f}%/año", S["td"])],
        [Paragraph("Tasa de descuento", S["td_l"]),Paragraph(f"{discount_rate:.1f}%", S["td"]),
         Paragraph("O&M anual", S["td_l"]),        Paragraph(f"{om_pct:.1f}% inv.", S["td"])],
        [Paragraph("Vida útil", S["td_l"]),        Paragraph(f"{vida_util} años", S["td"]),
         Paragraph("TIR", S["td_l"]),              Paragraph(f"{tir:.1f}%" if tir else "—", S["td"])],
    ]
    story.append(Table(fin_rows, colWidths=[W*0.3, W*0.2, W*0.3, W*0.2], style=_table_style()))
    story.append(Spacer(1, 8))

    # ── Especificaciones técnicas y certificaciones ────────────────────────────
    story.append(_hr())
    story.append(Paragraph("ESPECIFICACIONES TÉCNICAS Y CERTIFICACIONES REQUERIDAS", S["section"]))

    specs = [
        ("MÓDULOS FV",         "IEC 61215 (calificación de diseño) · IEC 61730 (seguridad) · UL 61730 si aplica"),
        ("Efic. / Degradación","Eficiencia ≥ 21% · Degradación: ≤ 2% año 1; ≤ 0.4%/año (TOPCon) / ≤ 0.5%/año (PERC)"),
        ("Clase / IP",         "Clase A · ≥ 1000 V DC · IP67 en caja de conexiones"),
        ("Carga mecánica",     "Frontal ≥ 5400 Pa (IEC 61215)"),
        ("Garantías módulo",   "Producto ≥ 12 años · Potencia lineal ≥ 25 años (P90 ≥ 80% Pmax)"),
        ("Amoniaco / Salina",  "IEC 62716 (zonas agrícolas) · IEC 61701 Cat.6 (zonas costeras) — condicional"),
        ("INVERSORES",         "IEC 62109-1/-2 · Anti-islanding IEEE 1547 / NOM-001-SEDE · THD ≤ 3% (IEEE 519)"),
        ("IP / Protecciones",  "IP66 mínimo · AFCI NEC 690.11 · Rapid Shutdown NEC 690.12 (techos)"),
        ("Eficiencia inversor","≥ 98% Euro o CEC a potencia nominal"),
        ("Garantía inversor",  "Mínimo 5 años · extendible a 10–20 años"),
        (">500 kWp topología", "Justificar string / centralizado / optimizadores según sombra y orientación"),
        ("ESTRUCTURA",         "Aluminio 6005-T5 o acero ASTM A653 ≥ ZF275 · Niebla salina ≥ 1000 h (ISO 9227)"),
        ("Diseño estructural", "ASCE 7 / MDOC-CFE (viento, sismo) · Firmado por DRO o Perito acreditado"),
        ("CABLEADO DC",        "H1Z2Z2-K o equiv. IEC 62930 · Tensión ≥ 1500 V"),
        ("PROTECCIONES AC",    "NOM-001-SEDE / NEC 690 · Puesta a tierra NOM-022-STPS y NEC 690.47"),
        ("MONITOREO",          "Datalogger SCADA/cloud · Resolución ≤ 15 min · Modbus TCP o SunSpec · SLA ≥ 99%"),
        ("Variables monitor.", "Potencia AC/DC, energía acumulada, V/I por MPPT, temp. módulo, irradiancia (piranómetro clase 2), alarmas"),
    ]
    from reportlab.platypus import Table as RLTable
    W_page = letter[0] - 3.6*cm
    spec_data = [[Paragraph("Ítem", S["th"]), Paragraph("Requisito", S["th"])]]
    for item, req in specs:
        spec_data.append([Paragraph(item, S["td_l"]), Paragraph(req, S["td_l"])])
    spec_tbl = RLTable(spec_data, colWidths=[W_page*0.28, W_page*0.72])
    spec_tbl.setStyle(_table_style(header_rows=1))
    story.append(spec_tbl)
    story.append(Spacer(1, 8))

    # ── Verificaciones e inspecciones ──────────────────────────────────────────
    story.append(Paragraph("VERIFICACIONES E INSPECCIONES REQUERIDAS", S["section"]))
    insp_data = [
        [Paragraph("Fase", S["th"]), Paragraph("Verificación / Inspección", S["th"]), Paragraph("Norma de referencia", S["th"])],
        [Paragraph("Pre-puesta en marcha", S["td_l"]),
         Paragraph("Polaridad, continuidad y aislamiento DC ≥ 1 MΩ a 500 V", S["td_l"]),
         Paragraph("IEC 62446-1 · IEC 61557-2", S["td"])],
        [Paragraph("Pre-puesta en marcha", S["td_l"]),
         Paragraph("Inspección visual montaje, torque conectores MC4, sellado", S["td_l"]),
         Paragraph("IEC 62446-1", S["td"])],
        [Paragraph("Pre-puesta en marcha", S["td_l"]),
         Paragraph("Continuidad puesta a tierra (< 1 Ω) y protecciones AC", S["td_l"]),
         Paragraph("NOM-001-SEDE · NEC 690.47", S["td"])],
        [Paragraph("Commissioning", S["td_l"]),
         Paragraph("Curvas I-V por string vs STC (tolerancia ± 3%)", S["td_l"]),
         Paragraph("IEC 62446-1", S["td"])],
        [Paragraph("Commissioning", S["td_l"]),
         Paragraph("Termografía IR de módulos y conexiones bajo carga", S["td_l"]),
         Paragraph("IEC TS 62446-3", S["td"])],
        [Paragraph("Commissioning", S["td_l"]),
         Paragraph("Anti-islanding y protecciones de reconexión", S["td_l"]),
         Paragraph("IEEE 1547 · NOM-001-SEDE", S["td"])],
        [Paragraph("Commissioning", S["td_l"]),
         Paragraph("Validación monitoreo con registro mínimo 48 h", S["td_l"]),
         Paragraph("IEC 62446-2", S["td"])],
        [Paragraph("Anual", S["td_l"]),
         Paragraph("Termografía IR módulos, inversores y cableado", S["td_l"]),
         Paragraph("IEC TS 62446-3", S["td"])],
        [Paragraph("Anual", S["td_l"]),
         Paragraph("Prueba de aislamiento DC y revisión de estructura/anclajes", S["td_l"]),
         Paragraph("IEC 62446-1", S["td"])],
        [Paragraph("Anual", S["td_l"]),
         Paragraph("Análisis de rendimiento vs P90 (PR objetivo ≥ 75%)", S["td_l"]),
         Paragraph("IEC 61724-1", S["td"])],
        [Paragraph("Semestral", S["td_l"]),
         Paragraph("Limpieza de módulos y revisión de firmware inversores", S["td_l"]),
         Paragraph("Manual fabricante", S["td"])],
    ]
    W_page2 = letter[0] - 3.6*cm
    insp_tbl = RLTable(insp_data, colWidths=[W_page2*0.22, W_page2*0.52, W_page2*0.26])
    insp_tbl.setStyle(_table_style(header_rows=1))
    story.append(insp_tbl)
    story.append(Spacer(1, 8))

    story.append(Paragraph("ENTREGABLES OBLIGATORIOS DEL PROVEEDOR", S["section"]))
    entregables = [
        "Simulación PVSyst / Helioscope (P50 y P90) con sombras 3D, estudio de soiling y desglose de pérdidas",
        "Planos eléctricos y estructurales firmados por DRO o Perito acreditado",
        "Memoria de cálculo estructural del sistema de montaje",
        "Dictamen UVIE (NOM-001-SEDE) — requerido para contrato CFE en mayoría de estados",
        "Inspección municipal / STPS si aplica (uso industrial o comercial mayor)",
        "Trámite de interconexión CFE (Pequeña Escala o GD) · Permiso CRE si > 0.5 MW",
        "Garantía de ejecución EPC ≥ 2 años post-puesta en marcha (financiamientos exigen 3 años)",
        "Seguro de Responsabilidad Civil vigente durante la obra",
        "Performance bond / fianza de cumplimiento ≥ 10% del contrato (obligatorio > 100 kWp)",
        "Manual de O&M con plan de inspecciones periódicas",
        "Protocolo de commissioning firmado: curvas I-V, termografía IR, aislamiento DC, PR",
    ]
    for e in entregables:
        story.append(Paragraph(f"• {e}", S["note"]))
        story.append(Spacer(1, 2))
    story.append(Spacer(1, 8))

    # ── Consideraciones ────────────────────────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("CONSIDERACIONES GENERALES", S["section"]))
    notas = [
        "Los valores son estimados de pre-sizing (±15%). El proveedor deberá realizar diseño detallado con software especializado (PVSyst, Helioscope, etc.).",
        f"El P90 representa el percentil 10 de la distribución de generaciones anuales simuladas con {NASA_END - NASA_START + 1} años de irradiancia real NASA POWER ({NASA_START}–{NASA_END}).",
        "Verificar disponibilidad de red y trámites CFE / CRE antes de proceder.",
        f"Factor de emisión CO₂ de referencia: {CO2_FACTOR_KG_KWH} kg CO₂e/kWh — Factor de Emisión del SEN 2024 (SEMARNAT/CRE, aviso 28-Feb-2025). Actualizar con el aviso anual más reciente.",
    ]
    for n in notas:
        story.append(Paragraph(f"• {n}", S["note"]))
        story.append(Spacer(1, 3))

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return buf.getvalue()


def build_pdf_ppa(
    proj_loc, ppa_kwp, ppa_gen_anual, ppa_inversion_usd, usd_to_mxn,
    ppa_wacc, ppa_inflacion_tarifa, ppa_degradacion, ppa_om_pct,
    ppa_seguros_pct, ppa_tarifa_cliente, ppa_inflacion_cfe,
    ppa_precio_manual, ppa_plazo_minimo, ppa_plazos,
    resultados, descuento_vs_cfe,
    ro,                    # resultados[ppa_plazo_minimo]
    ahorro_total, ahorro_y, cfe_y, pago_ppa, pago_cfe,
    pm_obj, viable,
) -> bytes:
    """Genera el PDF del análisis PPA y devuelve bytes."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.6*cm, bottomMargin=1.8*cm,
    )
    S = _pdf_styles()
    story = []
    W = letter[0] - 3.6*cm

    # ── Encabezado ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Sizing Tool", S["title"]))
    story.append(Paragraph(
        f"Análisis PPA · Venta al Cliente &nbsp;·&nbsp; {proj_loc} &nbsp;·&nbsp; "
        f"Plazo objetivo: {ppa_plazo_minimo} años",
        S["subtitle"]))
    story.append(_hr())

    # ── KPIs hero ──────────────────────────────────────────────────────────────
    story.append(Paragraph("RESUMEN EJECUTIVO", S["section"]))
    pm_str  = f"${pm_obj:.4f}/kWh" if pm_obj else "No viable"
    viable_color = "#4ade80" if viable else "#f43f5e"
    story.append(_kpi_table([
        ("Precio PPA año 1",     f"${ppa_precio_manual:.4f}/kWh", "Evaluado"),
        ("Precio minimo viable", pm_str,                           f"VPN=0 a {ppa_plazo_minimo}a"),
        ("Descuento vs CFE hoy", f"{descuento_vs_cfe:+.1f}%",     "precio PPA vs tarifa"),
        ("Ahorro total cliente", f"${ahorro_total:,.0f}",          f"MXN en {ppa_plazo_minimo} años"),
    ], S))
    story.append(Spacer(1, 6))
    story.append(_kpi_table([
        ("VPN desarrollador",  f"${ro['vpn']:,.0f} MXN",         f"WACC {ppa_wacc:.1f}%"),
        ("TIR equity",         f"{ro['tir']:.1f}%" if ro["tir"] else "—", "sobre capital propio"),
        ("Payback simple",     f"{ro['pb']} años" if ro["pb"] else f">{ppa_plazo_minimo}a", "nominal s/descontar"),
        ("Payback descontado",  f"{ro['pb_disc']} años" if ro.get("pb_disc") else f">{ppa_plazo_minimo}a", f"WACC {ppa_wacc:.1f}%"),
        ("Valor residual",     f"${ro.get('valor_residual',0):,.0f} MXN",
                               f"{max(0,25-ppa_plazo_minimo)}a restantes"),
    ], S))
    story.append(Spacer(1, 8))

    # ── Comparativo de plazos ─────────────────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("COMPARATIVO DE PLAZOS", S["section"]))
    plazos_hdr = [Paragraph("Métrica", S["th"])] + [Paragraph(f"{pl}a", S["th"]) for pl in ppa_plazos]
    metricas = [
        ("VPN (MXN)",       lambda r: f"${r['vpn']:,.0f}"),
        ("TIR equity",      lambda r: f"{r['tir']:.1f}%" if r["tir"] else "N/A"),
        ("Payback simple",  lambda r: f"{r['pb']}" if r["pb"] else f">{ppa_plazo_minimo}a"),
        ("Payback desc.",    lambda r: f"{r['pb_disc']}" if r.get("pb_disc") else f">{ppa_plazo_minimo}a"),
        ("Ingreso total",   lambda r: f"${r['ing_total']:,.0f}"),
        ("Precio minimo",   lambda r: f"${r['pm']:.4f}" if r.get("pm") else "N/V"),
        ("Valor residual",  lambda r: f"${r.get('valor_residual',0):,.0f}"),
    ]
    plazos_rows = [plazos_hdr]
    for label, fn in metricas:
        row = [Paragraph(label, S["td_l"])] + [Paragraph(fn(resultados[pl]), S["td"]) for pl in ppa_plazos]
        plazos_rows.append(row)
    col_w_pl = [W * 0.28] + [W * 0.72 / len(ppa_plazos)] * len(ppa_plazos)
    story.append(Table(plazos_rows, colWidths=col_w_pl, style=_table_style()))
    story.append(Spacer(1, 8))

    # ── Parámetros del modelo ──────────────────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("PARÁMETROS DEL MODELO", S["section"]))
    param_rows = [
        [Paragraph("Parámetro", S["th"]), Paragraph("Valor", S["th"]),
         Paragraph("Parámetro", S["th"]), Paragraph("Valor", S["th"])],
        [Paragraph("Capacidad sistema", S["td_l"]),    Paragraph(f"{ppa_kwp:.1f} kWp", S["td"]),
         Paragraph("Generación año 1", S["td_l"]),     Paragraph(f"{ppa_gen_anual:,.0f} kWh", S["td"])],
        [Paragraph("Inversión total", S["td_l"]),      Paragraph(f"${ppa_inversion_usd:,.0f} USD", S["td"]),
         Paragraph("Tipo de cambio", S["td_l"]),       Paragraph(f"${usd_to_mxn:.2f} MXN/USD", S["td"])],
        [Paragraph("WACC", S["td_l"]),                 Paragraph(f"{ppa_wacc:.1f}%", S["td"]),
         Paragraph("Escalador PPA", S["td_l"]),        Paragraph(f"{ppa_inflacion_tarifa:.1f}%/año", S["td"])],
        [Paragraph("Degradacion paneles", S["td_l"]), Paragraph(f"{ppa_degradacion:.2f}%/año", S["td"]),
         Paragraph("O&M anual", S["td_l"]),            Paragraph(f"{ppa_om_pct:.1f}% inv.", S["td"])],
        [Paragraph("Seguros", S["td_l"]),              Paragraph(f"{ppa_seguros_pct:.2f}% inv.", S["td"]),
         Paragraph("Tarifa CFE cliente", S["td_l"]),  Paragraph(f"${ppa_tarifa_cliente:.4f}/kWh", S["td"])],
        [Paragraph("Inflacion CFE", S["td_l"]),        Paragraph(f"{ppa_inflacion_cfe:.1f}%/año", S["td"]),
         Paragraph("Inflacion O&M", S["td_l"]),        Paragraph("—", S["td"])],
    ]
    story.append(Table(param_rows, colWidths=[W*0.3, W*0.2, W*0.3, W*0.2], style=_table_style()))
    story.append(Spacer(1, 8))

    # ── Tabla flujos anuales (plazo objetivo) ──────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph(
        f"FLUJOS ANUALES — PLAZO {ppa_plazo_minimo} AÑOS", S["section"]))
    flujo_hdr = [
        Paragraph("Año", S["th"]),
        Paragraph("Gen. (MWh)", S["th"]),
        Paragraph("Precio PPA", S["th"]),
        Paragraph("Ingreso PPA", S["th"]),
        Paragraph("CFE equiv.", S["th"]),
        Paragraph("Ahorro cliente", S["th"]),
        Paragraph("Flujo neto", S["th"]),
    ]
    flujo_rows = [flujo_hdr]
    gen_cl  = ro["gen_y"]
    prec_cl = ro["prec_y"]
    for i in range(ppa_plazo_minimo):
        fn_v = ro["fn_y"][i]
        fn_color = "#4ade80" if fn_v >= 0 else "#f87171"
        flujo_rows.append([
            Paragraph(str(ro["years"][i]), S["td"]),
            Paragraph(f"{gen_cl[i]/1000:.2f}", S["td"]),
            Paragraph(f"${prec_cl[i]:.4f}", S["td"]),
            Paragraph(f"${pago_ppa[i]:,.0f}", S["td"]),
            Paragraph(f"${cfe_y[i]:.4f}", S["td"]),
            Paragraph(f"${ahorro_y[i]:,.0f}", S["td"]),
            Paragraph(f'<font color="{fn_color}">${fn_v:,.0f}</font>', S["td"]),
        ])
    col_w_fl = [W*0.07, W*0.12, W*0.12, W*0.16, W*0.12, W*0.20, W*0.21]
    story.append(Table(flujo_rows, colWidths=col_w_fl, style=_table_style()))
    story.append(Spacer(1, 8))

    # ── Nota final ─────────────────────────────────────────────────────────────
    story.append(_hr())
    story.append(Paragraph("NOTAS Y DESCARGOS", S["section"]))
    notas_ppa = [
        "Este documento es una estimación financiera de pre-venta. Los valores reales dependerán del desempeño del sistema, condiciones climáticas y acuerdos contractuales definitivos.",
        "El valor residual se calcula usando anuidad con crecimiento compuesto (fórmula de Gordon) incorporando la degradación anual del panel y el escalador PPA.",
        "El precio mínimo viable (VPN=0) se calcula por bisección con 80 iteraciones sobre la función de VPN.",
        "Se recomienda contratar una auditoría energética y simulación detallada (PVSyst/Helioscope) antes de firmar el contrato PPA.",
    ]
    for n in notas_ppa:
        story.append(Paragraph(f"• {n}", S["note"]))
        story.append(Spacer(1, 3))

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return buf.getvalue()


# ── Session state initialization ──────────────────────────────────────────────
if "nasa_irradiance" not in st.session_state:
    st.session_state.nasa_irradiance = DEFAULT_IRR.copy()

if "nasa_irr_por_anio" not in st.session_state:
    st.session_state.nasa_irr_por_anio = {}   # vacío → P90 no disponible sin NASA

if "nasa_source_label" not in st.session_state:
    st.session_state.nasa_source_label = None



## ═════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    import os as _os, base64 as _b64
    _logo_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "logo.png")
    if _os.path.exists(_logo_path):
        with open(_logo_path, "rb") as _lf:
            _logo_b64 = _b64.b64encode(_lf.read()).decode()
        st.markdown(f'<img src="data:image/png;base64,{_logo_b64}" style="width:100%;max-width:275px;margin-bottom:8px;">', unsafe_allow_html=True)
    else:
        st.markdown('<div style="font-size:20px;font-weight:700;color:#f59e0b;padding:8px 0 12px;">⚡ Sizing Tool</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### PARÁMETROS CLAVE")
    st.markdown("---")

    # ── Ubicación con Mapa (solo coordenadas en sidebar) ─────────────────────
    st.markdown("#### Ubicación")
    st.markdown(f"""
    <style>
        div[key="lat_input"] input,
        div[key="lon_input"] input {{
            background-color: #000000 !important;
            color: #ffffff !important;
        }}
    </style>
""", unsafe_allow_html=True)
    col_lat, col_lon = st.columns(2)
    lat = col_lat.number_input("Latitud", -90.0, 90.0, 19.4326, format="%.4f", key="lat_input")
    lon = col_lon.number_input("Longitud", -180.0, 180.0, -99.1332, format="%.4f", key="lon_input")

    st.caption(f"📍 Coordenadas actuales: **{lat:.4f}**, **{lon:.4f}**")

    # Botón único para NASA POWER
    if st.button("🌍 Obtener irradiancia NASA POWER (2005–2024)", 
                 type="primary", 
                 use_container_width=True, 
                 key="nasa_button"):
        with st.spinner("Consultando datos de NASA POWER..."):
            try:
                irr_media, irr_por_anio = get_nasa_power_irradiance(lat, lon)
                st.session_state.nasa_irradiance   = irr_media
                st.session_state.nasa_irr_por_anio = irr_por_anio
                st.session_state.nasa_source_label = f"({lat:.4f}, {lon:.4f})"
                n_anios = len(irr_por_anio)
                st.success(f"✅ Datos de NASA POWER cargados — {n_anios} años históricos ({NASA_START}–{NASA_END})")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                st.session_state.nasa_irradiance   = DEFAULT_IRR.copy()
                st.session_state.nasa_irr_por_anio = {}
                st.session_state.nasa_source_label = None

    lbl = st.session_state.get("nasa_source_label", None)
    st.caption(f"Fuente: NASA POWER (2005–2024) · {lbl}" if lbl else "Valores por defecto: CDMX")

    st.markdown("---")

    # ── Ficha técnica del panel ───────────────────────────────────────────────
    st.markdown("#### 📋 Panel (datos básicos)")
    panel_wp           = st.number_input("Potencia pico Pmax (Wp)", 100, 900, 650, 5)
    panel_eff_declared = st.number_input("Eficiencia (%)", 10.0, 26.0, 24.1, 0.01)
    panel_largo_mm     = st.number_input("Largo (mm)", 1000, 2500, 2382, 1)
    panel_ancho_mm     = st.number_input("Ancho (mm)", 700, 1300, 1134, 1)
    panel_peso_kg      = st.number_input("Peso (kg)", 5.0, 40.0, 32.7, .1)

    st.markdown("---")

    # ── PR del Sistema ───────────────────────────────────────────────────────
    st.markdown("#### ⚙️ Performance Ratio (PR) del sistema")

    st.markdown("""
    <div style="font-size:13px; color:#9ca3af; margin-bottom:10px;">
        PR global del sistema (incluye inversor, cableado, suciedad, mismatch, temperatura, etc.)
    </div>
    """, unsafe_allow_html=True)

    effective_pr = st.slider(
        "Performance Ratio (PR)",
        min_value=0.60,
        max_value=0.95,
        value=0.78,
        step=0.01,
        format="%.2f"
    )

    pr_pct = effective_pr * 100

    if pr_pct >= 82:
        badge_class = "pr-green"
        badge_text = "● Excelente"
    elif pr_pct >= 75:
        badge_class = "pr-yellow"
        badge_text = "● Bueno"
    else:
        badge_class = "pr-red"
        badge_text = "● Bajo — revisar diseño"

    st.markdown(f"""
    <div class="pr-badge {badge_class}" style="margin:10px 0 8px 0;">
        {badge_text} — PR {pr_pct:.1f}%
    </div>
    """, unsafe_allow_html=True)

    st.caption("Valor típico residencial en México: 0.75 – 0.82")

    # Degradación anual
    st.markdown("---")
    panel_degradation = st.slider("Degradación anual (%/año)", 0.3, 1.0, 0.5, 0.1, key="degradacion_anual")

    st.markdown("---")
    st.markdown("#### 💰 Referencia financiera")
    tarifa    = st.slider("Tarifa ref. área (MXN/kWh)", 1.0, 8.0, 2.80, 0.10, key="tarifa",
                          help="Usada en modo 'Por área'. En modo recibo CFE se usa la tarifa del recibo.")
    inflation = st.slider("Inflación tarifa anual (%)", 0.0, 8.0, 3.0, 0.5, key="inflation")
    discount_rate = st.slider("Tasa de descuento (%)", 0.0, 30.0, 15.0, 0.5, key="discount_rate",
                              help="Tasa usada para evaluación")
    usd_to_mxn    = st.slider("Tipo de Cambio (MXN por USD)", 16.0, 22.0, 17.50, 0.1, key="usd_to_mxn",
                              help="Tipo de cambio para evaluación financiera")
    vida_util = st.slider("Vida útil (años)", 10, 30, 25, 1, key="vida_util")
    costo_kwp = st.slider("Costo ref. instalación (USD/kWp)", 500, 2000, 1000, 50, key="costo_kwp")
    om_pct_sidebar = st.slider("O&M anual (% inversión MXN)", 0.5, 3.0, 1.0, 0.1, key="om_pct_sidebar",
                               help="Operación y mantenimiento anual como % de la inversión en MXN")

    st.markdown("---")
    st.markdown(f"<div style='font-size:11px;color:#4b5563;'>v3.0 · NASA POWER {NASA_START}–{NASA_END} · México</div>",
                unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# Derived panel values (se calculan después del sidebar)
# ═════════════════════════════════════════════════════════════════════════════
panel_largo_m  = panel_largo_mm / 1000
panel_ancho_m  = panel_ancho_mm / 1000
panel_area     = panel_largo_m * panel_ancho_m
panel_eff_calc = (panel_wp / (panel_area * 1000)) * 100
eff_delta      = panel_eff_calc - panel_eff_declared
eff_ok         = abs(eff_delta) <= 0.5
eff_color      = "#14b8a6" if eff_ok else "#f43f5e"
eff_note       = "✓" if eff_ok else f"{'↑' if eff_delta>0 else '↓'}{abs(eff_delta):.1f}% vs declarada"


# ═════════════════════════════════════════════════════════════════════════════
# HEADER
# ═════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="app-title">Sizing Tool</div>', unsafe_allow_html=True)


tab1, tab3, tab4, tab5, tab6 = st.tabs(["  Turnkey Solar", "  PPA Solar", "  Turnkey BESS", "  PPA Bess", "⚡  Sistema Híbrido"])
active_irr         = st.session_state.nasa_irradiance
active_irr_por_anio = st.session_state.nasa_irr_por_anio


def irr_source_banner():
    lbl = st.session_state.nasa_source_label
    if lbl:
        st.markdown(
            f'<div class="nasa-box">🌍 NASA POWER · Climatología {NASA_START}–{NASA_END} · {lbl} · Editable</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            '<div class="info-box">Valores por defecto CDMX. Busca coordenada en el sidebar para modificar el sitio.</div>',
            unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS VISUALIZACIÓN 8760h — usados en Tab 1 (modo CSV) y Tab 6
# ═════════════════════════════════════════════════════════════════════════════

def build_heatmap_fig(values_8760: list, title: str, unit: str, colorscale: str = "YlOrRd") -> go.Figure:
    """Construye heatmap 24h × 365d desde una serie de 8760 valores."""
    n = min(len(values_8760), 8760)
    padded = list(values_8760[:n]) + [0.0] * (8760 - n)
    mat = np.array(padded[:8760]).reshape(365, 24)
    lay = copy.deepcopy(PLOT_LAYOUT)
    fig = go.Figure(go.Heatmap(
        z=mat,
        x=[f"{h:02d}h" for h in range(24)],
        y=[f"Día {d+1}" for d in range(365)],
        colorscale=colorscale,
        colorbar=dict(title=unit, tickfont=dict(size=9)),
        hovertemplate="Día %{y} · %{x}<br>" + f"{title}: %{{z:.1f}} {unit}<extra></extra>",
    ))
    lay.update({"height": 340, "title": dict(text=title, font=dict(size=12)),
                "xaxis": dict(title="Hora del día", tickfont=dict(size=8)),
                "yaxis": dict(title="", showticklabels=False),
                "margin": dict(l=10, r=60, t=40, b=40)})
    fig.update_layout(**lay)
    return fig


def duration_curve_fig(values_8760: list, label: str, color: str, unit: str = "kW") -> go.Figure:
    """Curva de duración (load duration curve) de una serie 8760h."""
    sorted_vals = sorted(values_8760, reverse=True)
    hours = list(range(1, len(sorted_vals) + 1))
    lay = copy.deepcopy(PLOT_LAYOUT)
    fig = go.Figure(go.Scatter(
        x=hours, y=sorted_vals, mode="lines", name=label,
        line=dict(color=color, width=2),
        fill="tozeroy", fillcolor=f"rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},0.10)",
    ))
    lay.update({"height": 280,
                "xaxis": dict(title="Horas/año (mayor a menor)", gridcolor="#1e2230"),
                "yaxis": dict(title=unit, gridcolor="#1e2230"),
                "margin": dict(l=10, r=10, t=30, b=40)})
    fig.update_layout(**lay)
    return fig


# ═════════════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES — 8760h (TMY, PV, BESS, CHP, financiero)
# Definidas aquí para estar disponibles en todos los tabs.
# ═════════════════════════════════════════════════════════════════════════════
def get_nasa_hourly(lat: float, lon: float, year: int = 2023) -> tuple[pd.DataFrame | None, str]:
    """
    Descarga datos horarios NASA POWER para un año dado.
    Parámetros: ALLSKY_SFC_SW_DWN (W/m²), T2M (°C), WS2M (m/s).
    Devuelve (DataFrame 8760 filas, mensaje_error).
    API hourly disponible desde 2001.
    """
    params = "ALLSKY_SFC_SW_DWN,T2M,WS2M"
    url = (
        "https://power.larc.nasa.gov/api/temporal/hourly/point"
        f"?parameters={params}&community=RE"
        f"&longitude={lon}&latitude={lat}&format=JSON"
        f"&start={year}0101&end={year}1231"
    )
    try:
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        data = r.json()
        raw = (data.get("properties", {})
                   .get("parameter", {}))
        if not raw:
            return None, "NASA POWER no devolvió datos horarios."

        irr_raw  = raw.get("ALLSKY_SFC_SW_DWN", {})
        temp_raw = raw.get("T2M", {})
        wind_raw = raw.get("WS2M", {})

        rows = []
        for key in sorted(irr_raw.keys()):
            if len(key) != 10:   # YYYYMMDDHH
                continue
            try:
                dt_str = f"{key[:4]}-{key[4:6]}-{key[6:8]} {key[8:10]}:00"
                irr  = float(irr_raw.get(key, 0))
                temp = float(temp_raw.get(key, 20))
                wind = float(wind_raw.get(key, 0))
                # NASA usa -999 para datos inválidos
                if irr  < 0: irr  = 0.0
                if temp < -50: temp = 20.0
                if wind < 0: wind = 0.0
                rows.append({"datetime": dt_str,
                             "irradiance_Wm2": irr,
                             "temp_C": temp,
                             "wind_ms": wind})
            except (ValueError, TypeError):
                continue

        if not rows:
            return None, "No se pudieron parsear los datos horarios."

        df = pd.DataFrame(rows)
        df["datetime"] = pd.to_datetime(df["datetime"])

        # Asegurar exactamente 8760 horas (año no bisiesto)
        df = df.head(8760).reset_index(drop=True)
        if len(df) < 8000:
            return None, f"Solo se obtuvieron {len(df)} horas. Verifica coordenadas."

        return df, ""

    except requests.exceptions.Timeout:
        return None, "⏱️ Timeout: NASA POWER tardó demasiado. Intenta de nuevo."
    except requests.exceptions.ConnectionError:
        return None, "🌐 Sin conexión. Verifica tu internet."
    except Exception as e:
        return None, f"❌ Error NASA POWER: {str(e)[:120]}"



def get_nasa_monthly_temp(lat: float, lon: float) -> tuple[list, str]:
    """
    Descarga temperatura media mensual (T2M) de NASA POWER 2005-2024.
    Devuelve (lista 12 valores °C, mensaje).
    Reutiliza el mismo endpoint mensual que get_nasa_power_irradiance().
    """
    url = (
        "https://power.larc.nasa.gov/api/temporal/monthly/point"
        "?parameters=T2M&community=RE"
        f"&longitude={lon}&latitude={lat}&format=JSON"
        f"&start={NASA_START}&end={NASA_END}"
    )
    try:
        r = requests.get(url, timeout=25)
        r.raise_for_status()
        raw = (r.json().get("properties", {})
                       .get("parameter", {})
                       .get("T2M", {}))
        if not raw:
            return [20.0]*12, "Sin datos de temperatura — usando 20°C fijo."

        monthly_sum   = [0.0]*12
        monthly_count = [0]*12
        for key, val in raw.items():
            if len(key) != 6: continue
            try:
                m_idx = int(key[4:6]) - 1
                if 0 <= m_idx <= 11 and val is not None and val > -100:
                    monthly_sum[m_idx]   += float(val)
                    monthly_count[m_idx] += 1
            except (ValueError, TypeError):
                continue

        temp_media = [
            round(monthly_sum[i]/monthly_count[i], 2) if monthly_count[i] > 0 else 20.0
            for i in range(12)
        ]
        return temp_media, f"Temperatura media mensual 2005-2024 · Media anual: {sum(temp_media)/12:.1f}°C"
    except Exception as e:
        return [20.0]*12, f"Error temperatura NASA: {str(e)[:80]} — usando 20°C."



def build_tmy_8760(
    lat: float,
    lon: float,
    irr_media_mensual: tuple,   # 12 valores climatológicos 2005-2024 (kWh/m²/día)
    temp_media_mensual: tuple = None,   # 12 valores °C promedio mensual (opcional)
    year_ref: int = 2023,       # año para obtener perfil intradiario de forma
) -> tuple[pd.DataFrame | None, str]:
    """
    Construye un perfil horario TMY (Typical Meteorological Year) de 8760h
    combinando dos fuentes de NASA POWER:

    1. MAGNITUD MENSUAL: promedio climatológico 2005-2024 ya calculado por
       get_nasa_power_irradiance() — 20 años, estadísticamente robusto.

    2. FORMA INTRADIARIA: distribución horaria de un año de referencia obtenida
       de la API horaria NASA POWER — captura la curva solar real del sitio.

    Metodología (equivalente a TMY simplificado NREL/SAM):
      Para cada mes m:
        factor_m = irr_media_mensual[m] * 1000/24 / mean(irr_horaria_año_ref[mes_m])
        irr_tmy[h] = irr_horaria_año_ref[h] * factor_m   (para h ∈ mes_m)

    Esto garantiza que:
      - La irradiancia integrada mensual coincide con el P50 de 20 años
      - La forma de la curva solar (ascenso/descenso, horas pico) es realista
      - La temperatura usa el promedio mensual si se proporciona

    Devuelve (DataFrame 8760h, mensaje_error).
    """
    MONTH_DAYS_TMY = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    # ── 1. Obtener perfil intradiario del año de referencia ─────────────────
    df_ref, err = get_nasa_hourly(lat, lon, year_ref)
    if err or df_ref is None:
        return None, f"No se pudo obtener perfil horario de referencia: {err}"

    irr_ref = df_ref["irradiance_Wm2"].values  # W/m²
    temp_ref = df_ref["temp_C"].values
    wind_ref = df_ref["wind_ms"].values

    # ── 2. Escalar mes a mes para que la magnitud coincida con el P50 20a ──
    tmy_irr  = irr_ref.copy().astype(float)
    tmy_temp = temp_ref.copy().astype(float)

    hour_cursor = 0
    for m in range(12):
        n_hours = MONTH_DAYS_TMY[m] * 24
        h_start = hour_cursor
        h_end   = min(hour_cursor + n_hours, len(irr_ref))
        h_slice = slice(h_start, h_end)

        # Media de la irradiancia horaria del mes de referencia (W/m²)
        irr_slice = irr_ref[h_slice]
        mean_ref_wm2 = irr_slice.mean()  # W/m² media horaria del mes

        # Target: convertir irr_media_mensual[m] (kWh/m²/día) a W/m² media horaria
        # kWh/m²/día = Wh/m²/día / 1000
        # W/m² media 24h = kWh/m²/día * 1000 / 24
        target_wm2 = irr_media_mensual[m] * 1000 / 24

        # Factor de escala
        if mean_ref_wm2 > 1.0:   # evitar división por cero en meses sin sol
            factor = target_wm2 / mean_ref_wm2
        else:
            factor = 1.0

        tmy_irr[h_slice] = np.clip(irr_ref[h_slice] * factor, 0, 1500)

        # Temperatura: si se proporcionan promedios mensuales climatológicos, centrar
        if temp_media_mensual is not None:
            t_offset = temp_media_mensual[m] - temp_ref[h_slice].mean()
            tmy_temp[h_slice] = temp_ref[h_slice] + t_offset

        hour_cursor = h_end

    # ── 3. Construir DataFrame TMY ──────────────────────────────────────────
    # Generar timestamps para un año no bisiesto (2023)
    ts = pd.date_range(f"{year_ref}-01-01 00:00", periods=len(tmy_irr), freq="h")
    n  = min(len(tmy_irr), 8760)

    df_tmy = pd.DataFrame({
        "datetime":        ts[:n],
        "irradiance_Wm2":  tmy_irr[:n].round(2),
        "temp_C":          tmy_temp[:n].round(2),
        "wind_ms":         wind_ref[:n].round(2),
    })

    # Verificación: energía total debe estar cerca del esperado
    irr_total_kwh_m2 = df_tmy["irradiance_Wm2"].sum() / 1000   # kWh/m²/año
    irr_expected     = sum(irr_media_mensual[m] * MONTH_DAYS_TMY[m] for m in range(12))
    error_pct        = abs(irr_total_kwh_m2 - irr_expected) / max(irr_expected, 1) * 100

    meta = (f"TMY 8760h construido · Irrad. total: {irr_total_kwh_m2:.0f} kWh/m²/año "
            f"(esperado {irr_expected:.0f}, error {error_pct:.1f}%) · "
            f"Temp. media: {df_tmy['temp_C'].mean():.1f}°C · "
            f"Fuente climatológica: NASA POWER 2005–2024")

    return df_tmy, meta



def validate_user_csv(df: pd.DataFrame) -> tuple[bool, str, pd.DataFrame]:
    """
    Valida el CSV subido por el usuario.
    Requiere: columna load_kW + (datetime o timestamp) + ≥8760 filas.
    Retorna (ok, mensaje, df_limpio).
    """
    # Normalizar nombres de columnas
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # Columna datetime
    dt_col = None
    for candidate in ["datetime", "timestamp", "date", "hora", "time"]:
        if candidate in df.columns:
            dt_col = candidate
            break
    if dt_col is None:
        return False, "El CSV debe tener una columna 'datetime' o 'timestamp'.", df

    # Columna de carga
    load_col = None
    for candidate in ["load_kw", "load", "consumo_kw", "demand_kw", "kw", "potencia_kw"]:
        if candidate in df.columns:
            load_col = candidate
            break
    if load_col is None:
        return False, "El CSV debe tener una columna 'load_kW' (consumo horario en kW).", df

    if len(df) < 8760:
        return False, f"El CSV tiene {len(df)} filas. Se requieren ≥ 8760 (una por hora del año).", df

    df = df.copy()
    df["datetime"] = pd.to_datetime(df[dt_col], errors="coerce")
    df["load_kW"]  = pd.to_numeric(df[load_col], errors="coerce").fillna(0.0).clip(lower=0)

    # Columnas opcionales
    if "irradiance_wm2" in df.columns:
        df["irradiance_Wm2"] = pd.to_numeric(df["irradiance_wm2"], errors="coerce").fillna(0).clip(lower=0)
    if "temp_c" in df.columns:
        df["temp_C"] = pd.to_numeric(df["temp_c"], errors="coerce").fillna(20)

    df = df.head(8760).reset_index(drop=True)
    return True, "", df



def merge_csv_with_nasa(user_df: pd.DataFrame, nasa_df: pd.DataFrame) -> pd.DataFrame:
    """
    Une el CSV del usuario con los datos NASA.
    Las columnas NASA se usan como fallback si no vienen en el CSV del usuario.
    """
    df = user_df.copy().reset_index(drop=True)
    nasa = nasa_df.copy().reset_index(drop=True)

    if "irradiance_Wm2" not in df.columns:
        df["irradiance_Wm2"] = nasa["irradiance_Wm2"].values[:len(df)]
    if "temp_C" not in df.columns:
        df["temp_C"] = nasa["temp_C"].values[:len(df)]
    if "wind_ms" not in df.columns:
        df["wind_ms"] = nasa["wind_ms"].values[:len(df)]

    return df



def simulate_pv_8760(
    irr_8760: tuple,      # W/m²  por hora
    temp_8760: tuple,     # °C por hora
    kwp: float,
    pr_base: float,       # PR base (0-1)
    panel_temp_coeff: float = -0.004,  # %/°C pérdida por temperatura (típico -0.4%/°C)
    t_ref: float = 25.0,              # Temperatura STC
    t_noct: float = 45.0,             # NOCT del panel
) -> tuple:
    """
    Simula generación PV hora a hora con corrección de temperatura.
    Fórmula estándar IEC 61724:
      T_cell = T_amb + (NOCT-20)/800 * G
      PR_efectivo = PR_base * (1 + coeff*(T_cell - T_ref))
      P_hora = kwp * (G/1000) * PR_efectivo
    PR clampado a [0, PR_base]: 0 es el límite físico correcto (potencia nunca negativa);
    PR_base es el techo ya que el coeff térmico solo penaliza (coeff < 0).
    Devuelve tuple de 8760 valores en kW.
    """
    gen = []
    for i in range(len(irr_8760)):
        g    = max(0.0, irr_8760[i])          # W/m²
        t    = temp_8760[i]
        t_cell = t + (t_noct - 20) / 800 * g   # temperatura de celda
        pr_eff = pr_base * (1 + panel_temp_coeff * (t_cell - t_ref))
        # Límite inferior 0.0 (física real): el panel nunca genera energía negativa.
        # El límite anterior de 0.5 era arbitrario y subestimaba pérdidas en climas extremos.
        pr_eff = max(0.0, min(pr_eff, pr_base))
        p_kw   = kwp * (g / 1000) * pr_eff
        gen.append(round(p_kw, 4))
    return tuple(gen)



def dispatch_pv_bess_8760(
    load_8760: tuple,      # kW demanda
    pv_8760: tuple,        # kW generación PV
    bess_kwh: float,       # capacidad útil BESS
    bess_kw: float,        # potencia PCS
    rte: float = 0.92,     # eficiencia round-trip
    dod: float = 0.90,
    tariff_8760: tuple = None,   # $/kWh por hora (para arbitraje tarifario)
) -> dict:
    """
    Despacho horario PV+BESS 8760h con estrategia solar-first.
    Prioridades:
      1. PV cubre carga directamente (autoconsumo directo)
      2. Excedente PV carga batería
      3. Batería descarga cuando PV < carga
      4. Red cubre déficit restante
    Devuelve métricas horarias y anuales completas.
    """
    n = len(load_8760)
    soc      = bess_kwh * 0.5      # estado inicial 50%
    soc_min  = 0.0
    soc_max  = bess_kwh

    grid_import   = []   # kWh desde la red
    grid_export   = []   # kWh excedente a la red (curtailment si no hay net metering)
    batt_charge   = []   # kWh cargados a batería
    batt_discharge= []   # kWh descargados de batería (útil, ya con RTE)
    pv_direct     = []   # kWh PV directo a carga
    soc_curve     = []
    total_cycles  = 0.0

    for h in range(n):
        load = max(0.0, load_8760[h])
        pv   = max(0.0, pv_8760[h])

        # 1. PV directo a carga
        pv_dir = min(pv, load)
        net_load = load - pv_dir      # demanda residual tras PV directo
        pv_exc   = pv - pv_dir        # excedente PV tras cubrir carga

        batt_ch  = 0.0
        batt_dis = 0.0

        # 2. Excedente PV → cargar batería
        if pv_exc > 0 and soc < soc_max:
            can_charge = min(bess_kw, soc_max - soc, pv_exc)
            soc += can_charge
            batt_ch = can_charge
            pv_exc -= can_charge

        # 3. Déficit → descargar batería
        if net_load > 0 and soc > soc_min:
            can_discharge = min(bess_kw, (soc - soc_min), net_load / rte)
            discharged_useful = can_discharge * rte
            soc -= can_discharge
            batt_dis  = discharged_useful
            net_load -= discharged_useful
            total_cycles += can_discharge / bess_kwh if bess_kwh > 0 else 0

        # 4. Red cubre déficit restante
        grid_i = max(0.0, net_load)
        grid_e = max(0.0, pv_exc)   # excedente no almacenado → export o curtailment

        grid_import.append(grid_i)
        grid_export.append(grid_e)
        batt_charge.append(batt_ch)
        batt_discharge.append(batt_dis)
        pv_direct.append(pv_dir)
        soc_curve.append(round(soc, 3))

    pv_total    = sum(pv_8760)
    load_total  = sum(load_8760)
    grid_i_tot  = sum(grid_import)
    grid_e_tot  = sum(grid_export)
    pv_dir_tot  = sum(pv_direct)
    batt_dis_tot= sum(batt_discharge)
    batt_ch_tot = sum(batt_charge)

    # Verificación de balance energético (tolerancia numérica de punto flotante):
    # Demanda = PV directo + descarga BESS + red importada
    # PV generado = PV directo + carga BESS + exportado a red
    balance_carga = abs((pv_dir_tot + batt_dis_tot + grid_i_tot) - load_total)
    balance_pv    = abs((pv_dir_tot + batt_ch_tot  + grid_e_tot) - pv_total)
    if balance_carga > max(1.0, load_total * 1e-4):
        import warnings
        warnings.warn(f"Despacho: balance de carga fuera de tolerancia: {balance_carga:.2f} kWh")
    if balance_pv > max(1.0, pv_total * 1e-4):
        import warnings
        warnings.warn(f"Despacho: balance PV fuera de tolerancia: {balance_pv:.2f} kWh")

    autoconsumo  = (pv_dir_tot + batt_dis_tot) / max(pv_total, 1) * 100
    autosufic    = (pv_dir_tot + batt_dis_tot) / max(load_total, 1) * 100
    curtailment  = grid_e_tot / max(pv_total, 1) * 100

    return dict(
        grid_import=grid_import,
        grid_export=grid_export,
        batt_charge=batt_charge,
        batt_discharge=batt_discharge,
        pv_direct=pv_direct,
        soc_curve=soc_curve,
        # Anuales
        pv_total_kwh=round(pv_total, 1),
        load_total_kwh=round(load_total, 1),
        grid_import_kwh=round(grid_i_tot, 1),
        grid_export_kwh=round(grid_e_tot, 1),
        autoconsumo_pct=round(autoconsumo, 2),
        autosuficiencia_pct=round(autosufic, 2),
        curtailment_pct=round(curtailment, 2),
        total_cycles=round(total_cycles, 1),
    )



def simulate_chp_8760(
    load_8760: tuple,       # kW demanda eléctrica
    kwp_chp: float,         # kW eléctricos nominales CHP
    eff_elec: float = 0.35, # eficiencia eléctrica CHP
    eff_term: float = 0.45, # eficiencia térmica CHP
    op_hours: int = 6000,   # horas de operación anuales
    min_load_pct: float = 0.50,  # carga mínima CHP (50% nominal)
) -> dict:
    """
    Simula un CHP (cogeneración gas/biogás) que opera en baseload o load-following.
    Estrategia: opera cuando la demanda supera min_load_pct × kwp_chp.
    Devuelve generación eléctrica y térmica horaria.

    Balance energético (correcto):
      combustible [kWh] = p_eléctrica / eff_elec
      calor_recuperado  = combustible × eff_term
      eficiencia global = eff_elec + eff_term  (típico 80–85% LHV para gas natural)
    """
    n = min(len(load_8760), 8760)
    min_load = kwp_chp * min_load_pct

    # Ordenar horas por demanda desc para asignar op_hours a las más altas
    load_sorted_idx = sorted(range(n), key=lambda i: load_8760[i], reverse=True)
    op_set = set(load_sorted_idx[:op_hours])

    chp_elec = []
    chp_term = []
    fuel_kwh = []

    for h in range(n):
        if h in op_set and load_8760[h] >= min_load:
            # CHP opera al mínimo entre demanda y capacidad nominal
            p_e  = min(kwp_chp, load_8760[h])
            # Balance energético correcto:
            #   combustible = p_e / eff_elec
            #   calor recuperado = combustible * eff_term
            # (NO p_e * eff_term/eff_elec, que viola balance cuando eff_e + eff_t > 1
            #  al no respetar que ambas eficiencias se aplican sobre el mismo input)
            fuel = p_e / eff_elec
            p_t  = fuel * eff_term
        else:
            p_e = 0.0; p_t = 0.0; fuel = 0.0

        chp_elec.append(p_e)
        chp_term.append(p_t)
        fuel_kwh.append(fuel)

    return dict(
        chp_elec=chp_elec,
        chp_term=chp_term,
        fuel_kwh=fuel_kwh,
        chp_elec_total=round(sum(chp_elec), 1),
        chp_term_total=round(sum(chp_term), 1),
        fuel_total=round(sum(fuel_kwh), 1),
        real_op_hours=sum(1 for x in chp_elec if x > 0),
    )



def financial_8760(
    grid_import_kwh: float,
    tarifa_mxn_kwh: float,
    inflation_pct: float,
    capex_usd: float,
    usd_to_mxn: float,
    discount_rate_pct: float,
    vida_util: int,
    om_pct: float,
    panel_degradation_pct: float,
    baseline_cost_mxn: float,      # gasto CFE sin proyecto (año 1)
    # BESS reemplazo (opcional)
    bess_capex_energia_usd: float = 0.0,
    bess_año_reemplazo: int = 999,
    bess_reemplazo_pct: float = 0.65,
) -> dict:
    """
    Modelo financiero adaptado para proyectos 8760h.
    Ahorro = baseline_cost - costo_red_con_proyecto (con inflación).
    """
    r       = discount_rate_pct / 100
    inf     = inflation_pct / 100
    inv_mxn = capex_usd * usd_to_mxn
    deg     = panel_degradation_pct / 100

    years   = list(range(1, vida_util + 1))
    costo_red_y1 = grid_import_kwh * tarifa_mxn_kwh

    ahorro_y = []
    om_y     = []
    rep_y    = []
    fn_y     = []
    fd_y     = []

    for y in years:
        # Degradación compuesta año a año (modelo IEC 61724 / NREL):
        # factor = (1 - deg)^(y-1)  — no lineal, porque cada año degrada sobre el anterior.
        # El modelo lineal (1 - deg*(y-1)) sobreestima la generación en años tardíos
        # y puede volverse negativo para degradaciones altas y horizontes largos.
        factor_deg  = max(0.0, (1 - deg) ** (y - 1))
        # Ahorro: baseline CFE − costo red con proyecto (ambos con inflación)
        base_y = baseline_cost_mxn * (1 + inf) ** (y - 1)
        red_y  = costo_red_y1 * factor_deg * (1 + inf) ** (y - 1)
        ahorro = base_y - red_y
        om     = inv_mxn * (om_pct / 100) * (1 + inf) ** (y - 1)
        # Reemplazo BESS si aplica
        rep    = (bess_capex_energia_usd * bess_reemplazo_pct * usd_to_mxn
                  if y == bess_año_reemplazo else 0.0)
        fn     = ahorro - om - rep
        fd     = fn / (1 + r) ** y
        ahorro_y.append(ahorro)
        om_y.append(om)
        rep_y.append(rep)
        fn_y.append(fn)
        fd_y.append(fd)

    vpn = -inv_mxn + sum(fd_y)

    # TIR — función compartida a nivel de módulo (_bisection_irr)
    tir = _bisection_irr([-inv_mxn] + fn_y)

    acum = -inv_mxn
    pb_s = None
    for i, fn in enumerate(fn_y):
        prev = acum; acum += fn
        if acum >= 0 and pb_s is None:
            pb_s = round(years[i] - 1 + (-prev) / (acum - prev), 1)

    acum_d = -inv_mxn
    pb_d   = None
    acum_desc_list = []
    for i, fd in enumerate(fd_y):
        prev = acum_d; acum_d += fd
        acum_desc_list.append(acum_d)
        if acum_d >= 0 and pb_d is None:
            pb_d = round(years[i] - 1 + (-prev) / (acum_d - prev), 1)

    acum_nom = []
    run = -inv_mxn
    for fn in fn_y:
        run += fn; acum_nom.append(run)

    # LCOE / LCOS-like: costo nivelado por kWh ahorrado.
    # FIX: se usa degradación exponencial (1-deg)^(y-1) — consistente con el cálculo
    # de factor_deg en el loop de flujos. El modelo lineal (1 - deg*(y-1)) sobreestima
    # el denominador y subestima el LCOE.
    pv_kwh = sum(
        (grid_import_kwh * max(0.0, (1 - deg) ** (y - 1))) / (1 + r) ** y
        for y in years
    )
    pv_cost = inv_mxn + sum(om_y[i] / (1+r)**years[i] for i in range(len(years)))
    if bess_año_reemplazo <= vida_util:
        pv_cost += bess_capex_energia_usd * bess_reemplazo_pct * usd_to_mxn / (1+r)**bess_año_reemplazo
    lcoe = pv_cost / pv_kwh if pv_kwh > 0 else 0

    return dict(
        vpn=round(vpn, 0), tir=tir, pb_simple=pb_s, pb_disc=pb_d, lcoe=round(lcoe, 4),
        capex_mxn=round(inv_mxn, 0),
        years=years, ahorro_y=ahorro_y, om_y=om_y, rep_y=rep_y,
        fn_y=fn_y, fd_y=fd_y,
        acum_nom=acum_nom, acum_desc=acum_desc_list,
    )


# ═════════════════════════════════════════════════════════════════════════════
# TAB 1 — PRE-SIZING / TOR
# ═════════════════════════════════════════════════════════════════════════════
with tab1:
    col_p, col_r = st.columns([1, 1.8], gap="large")

    with col_p:
        # ── Datos del proyecto ─────────────────────────────────────────────────
        st.markdown('<div class="section-header">Datos del proyecto</div>', unsafe_allow_html=True)
        proj_loc  = st.text_input("Ubicación / dirección", value=f"{lat:.4f}, {lon:.4f}")

        # ── Mapa de ubicación ─────────────────────────────────────────────────
        st.markdown('<div class="section-header">Mapa de ubicación</div>', unsafe_allow_html=True)
        fig_map = go.Figure(go.Scattermapbox(
            lat=[lat], lon=[lon],
            mode="markers",
            marker=dict(size=14, color=AMBER, opacity=0.95),
            text=["Ubicación del proyecto"],
            hovertemplate="<b>%{text}</b><br>Lat: %{lat:.4f}<br>Lon: %{lon:.4f}<extra></extra>",
        ))
        fig_map.update_layout(
            mapbox=dict(
                style="carto-positron",
                center=dict(lat=lat, lon=lon),
                zoom=11,
            ),
            margin=dict(l=0, r=0, t=0, b=0),
            height=220,
            paper_bgcolor="#0a0c10",
            plot_bgcolor="#ffffff",
            showlegend=False,
        )
        st.plotly_chart(fig_map, use_container_width=True, config={
            "displayModeBar": True,
            "displaylogo": False,
            "modeBarButtonsToRemove": ["toImage", "select2d", "lasso2d"],
            "scrollZoom": True,
        })

        # ── Modo de dimensionamiento ───────────────────────────────────────────
        st.markdown('<div class="section-header">Modo de dimensionamiento</div>', unsafe_allow_html=True)
        sizing_mode = st.radio(
            "Método de sizing",
            ["📐 Por área disponible", "🧾 Por datos del recibo CFE", "⚡ Alta precisión 8760h (CSV medidor)"],
            horizontal=False,
            help="'Por área': calcula cuántos paneles caben. 'Por recibo': dimensiona con consumo mensual real. '8760h': simulación horaria con tu CSV de medidor + TMY NASA."
        )
        uso_area    = sizing_mode == "📐 Por área disponible"
        uso_8760csv = sizing_mode == "⚡ Alta precisión 8760h (CSV medidor)"

        if uso_area:
            # ── Área disponible ────────────────────────────────────────────────
            st.markdown('<div class="section-header">Área disponible</div>', unsafe_allow_html=True)
            area_total = st.number_input("Área total disponible (m²)", 10.0, 50000.0, 200.0, 10.0)
            occ_factor = st.slider("Factor de ocupación (%)", 40, 95, 75, 5,
                help="% del área realmente aprovechable (sin obstáculos, accesos, bordes de seguridad)")
        else:
            # ── Datos del recibo CFE — histórico mensual ──────────────────────
            st.markdown('<div class="section-header">Histórico mensual (12 meses)</div>', unsafe_allow_html=True)
            st.markdown('<div class="nasa-box">📅 Ingresa el consumo y tarifa de cada mes. El importe se calcula automáticamente. Puedes obtener los datos de tus recibos bimestrales (divide entre 2) o del portal CFE.</div>', unsafe_allow_html=True)

            uso_historico = True  # único modo disponible

            # Defaults razonables
            cons_default = [500.0, 480.0, 460.0, 450.0, 470.0, 550.0,
                            600.0, 580.0, 510.0, 470.0, 460.0, 520.0]
            tar_default  = [2.80] * 12

            st.markdown('<div class="section-header">Consumo y tarifa mensual</div>', unsafe_allow_html=True)
            st.caption("Edita Consumo y Tarifa — el Importe se recalcula automáticamente.")

            # Construir df base
            # ── Tabla editable: solo Consumo y Tarifa son editables ──────────
            # El Importe se recalcula automáticamente cada render.
            df_input = pd.DataFrame({
                "Mes":              MONTHS,
                "Consumo (kWh)":    cons_default,
                "Tarifa (MXN/kWh)": tar_default,
            })

            df_edit = st.data_editor(
                df_input,
                column_config={
                    "Mes":              st.column_config.TextColumn(disabled=True),
                    "Consumo (kWh)":    st.column_config.NumberColumn(
                        min_value=0.0, max_value=2_000_000.0, step=10.0, format="%.0f"),
                    "Tarifa (MXN/kWh)": st.column_config.NumberColumn(
                        min_value=0.0, max_value=50.0, step=0.001, format="%.3f",
                        help="Precio medio pagado ese mes"),
                },
                hide_index=True, use_container_width=True, key="hist_mensual",
                num_rows="fixed",
            )

            cons_edit = [float(v) for v in df_edit["Consumo (kWh)"].tolist()]
            tar_edit  = [float(v) for v in df_edit["Tarifa (MXN/kWh)"].tolist()]
            imp_calc  = [round(cons_edit[i] * tar_edit[i], 0) for i in range(12)]

            # Mostrar tabla con importe calculado (read-only, siempre actualizada)
            df_show = pd.DataFrame({
                "Mes":              MONTHS,
                "Consumo (kWh)":    [f"{v:,.0f}" for v in cons_edit],
                "Tarifa (MXN/kWh)": [f"${v:.3f}" for v in tar_edit],
                "Importe (MXN)":    [f"${v:,.0f}" for v in imp_calc],
            })
            st.dataframe(df_show, use_container_width=True, hide_index=True)
            st.caption("↑ Importe = Consumo × Tarifa (actualizado automáticamente)")

            monthly_cons_input = tuple(cons_edit)
            monthly_tar_input  = tuple(tar_edit)

            # Resumen rápido
            consumo_anual_hist = sum(monthly_cons_input)
            gasto_anual_hist   = sum(imp_calc)
            tar_media_hist     = gasto_anual_hist / consumo_anual_hist if consumo_anual_hist > 0 else 0
            mes_max_idx        = monthly_cons_input.index(max(monthly_cons_input))
            mes_min_idx        = monthly_cons_input.index(min(monthly_cons_input))
            st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-top:8px;">
  <div class="snap-card" style="min-height:72px;padding:10px 8px;">
    <div class="sc-label">Consumo anual</div>
    <div class="sc-val" style="font-size:clamp(12px,1.1vw,16px);">{consumo_anual_hist:,.0f}</div>
    <div class="sc-sub">kWh/año</div>
  </div>
  <div class="snap-card" style="min-height:72px;padding:10px 8px;">
    <div class="sc-label">Gasto anual CFE</div>
    <div class="sc-val" style="font-size:clamp(12px,1.1vw,16px);">${gasto_anual_hist:,.0f}</div>
    <div class="sc-sub">MXN/año</div>
  </div>
  <div class="snap-card" style="min-height:72px;padding:10px 8px;">
    <div class="sc-label">Tarifa media real</div>
    <div class="sc-val" style="color:#f59e0b;font-size:clamp(12px,1.1vw,16px);">${tar_media_hist:.3f}</div>
    <div class="sc-sub">MXN/kWh ponderada</div>
  </div>
  <div class="snap-card" style="min-height:72px;padding:10px 8px;">
    <div class="sc-label">Mes pico / valle</div>
    <div class="sc-val" style="font-size:clamp(11px,1vw,14px);">{MONTHS[mes_max_idx]} / {MONTHS[mes_min_idx]}</div>
    <div class="sc-sub">{max(monthly_cons_input):,.0f} / {min(monthly_cons_input):,.0f} kWh</div>
  </div>
</div>
""", unsafe_allow_html=True)


            # ── Slider kWp manual ──────────────────────────────────────────────
            st.markdown('<div class="section-header">Tamaño del sistema</div>', unsafe_allow_html=True)

            # Referencia: kWp para ~80% cobertura con irradiancia media
            _irr_prom = sum(active_irr) / 12 if sum(active_irr) > 0 else 5.0
            _kwp_ref  = consumo_anual_hist * 0.80 / (_irr_prom * effective_pr * 365)
            _kwp_ref  = max(1.0, _kwp_ref)
            _kwp_max  = max(20.0, round(_kwp_ref * 2.5 / 5) * 5)

            kwp_manual = st.slider(
                "Capacidad del sistema (kWp)",
                min_value=1.0, max_value=float(_kwp_max),
                value=float(max(1.0, round(_kwp_ref / 5) * 5)),
                step=0.5,
                help="Mueve para explorar la cobertura solar. El sistema se redondea al número entero de paneles.",
                key="kwp_slider_recibo",
            )
            # Preview instantáneo — kWp y cobertura con fórmula correcta (capeada por mes)
            _n_prev   = max(1, round(kwp_manual * 1000 / panel_wp))
            _kwp_prev = _n_prev * panel_wp / 1000
            _gen_prev = [_kwp_prev * active_irr[m] * effective_pr * MONTH_DAYS[m] for m in range(12)]
            # Cobertura correcta: min(gen, consumo) mes a mes / consumo total
            # No usar gen_total/consumo_total — sobreestima cuando hay excedentes mensuales
            _cob_prev = (sum(min(_gen_prev[m], monthly_cons_input[m]) for m in range(12))
                         / max(consumo_anual_hist, 1) * 100)
            _cc = "#4ade80" if _cob_prev >= 80 else ("#facc15" if _cob_prev >= 50 else "#f87171")
            st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:6px 0 4px;">
  <div class="snap-card" style="min-height:72px;">
    <div class="sc-label">Paneles</div>
    <div class="sc-val" style="font-size:16px;color:#f1f5f9;">{_n_prev}</div>
    <div class="sc-sub">{panel_wp} Wp c/u</div>
  </div>
  <div class="snap-card" style="min-height:72px;">
    <div class="sc-label">kWp instalados</div>
    <div class="sc-val" style="font-size:16px;color:#f1f5f9;">{_kwp_prev:.1f}</div>
    <div class="sc-sub">kWp</div>
  </div>
  <div class="snap-card" style="min-height:72px;">
    <div class="sc-label">Cobertura anual</div>
    <div class="sc-val" style="font-size:16px;color:{_cc};">{_cob_prev:.1f}%</div>
    <div class="sc-sub">del consumo total</div>
  </div>
</div>
""", unsafe_allow_html=True)

            area_total      = 0.0
            occ_factor      = 75
            solar_pct       = 80          # legacy — no usado en este modo
            sizing_strategy = "Promedio anual (económico)"  # legacy

        # ══════════════════════════════════════════════════════════════════════
        # MODO 3 — ALTA PRECISIÓN 8760h (CSV MEDIDOR + TMY NASA)
        # ══════════════════════════════════════════════════════════════════════
        if uso_8760csv:
            st.markdown('<div class="section-header">CSV de demanda horaria (medidor)</div>',
                        unsafe_allow_html=True)
            st.markdown("""
<div class="nasa-box">
  ⚡ <b>Modo 8760h:</b> Sube el CSV de tu medidor (una fila por hora del año).
  La irradiancia horaria se descarga automáticamente de NASA POWER con las coordenadas del sidebar.
  Columnas requeridas: <code>datetime</code> y <code>load_kW</code>.
  Formato de fecha soportado: DD/MM/YYYY HH:MM o YYYY-MM-DD HH:MM.
</div>
""", unsafe_allow_html=True)

            csv_file_8760 = st.file_uploader(
                "Sube tu CSV de demanda 8760h",
                type=["csv"],
                key="csv_8760_tab1",
                help="CSV con columnas datetime y load_kW — mínimo 8,760 filas (1 por hora del año)"
            )

            # Parámetros del sistema — siempre visibles para que el usuario los ajuste
            st.markdown('<div class="section-header">Tamaño del sistema</div>', unsafe_allow_html=True)
            _irr_prom_8760 = sum(active_irr) / 12 if sum(active_irr) > 0 else 5.0

            kwp_8760 = st.slider(
                "Capacidad PV (kWp)",
                min_value=10.0, max_value=50000.0,
                value=500.0, step=10.0,
                key="kwp_slider_8760",
                help="Mueve para explorar diferentes tamaños de sistema"
            )
            tarifa_8760 = st.number_input(
                "Tarifa CFE promedio (MXN/kWh)",
                min_value=0.5, max_value=20.0, value=float(tarifa), step=0.1,
                key="tarifa_8760",
                help="Tarifa media para calcular el ahorro económico"
            )

            # Inicializar variables del modo 8760h
            df_8760_valid     = None
            tmy_df_8760       = None
            pv_gen_8760_csv   = None
            dispatch_8760_csv = None
            _8760_ready       = False
            _8760_error       = ""

            if csv_file_8760 is not None:
                try:
                    df_raw = pd.read_csv(csv_file_8760)
                    df_raw.columns = [c.strip().lower().replace(" ", "_") for c in df_raw.columns]

                    # Detectar columna datetime
                    dt_col = next((c for c in ["datetime", "timestamp", "date", "hora", "time", "fecha"]
                                   if c in df_raw.columns), None)
                    # Detectar columna de carga
                    load_col = next((c for c in ["load_kw", "load", "consumo_kw", "demand_kw", "kw",
                                                  "potencia_kw", "demanda_kw"]
                                     if c in df_raw.columns), None)

                    if dt_col is None:
                        _8760_error = "No se encontró columna de fecha. Nombra la columna 'datetime'."
                    elif load_col is None:
                        _8760_error = "No se encontró columna de demanda. Nombra la columna 'load_kW'."
                    elif len(df_raw) < 8760:
                        _8760_error = f"El CSV tiene {len(df_raw)} filas — se requieren ≥ 8,760."
                    else:
                        df_clean = df_raw.copy()
                        # Parsear fecha con dayfirst automático
                        try:
                            df_clean["dt"] = pd.to_datetime(df_clean[dt_col], dayfirst=True, errors="coerce")
                        except Exception:
                            df_clean["dt"] = pd.to_datetime(df_clean[dt_col], errors="coerce")

                        df_clean["load_kW"] = pd.to_numeric(df_clean[load_col], errors="coerce").fillna(0.0).clip(lower=0)
                        df_clean = df_clean.sort_values("dt").head(8760).reset_index(drop=True)

                        # Detectar y reportar ceros (paros)
                        n_zeros = (df_clean["load_kW"] == 0).sum()
                        n_near  = (df_clean["load_kW"] < df_clean["load_kW"].quantile(0.01) * 0.3).sum()

                        # Estadísticas rápidas
                        dem_mean = df_clean["load_kW"].mean()
                        dem_max  = df_clean["load_kW"].max()
                        dem_min_nonzero = df_clean[df_clean["load_kW"] > 0]["load_kW"].min()
                        consumo_anual_8760 = df_clean["load_kW"].sum()  # kWh (cada fila = 1h)

                        # Dashboard diagnóstico del CSV
                        st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:8px 0;">
  <div class="snap-card" style="min-height:68px;padding:8px;">
    <div class="sc-label">Consumo anual</div>
    <div class="sc-val" style="font-size:clamp(12px,1.1vw,16px);">{consumo_anual_8760:,.0f}</div>
    <div class="sc-sub">kWh/año</div>
  </div>
  <div class="snap-card" style="min-height:68px;padding:8px;">
    <div class="sc-label">Demanda media</div>
    <div class="sc-val" style="font-size:clamp(12px,1.1vw,16px);">{dem_mean:,.0f}</div>
    <div class="sc-sub">kW promedio</div>
  </div>
  <div class="snap-card" style="min-height:68px;padding:8px;">
    <div class="sc-label">Demanda pico</div>
    <div class="sc-val" style="font-size:clamp(12px,1.1vw,16px);color:#f59e0b;">{dem_max:,.0f}</div>
    <div class="sc-sub">kW máximo</div>
  </div>
  <div class="snap-card" style="min-height:68px;padding:8px;">
    <div class="sc-label">Horas de paro</div>
    <div class="sc-val" style="font-size:clamp(12px,1.1vw,16px);color:{"#f87171" if n_zeros>0 else "#4ade80"};">{n_zeros}</div>
    <div class="sc-sub">horas con carga = 0</div>
  </div>
</div>
""", unsafe_allow_html=True)

                        if n_zeros > 0:
                            st.caption(f"ℹ️ {n_zeros} horas con demanda = 0 detectadas (paros de planta). "
                                       f"Se conservan tal cual — el solar en esas horas irá a excedente.")

                        df_8760_valid = df_clean
                        _8760_ready   = True

                except Exception as e:
                    _8760_error = f"Error al leer el CSV: {str(e)[:120]}"

            if _8760_error:
                st.error(f"⚠️ {_8760_error}")

            # Variables legacy para que el resto del tab no rompa
            area_total      = 0.0
            occ_factor      = 75
            solar_pct       = 80
            sizing_strategy = "Alta precisión 8760h"
            monthly_cons_input = tuple([0.0] * 12)
            monthly_tar_input  = tuple([tarifa_8760] * 12)
            uso_historico      = False
            kwp_manual         = kwp_8760

        # ── Irradiancia (siempre visible) ──────────────────────────────────────
        st.markdown('<div class="section-header">Irradiancia mensual (kWh/m²/día)</div>',
                    unsafe_allow_html=True)
        irr_source_banner()
        irr_df1 = pd.DataFrame({"Mes": MONTHS, "Irradiancia (kWh/m²/día)": active_irr})
        irr_ed1 = st.data_editor(irr_df1, column_config={
            "Mes": st.column_config.TextColumn(disabled=True),
            "Irradiancia (kWh/m²/día)": st.column_config.NumberColumn(
                min_value=0.0, max_value=10.0, step=0.0001, format="%.4f"),
        }, hide_index=True, use_container_width=True, key="irr1")
        irr_vals = irr_ed1["Irradiancia (kWh/m²/día)"].tolist()

    # ── Calculations — todo cacheado ───────────────────────────────────────────
    with col_r:

        irr_tuple = tuple(irr_vals)

        # ── MODO 8760h: simulación horaria completa ───────────────────────────
        if uso_8760csv:
            if not _8760_ready or df_8760_valid is None:
                st.info("📂 Sube tu CSV de demanda en el panel izquierdo para ver resultados.")
                st.stop()

            # Construir TMY 8760h desde NASA (mismo motor que Tab 6)
            with st.spinner("🌍 Descargando TMY NASA POWER y simulando 8760h..."):
                tmy_df, tmy_err = build_tmy_8760(
                    lat, lon,
                    irr_media_mensual=tuple(active_irr),
                )

            if tmy_err and tmy_df is None:
                st.error(f"❌ No se pudo construir el TMY: {tmy_err}")
                st.stop()

            if tmy_err:
                st.caption(f"ℹ️ TMY: {tmy_err}")

            irr_8760_csv  = tuple(tmy_df["irradiance_Wm2"].values[:8760].tolist())
            temp_8760_csv = tuple(tmy_df["temp_C"].values[:8760].tolist())
            load_8760_csv = tuple(df_8760_valid["load_kW"].values[:8760].tolist())

            # Simular generación PV hora a hora
            n_panels_8760 = max(1, round(kwp_8760 * 1000 / panel_wp))
            kwp_8760_real = n_panels_8760 * panel_wp / 1000
            area_used_8760 = n_panels_8760 * panel_area

            pv_gen_8760_csv = simulate_pv_8760(
                irr_8760=irr_8760_csv,
                temp_8760=temp_8760_csv,
                kwp=kwp_8760_real,
                pr_base=effective_pr,
            )

            # Despacho PV puro (sin batería)
            dispatch_8760_csv = dispatch_pv_bess_8760(
                load_8760=load_8760_csv,
                pv_8760=pv_gen_8760_csv,
                bess_kwh=0.001, bess_kw=0.001,
                rte=1.0, dod=1.0,
            )

            # Métricas derivadas
            gen_anual_8760     = dispatch_8760_csv["pv_total_kwh"]
            autoconsumo_kwh    = sum(dispatch_8760_csv["pv_direct"])
            excedente_kwh      = dispatch_8760_csv["grid_export_kwh"]
            grid_import_kwh    = dispatch_8760_csv["grid_import_kwh"]
            autoconsumo_pct    = dispatch_8760_csv["autoconsumo_pct"]
            autosufic_pct      = dispatch_8760_csv["autosuficiencia_pct"]
            curtail_pct        = dispatch_8760_csv["curtailment_pct"]
            consumo_anual_8760 = dispatch_8760_csv["load_total_kwh"]

            ahorro_8760        = autoconsumo_kwh * tarifa_8760
            inversion_usd_8760 = kwp_8760_real * costo_kwp
            inversion_mxn_8760 = inversion_usd_8760 * usd_to_mxn
            co2_saved_8760     = gen_anual_8760 * CO2_FACTOR_KG_KWH / 1000  # toneladas

            # ── HERO CARD 8760h ────────────────────────────────────────────────
            st.markdown(f"""
<div class="tor-hero">
  <div class="th-project">⚡ PRE-SIZING · ALTA PRECISIÓN 8760h · SIMULACIÓN HORARIA NASA</div>
  <div class="th-meta">
    {proj_loc} &nbsp;·&nbsp;
    <span class="pr-badge pr-green">● PR {pr_pct:.1f}%</span>
    &nbsp;·&nbsp; Tarifa: <b>${tarifa_8760:.3f}/kWh</b>
    &nbsp;·&nbsp; TMY NASA POWER 2005–{NASA_END}
  </div>
  <div class="th-grid">
    <div class="th-item">
      <span class="th-label">CAPACIDAD PICO</span>
      <span class="th-val">{kwp_8760_real:,.1f}</span>
      <span class="th-unit">kWp · {n_panels_8760} paneles</span>
    </div>
    <div class="th-item">
      <span class="th-label">GENERACIÓN ANUAL</span>
      <span class="th-val">{gen_anual_8760/1000:,.1f}</span>
      <span class="th-unit">MWh/año · simulación horaria</span>
    </div>
    <div class="th-item">
      <span class="th-label">AUTOCONSUMO REAL</span>
      <span class="th-val">{autoconsumo_kwh/1000:,.1f}</span>
      <span class="th-unit">MWh/año · {autoconsumo_pct:.1f}% de la generación</span>
    </div>
    <div class="th-item">
      <span class="th-label">AUTOSUFICIENCIA</span>
      <span class="th-val">{autosufic_pct:.1f}</span>
      <span class="th-unit">% del consumo cubierto</span>
    </div>
    <div class="th-item">
      <span class="th-label">EXCEDENTE A RED</span>
      <span class="th-val">{excedente_kwh/1000:,.1f}</span>
      <span class="th-unit">MWh/año · {curtail_pct:.1f}% de generación</span>
    </div>
    <div class="th-item">
      <span class="th-label">INVERSIÓN REF.</span>
      <span class="th-val">${inversion_usd_8760:,.0f}</span>
      <span class="th-unit">USD · ≈ ${inversion_mxn_8760:,.0f} MXN</span>
    </div>
    <div class="th-item">
      <span class="th-label">AHORRO AÑO 1</span>
      <span class="th-val">${ahorro_8760:,.0f}</span>
      <span class="th-unit">MXN · sobre autoconsumo real</span>
    </div>
    <div class="th-item">
      <span class="th-label">ÁREA REQUERIDA</span>
      <span class="th-val">{area_used_8760 / (occ_factor/100):,.0f}</span>
      <span class="th-unit">m² brutos · {area_used_8760:,.0f} m² netos (FC {occ_factor}%)</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

            # ── Nota importante sobre autoconsumo vs generación ───────────────
            if curtail_pct > 15:
                st.warning(f"⚠️ El {curtail_pct:.1f}% de la generación va a excedente de red. "
                           f"Considera reducir el kWp o agregar BESS para mejorar el autoconsumo.")

            # ── Gráfica: perfil semanal verano e invierno ─────────────────────
            st.markdown('<div class="section-header">Perfil horario — semana típica</div>',
                        unsafe_allow_html=True)

            load_list = list(load_8760_csv)
            pv_list   = list(pv_gen_8760_csv)
            grid_list = list(dispatch_8760_csv["grid_import"])
            ac_list   = list(dispatch_8760_csv["pv_direct"])

            # Verano: semana 26 (junio) · Invierno: semana 2 (enero)
            week_labels = {"Verano (jun)": 25*7*24, "Invierno (ene)": 1*7*24}
            fig_week = go.Figure()
            colors_season = {"Verano (jun)": AMBER, "Invierno (ene)": BLUE}
            for season, h_start in week_labels.items():
                h_end  = min(h_start + 7*24, 8760)
                horas  = list(range(h_start, h_end))
                fig_week.add_trace(go.Scatter(
                    x=horas, y=[load_list[h] for h in horas],
                    name=f"Demanda {season}", mode="lines",
                    line=dict(color=colors_season[season], width=1.5, dash="dot"),
                ))
                fig_week.add_trace(go.Scatter(
                    x=horas, y=[pv_list[h] for h in horas],
                    name=f"Solar {season}", mode="lines",
                    line=dict(color=TEAL if season.startswith("V") else "#60a5fa", width=2),
                    fill="tozeroy", fillcolor="rgba(20,184,166,0.08)" if season.startswith("V") else "rgba(96,165,250,0.06)",
                ))
            lyt_week = copy.deepcopy(PLOT_LAYOUT)
            lyt_week.update({
                "height": 280,
                "yaxis": dict(title="kW", gridcolor="#1e2230"),
                "xaxis": dict(title="Hora del año", gridcolor="#1e2230"),
                "legend": dict(orientation="h", y=1.12, bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=10, r=10, t=40, b=40),
            })
            fig_week.update_layout(**lyt_week)
            st.plotly_chart(fig_week, use_container_width=True)

            # ── Curva de duración de generación ──────────────────────────────
            st.markdown('<div class="section-header">Curva de duración de generación</div>',
                        unsafe_allow_html=True)
            st.plotly_chart(
                duration_curve_fig(pv_list, "Generación PV", AMBER),
                use_container_width=True
            )

            # ── Heatmap generación 24h × 365d ─────────────────────────────────
            st.markdown('<div class="section-header">Heatmap de generación horaria</div>',
                        unsafe_allow_html=True)
            st.plotly_chart(
                build_heatmap_fig(pv_list, "Generación PV", "kW", "YlOrBr"),
                use_container_width=True
            )

            # ── Modelo financiero (mismo que modos 1 y 2) ─────────────────────
            # El ahorro financiero se calcula sobre el autoconsumo real (no la generación total)
            st.markdown("---")
            fm_8760 = calc_financial_model(
                autoconsumo_kwh, kwp_8760_real, float(inversion_usd_8760),
                tarifa_8760, inflation, discount_rate,
                panel_degradation, vida_util, usd_to_mxn,
                om_pct=om_pct_sidebar
            )
            vpn_8   = fm_8760["vpn"]
            tir_8   = fm_8760["tir"]
            pb_s_8  = fm_8760["pb_simple"]
            pb_d_8  = fm_8760["pb_disc"]
            lcoe_8  = fm_8760["lcoe"]
            kc8     = "#4ade80" if vpn_8 > 0 else "#f87171"
            tir_s8  = f"{tir_8:.1f}%" if tir_8 else "N/A"
            pbs_s8  = f"{pb_s_8:.1f} años" if pb_s_8 else f">{vida_util} años"
            pbd_s8  = f"{pb_d_8:.1f} años" if pb_d_8 else f">{vida_util} años"

            st.markdown(f"""
<div style="margin-bottom:0.5rem">
  <div class="section-header">Modelo financiero · {vida_util} años · base autoconsumo horario real</div>
</div>
<div style="display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:12px;">
  <div class="snap-card"><div class="sc-label">VPN</div>
    <div class="sc-val" style="color:{kc8};">${vpn_8:,.0f}</div><div class="sc-sub">MXN</div></div>
  <div class="snap-card"><div class="sc-label">TIR</div>
    <div class="sc-val" style="color:#22d3ee;">{tir_s8}</div><div class="sc-sub">vs {discount_rate}% WACC</div></div>
  <div class="snap-card"><div class="sc-label">LCOE</div>
    <div class="sc-val" style="color:{VIOLET};">${lcoe_8:.2f}</div><div class="sc-sub">MXN/kWh</div></div>
  <div class="snap-card"><div class="sc-label">Payback simple</div>
    <div class="sc-val">{pbs_s8}</div><div class="sc-sub">nominal</div></div>
  <div class="snap-card"><div class="sc-label">Payback desc.</div>
    <div class="sc-val">{pbd_s8}</div><div class="sc-sub">descontado</div></div>
  <div class="snap-card"><div class="sc-label">CO₂ evitado</div>
    <div class="sc-val" style="color:#4ade80;">{co2_saved_8760:.1f}</div><div class="sc-sub">ton/año</div></div>
</div>
""", unsafe_allow_html=True)

            # Gráfica de flujos acumulados
            fig_acum8 = go.Figure()
            fig_acum8.add_trace(go.Scatter(
                x=fm_8760["years"], y=fm_8760["acum_nominal"],
                name="Acumulado nominal", mode="lines+markers",
                line=dict(color=AMBER, width=2),
                hovertemplate="<b>Año %{x}</b><br>Acum.: $%{y:,.0f} MXN<extra></extra>",
            ))
            fig_acum8.add_trace(go.Scatter(
                x=fm_8760["years"], y=fm_8760["acum_desc"],
                name="Acumulado descontado", mode="lines+markers",
                line=dict(color=TEAL, width=2, dash="dot"),
                hovertemplate="<b>Año %{x}</b><br>Acum. desc.: $%{y:,.0f} MXN<extra></extra>",
            ))
            fig_acum8.add_hline(y=0, line_color="#475569", line_dash="dash", line_width=1)
            lay8 = copy.deepcopy(PLOT_LAYOUT)
            lay8.update({
                "height": 280,
                "yaxis": dict(title="MXN acumulado", gridcolor="#1e2230", tickformat=","),
                "xaxis": dict(title="Año", gridcolor="#1e2230", tickmode="linear"),
                "legend": dict(orientation="h", y=1.12, bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=10, r=10, t=40, b=40),
            })
            fig_acum8.update_layout(**lay8)
            st.plotly_chart(fig_acum8, use_container_width=True)

            st.stop()   # El modo 8760h tiene su propio render — no cae al bloque común

        if uso_area:
            sz = calc_sizing_area(area_total, occ_factor, panel_wp, panel_area,
                                  irr_tuple, effective_pr)
            monthly_cons_ref = None
            monthly_tar_ref  = None
            tarifa_efectiva  = tarifa
            uso_historico_r  = False
        else:
            uso_historico_r = uso_historico
            if uso_historico:
                _ok_r, _msg_r = _validate_recibo_inputs(monthly_cons_input, monthly_tar_input)
                if not _ok_r:
                    st.error(f"⚠️ Datos de recibo inválidos: {_msg_r}")
                    st.stop()
                sz = calc_sizing_recibo_kwp(
                    monthly_cons_input, monthly_tar_input,
                    max(kwp_manual, 0.5),
                    panel_wp, panel_area,
                    irr_tuple, effective_pr,
                    occ_factor=occ_factor)  # FIX: pasar factor de ocupación del usuario
                monthly_cons_ref = sz["monthly_cons"]
                monthly_tar_ref  = sz["monthly_tarifas"]
                tarifa_efectiva  = sz["tarifa_media_pond"]


        n_panels   = sz["n_panels"]
        kwp        = sz["kwp"]
        area_util  = sz["area_util"]
        area_used  = sz["area_used"]
        monthly_gen = sz["monthly_gen"]
        annual_gen  = sz["annual_gen"]

        daily_avg   = annual_gen / 365
        co2_saved   = annual_gen * CO2_FACTOR_KG_KWH   # kg/año  (factor SEN SENER/CENACE)
        co2_saved_t = co2_saved / 1000     # toneladas/año
        # HSP promedio anual: promedio de los valores mensuales de irradiancia
        hsp_anual   = sum(irr_vals) / 12   # kWh/m²/día promedio anual
        if hsp_anual >= 5.5:
            hsp_nota = "✅ Excelente recurso solar (≥ 5.5 kWh/m²/día · referencia IEA/NREL)"
            hsp_color = "#4ade80"
        elif hsp_anual >= 4.5:
            hsp_nota = "🟡 Buen recurso solar (4.5–5.5 kWh/m²/día)"
            hsp_color = "#facc15"
        else:
            hsp_nota = "🔴 Recurso solar bajo (< 4.5 kWh/m²/día · validar viabilidad)"
            hsp_color = "#f87171"
        inversion_usd = kwp * costo_kwp
        inversion     = inversion_usd  # mantener alias USD para compatibilidad con TOR
        inversion_mxn = inversion_usd * usd_to_mxn
        # En modo histórico el ahorro real ya está calculado con tarifas mensuales reales
        if not uso_area and uso_historico_r and "ahorro_anual" in sz:
            ahorro1 = sz["ahorro_anual"]
        else:
            ahorro1 = annual_gen * tarifa_efectiva
        # FIX Bug 1: payback calculado en MXN/MXN → resultado en años (antes era USD/MXN)
        payback = inversion_mxn / ahorro1 if ahorro1 > 0 else 999
        tarifa_efectiva = _safe(tarifa_efectiva, fallback=2.80, min_val=0.01, label="tarifa efectiva")

        # ── P50 / P90 riguroso con serie interanual NASA POWER ────────────────
        # Si el usuario no ha cargado datos NASA, no hay serie histórica y
        # se muestra un aviso en lugar de un número inventado.
        p50_real, p90_real, gen_por_anio = compute_p90(
            active_irr_por_anio, kwp, effective_pr
        )
        has_p90 = p50_real is not None
        # Para el TOR usamos P50 = generación con irr media editada
        p50 = annual_gen
        p90 = p90_real if has_p90 else None

        # ── TOR HERO — resultados en encabezado ───────────────────────────────
        area_label = f"{area_used:.0f} m² de {area_util:.0f} útiles" if uso_area else f"{area_used:.0f} m² estimados"
        st.markdown(f"""
<div class="tor-hero">
  <div class="th-project">📋 PRE-SIZING • {"ÁREA DISPONIBLE" if uso_area else "RECIBO CFE"}</div>
  <div class="th-meta">
    {proj_loc} &nbsp;·&nbsp; <span class="pr-badge pr-green">● PR {pr_pct:.1f}%</span>
    {"" if uso_area else f"&nbsp;·&nbsp; Tarifa media: <b>${tarifa_efectiva:.3f}/kWh</b>"}
  </div>
  <div class="th-grid">
    <div class="th-item">
      <span class="th-label">CAPACIDAD PICO</span>
      <span class="th-val">{kwp:,.1f}</span>
      <span class="th-unit">kWp</span>
    </div>
    <div class="th-item">
      <span class="th-label">PANELES</span>
      <span class="th-val">{n_panels:,.0f}</span>
      <span class="th-unit">unidades • {panel_wp} Wp</span>
    </div>
    <div class="th-item">
      <span class="th-label">GENERACIÓN AÑO 1</span>
      <span class="th-val">{p50/1000:,.1f}</span>
      <span class="th-unit">MWh/año</span>
    </div>
    <div class="th-item">
      <span class="th-label">GENERACIÓN P90</span>
      <span class="th-val">{"—" if not has_p90 else f"{p90/1000:,.1f}"}</span>
      <span class="th-unit">{"Carga NASA" if not has_p90 else "MWh/año"}</span>
    </div>
    <div class="th-item">
      <span class="th-label">{"ÁREA UTILIZADA" if uso_area else "COBERTURA SOLAR"}</span>
      <span class="th-val">{"" if uso_area else f"{sz.get('cobertura_anual', 0):.1f}" if monthly_cons_ref else "—"}{f"{area_used:,.1f}" if uso_area else ""}</span>
      <span class="th-unit">{"m²" if uso_area else "% del consumo · min(gen,cons) por mes"}</span>
    </div>
    <div class="th-item">
      <span class="th-label">INVERSIÓN REF.</span>
      <span class="th-val">${inversion:,.0f}</span>
      <span class="th-unit">USD · ≈ ${inversion_mxn:,.0f} MXN</span>
    </div>
    <div class="th-item">
      <span class="th-label">AHORRO AÑO 1</span>
      <span class="th-val">${ahorro1:,.0f}</span>
      <span class="th-unit">MXN · tarifa ${tarifa_efectiva:.3f}/kWh</span>
    </div>
    <div class="th-item">
      <span class="th-label">PAYBACK SIMPLE</span>
      <span class="th-val">{payback:.1f}</span>
      <span class="th-unit">años · nominal<br><span style="font-size:9px;color:#475569;">Ver modelo financiero ↓ para payback descontado</span></span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── ALERTA REGULATORIA para Generación Distribuida ─────────────────────
        if kwp > 699:
            st.error("⚠️ **Proyecto mayor a 699 kWp** — supera el límite típico de Generación Distribuida en México.")
        if uso_area:
            max_kwp_gd = 699
            area_max_gd = max_kwp_gd * 1000 / panel_wp * panel_area / (occ_factor / 100)
            st.caption(f"📏 Área máx. para GD (<700 kWp): **{area_max_gd:,.0f} m²** (con {occ_factor}% ocupación)")
        else:
            consumo_anual_ref = sum(monthly_cons_ref) if monthly_cons_ref else 0
            cobertura_pct = annual_gen / max(consumo_anual_ref, 1) * 100
            st.caption(f"📊 Consumo anual estimado: **{consumo_anual_ref:,.0f} kWh** · Cobertura: **{cobertura_pct:.1f}%** · Tarifa media: **${tarifa_efectiva:.3f}/kWh**")
        # ── Panel activo (mini card) ────────────────────────────────────────────
        st.markdown(f"""
<div class="panel-card">
  <div class="pc-title">📋 Panel de referencia</div>
  <div class="pc-grid">
    <div class="pc-item"><span class="pc-label">Potencia</span>
      <span class="pc-val">{panel_wp} Wp</span></div>
    <div class="pc-item"><span class="pc-label">Eficiencia calc.</span>
      <span class="pc-val" style="color:{eff_color}">{panel_eff_calc:.2f}%
        <span style="font-size:10px;color:{eff_color}">{eff_note}</span></span></div>
    <div class="pc-item"><span class="pc-label">Dimensiones</span>
      <span class="pc-val">{panel_largo_mm}×{panel_ancho_mm} mm</span></div>
    <div class="pc-item"><span class="pc-label">Área unitaria</span>
      <span class="pc-val">{panel_area:.4f} m²</span></div>
    <div class="pc-item"><span class="pc-label">Peso</span>
      <span class="pc-val">{panel_peso_kg} kg</span></div>
    <div class="pc-item"><span class="pc-label">Densidad potencia</span>
      <span class="pc-val">{panel_wp/panel_area:.0f} Wp/m²</span></div>
  </div>
  <div style="border-top:1px solid #2a2d3a;margin-top:10px;padding-top:10px;">
    <div class="pc-title" style="margin-bottom:8px;">⚖️ Carga estructural estimada</div>
    <div class="pc-grid">
      <div class="pc-item">
        <span class="pc-label">Carga total (×1.35)</span>
        <span class="pc-val" style="color:#f87171;">{n_panels * panel_peso_kg * 1.35:,.0f} kg</span>
      </div>
      <div class="pc-item">
        <span class="pc-label">Carga total</span>
        <span class="pc-val" style="color:#f87171;">{n_panels * panel_peso_kg * 1.35 * 9.8 / 1000:,.2f} kN</span>
      </div>
      <div class="pc-item">
        <span class="pc-label">Carga por m² instalado</span>
        <span class="pc-val" style="color:#f87171;">{n_panels * panel_peso_kg * 1.35 * 9.8 / 1000 / max(n_panels * panel_area, 0.01):.3f} kN/m²</span>
      </div>
      <div class="pc-item">
        <span class="pc-label">Factor extra aplicado</span>
        <span class="pc-val" style="color:#9ca3af;">+35% montura/BOS</span>
      </div>
    </div>
  </div>

  <!-- HSP promedio anual -->
  <div style="border-top:1px solid #1e2230;margin-top:10px;padding-top:10px;">
    <div class="pc-title" style="margin-bottom:8px;">☀️ Recurso solar del sitio</div>
    <div class="pc-grid">
      <div class="pc-item">
        <span class="pc-label">HSP promedio anual</span>
        <span class="pc-val" style="color:{hsp_color};">{hsp_anual:.2f} kWh/m²/día</span>
      </div>
      <div class="pc-item">
        <span class="pc-label">Evaluación</span>
        <span class="pc-val" style="font-size:11px;color:{hsp_color};font-family:Inter,sans-serif;">{hsp_nota}</span>
      </div>
    </div>
  </div>

  <!-- CO2 -->
  <div style="border-top:1px solid #1e2230;margin-top:10px;padding-top:10px;">
    <div class="pc-title" style="margin-bottom:8px;">🌿 Impacto ambiental año 1</div>
    <div class="pc-grid">
      <div class="pc-item">
        <span class="pc-label">CO₂ evitado</span>
        <span class="pc-val" style="color:#4ade80;">{co2_saved_t:,.2f} ton/año</span>
      </div>
      <div class="pc-item">
        <span class="pc-label">Factor de emisión utilizado</span>
        <span class="pc-val" style="font-size:11px;color:#64748b;font-family:Inter,sans-serif;">{CO2_FACTOR_KG_KWH} kg CO₂e/kWh · SEN 2024 · SEMARNAT/CRE 28-Feb-2025</span>
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Si modo recibo: mostrar comparativa consumo vs generación ───────────
        if not uso_area and monthly_cons_ref:
            st.markdown('<div class="section-header">Generación vs Consumo mensual</div>', unsafe_allow_html=True)

            # Usar datos detallados si están disponibles
            if uso_historico_r and "ahorro_mensual" in sz:
                excedente_m = sz["excedente"]
                coverage_m  = sz["cobertura_pct"]
                ahorro_m    = sz["ahorro_mensual"]
                energy_cov  = sz["energia_cubierta"]
            else:
                excedente_m = [monthly_gen[m] - monthly_cons_ref[m] for m in range(12)]
                coverage_m  = [min(100, monthly_gen[m] / max(monthly_cons_ref[m], 1) * 100) for m in range(12)]
                tar_mes     = monthly_tar_ref if monthly_tar_ref else [tarifa_efectiva]*12
                ahorro_m    = [min(monthly_gen[m], monthly_cons_ref[m]) * tar_mes[m] for m in range(12)]
                energy_cov  = [min(monthly_gen[m], monthly_cons_ref[m]) for m in range(12)]

            fig_cv = go.Figure()
            fig_cv.add_trace(go.Bar(x=MONTHS, y=monthly_cons_ref, name="Consumo",
                marker_color="#374151",
                hovertemplate="<b>%{x}</b><br>Consumo: %{y:,.0f} kWh<extra></extra>"))
            fig_cv.add_trace(go.Bar(x=MONTHS, y=energy_cov, name="Cubierto solar",
                marker_color=AMBER,
                hovertemplate="<b>%{x}</b><br>Cubierto: %{y:,.0f} kWh<extra></extra>"))
            fig_cv.add_trace(go.Scatter(x=MONTHS, y=monthly_gen, mode="lines+markers",
                name="Generación total", line=dict(color=TEAL, width=2, dash="dot"),
                marker=dict(size=6, color=TEAL),
                hovertemplate="<b>%{x}</b><br>Generación: %{y:,.0f} kWh<extra></extra>"))
            lyt_cv = copy.deepcopy(PLOT_LAYOUT)
            lyt_cv.update({"height": 270, "barmode": "overlay",
                           "yaxis": dict(title="kWh", gridcolor="#2a2d3a"),
                           "legend": dict(orientation="h", y=1.1, bgcolor="rgba(0,0,0,0)"),
                           "margin": dict(l=20, r=20, t=30, b=40)})
            fig_cv.update_layout(**lyt_cv)
            st.plotly_chart(fig_cv, use_container_width=True)

            # ── Resumen de ahorro por mes (reemplaza gráfica de tarifa) ──────────
            if uso_historico_r and monthly_tar_ref:
                st.markdown('<div class="section-header">Ahorro mensual estimado</div>', unsafe_allow_html=True)
                ahorro_html = '<div style="display:grid;grid-template-columns:repeat(6,1fr);gap:6px;margin-bottom:0.75rem;">'
                for i, m in enumerate(MONTHS):
                    color = "#4ade80" if ahorro_m[i] > 0 else "#f87171"
                    ahorro_html += f"""
  <div style="background:#111318;border:0.5px solid #1e2230;border-radius:10px;padding:8px 6px;text-align:center;">
    <div style="font-size:10px;color:#475569;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:4px;">{m}</div>
    <div style="font-size:13px;font-weight:600;color:{color};font-family:'JetBrains Mono',monospace;">${ahorro_m[i]:,.0f}</div>
    <div style="font-size:9px;color:#475569;margin-top:2px;">${monthly_tar_ref[i]:.3f}/kWh</div>
  </div>"""
                ahorro_html += '</div>'
                ahorro_html += f"""<div style="display:flex;justify-content:flex-end;font-size:12px;color:#94a3b8;margin-top:2px;">
  Ahorro anual total: <span style="color:#4ade80;font-weight:600;margin-left:6px;">${sum(ahorro_m):,.0f} MXN</span>
</div>"""
                st.markdown(ahorro_html, unsafe_allow_html=True)

            # ── Tabla mensual ──────────────────────────────────────────────────
            st.markdown('<div class="section-header">Tabla mensual detallada</div>', unsafe_allow_html=True)
            tar_display = monthly_tar_ref if monthly_tar_ref else [tarifa_efectiva]*12
            df_tabla = pd.DataFrame({
                "Mes":               MONTHS,
                "Consumo (kWh)":    [f"{v:,.0f}" for v in monthly_cons_ref],
                "Generación (kWh)": [f"{v:,.0f}" for v in monthly_gen],
                "Cubierto (kWh)":   [f"{v:,.0f}" for v in energy_cov],
                "Cobertura (%)":    [f"{v:.1f}%" for v in coverage_m],
                "Excedente (kWh)":  [f"+{v:,.0f}" if v >= 0 else f"{v:,.0f}" for v in excedente_m],
                "Tarifa ($/kWh)":   [f"${t:.3f}" for t in tar_display],
                "Ahorro (MXN)":     [f"${v:,.0f}" for v in ahorro_m],
            })
            st.dataframe(df_tabla, use_container_width=True, hide_index=True)

            # cobertura_anual = Σ min(gen_mes, cons_mes) / Σ cons_mes × 100
            # Fórmula correcta: no cuenta excedente que va a red como cobertura del consumo
            cobertura_anual = sz.get("cobertura_anual",
                sum(energy_cov) / max(sum(monthly_cons_ref), 1) * 100)
            st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-top:8px;">
  <div class="snap-card" style="min-height:80px;">
    <div class="sc-label">Consumo anual</div>
    <div class="sc-val" style="font-size:14px;">{sum(monthly_cons_ref):,.0f}</div>
    <div class="sc-sub">kWh/año</div>
  </div>
  <div class="snap-card" style="min-height:80px;">
    <div class="sc-label">Cobertura solar</div>
    <div class="sc-val" style="color:#f59e0b;font-size:14px;">{cobertura_anual:.1f}%</div>
    <div class="sc-sub">del consumo anual</div>
  </div>
  <div class="snap-card" style="min-height:80px;">
    <div class="sc-label">Ahorro anual</div>
    <div class="sc-val" style="color:#4ade80;font-size:14px;">${sum(ahorro_m):,.0f}</div>
    <div class="sc-sub">MXN/año</div>
  </div>
  <div class="snap-card" style="min-height:80px;">
    <div class="sc-label">Tarifa media real</div>
    <div class="sc-val" style="font-size:14px;">${tarifa_efectiva:.3f}</div>
    <div class="sc-sub">MXN/kWh ponderada</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Gráfica Mensual + Variabilidad interanual ─────────────────────────
        st.markdown('<div class="section-header">Generación mensual estimada</div>',
                    unsafe_allow_html=True)

        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            x=MONTHS, y=monthly_gen,
            name="Generación P50 (irr. media)",
            marker_color=AMBER, opacity=0.95,
            hovertemplate="<b>%{x}</b><br>P50: %{y:,.0f} kWh<extra></extra>"
        ))
        fig1.add_trace(go.Scatter(
            x=MONTHS, y=irr_vals,
            mode="lines+markers",
            name="Irradiancia NASA (2005–2024)",
            yaxis="y2",
            line=dict(color=ROSE, width=3, dash="dot"),
            marker=dict(size=7, color=ROSE, line=dict(width=1.5, color="white")),
            hovertemplate="<b>%{x}</b><br>Irradiancia: %{y:.4f} kWh/m²/día<extra></extra>"
        ))
        layout1 = copy.deepcopy(PLOT_LAYOUT)
        layout1.update({
            "height": 380, "barmode": "group",
            "title": dict(text="Generación Mensual (P50 con irradiancia media)", font=dict(size=15)),
            "yaxis":  dict(title="kWh generados", gridcolor="#2a2d3a", tickformat=",", rangemode="tozero"),
            "yaxis2": dict(title="Irradiancia (kWh/m²/día)", overlaying="y", side="right",
                           range=[0, max(irr_vals) * 1.25],
                           tickfont=dict(color=ROSE), tickformat=".2f"),
            "legend": dict(orientation="h", y=-0.22, x=0.5, xanchor="center", yanchor="top",
                           font=dict(size=13), bgcolor="rgba(0,0,0,0)",
                           bordercolor="#2a2d3a", borderwidth=1),
            "margin": dict(l=20, r=80, t=60, b=100),
            "hovermode": "x unified",
        })
        fig1.update_layout(**layout1)
        st.plotly_chart(fig1, use_container_width=True)

        # ── Distribución interanual + P90 real ────────────────────────────────
        if has_p90:
            st.markdown('<div class="section-header">Variabilidad interanual · P90 riguroso</div>',
                        unsafe_allow_html=True)
            n_anios = len(gen_por_anio)

            # Aviso metodológico
            st.markdown(
                f'<div class="nasa-box">🔬 P90 calculado como percentil 10 de la generación anual '
                f'simulada con los {n_anios} años de irradiancia real NASA POWER '
                f'({NASA_START}–{NASA_END}). '
                f'El sistema supera el P90 el 90% de los años históricos.</div>',
                unsafe_allow_html=True)

            anios  = list(gen_por_anio.keys())
            gen_v  = [gen_por_anio[y] / 1000 for y in anios]   # MWh
            p50_mwh = p50_real / 1000
            p90_mwh = p90_real / 1000

            # Colores: rojo si el año está por debajo del P90, ámbar si debajo del P50
            bar_colors = [
                ROSE  if v < p90_mwh else
                AMBER if v < p50_mwh else
                TEAL
                for v in gen_v
            ]

            fig_p90 = go.Figure()

            # Barras con bordes suaves y hover enriquecido
            fig_p90.add_trace(go.Bar(
                x=anios, y=gen_v,
                name="Generación anual",
                marker=dict(
                    color=bar_colors,
                    line=dict(color="rgba(0,0,0,0)", width=0),
                    cornerradius=4,
                ),
                hovertemplate=(
                    "<b>%{x}</b><br>"
                    "Generación: <b>%{y:,.1f} MWh</b><extra></extra>"
                ),
                showlegend=False,
            ))

            # Línea P50 — sin anotación flotante
            fig_p90.add_hline(
                y=p50_mwh,
                line_color=AMBER,
                line_dash="dash",
                line_width=1.8,
            )

            # Línea P90 — sin anotación flotante
            fig_p90.add_hline(
                y=p90_mwh,
                line_color=ROSE,
                line_dash="dot",
                line_width=1.8,
            )

            layout_p90 = copy.deepcopy(PLOT_LAYOUT)
            layout_p90.update({
                "height": 360,
                "yaxis": dict(
                    title="MWh/año",
                    gridcolor="#1e2130",
                    zeroline=False,
                    tickfont=dict(size=11),
                ),
                "xaxis": dict(
                    tickmode="linear",
                    dtick=1,
                    gridcolor="rgba(0,0,0,0)",
                    tickfont=dict(size=11),
                    tickangle=-45,
                ),
                "bargap": 0.25,
                "legend": dict(orientation="h", y=-0.18, x=0.5, xanchor="center",
                               bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=20, r=20, t=20, b=60),
                "hovermode": "x unified",
            })
            fig_p90.update_layout(**layout_p90)
            st.plotly_chart(fig_p90, use_container_width=True)

            # Leyenda manual debajo de la gráfica (limpia, sin anotaciones encima)
            st.markdown(f"""
<div style="display:flex; gap:28px; justify-content:center; flex-wrap:wrap;
            margin-top:-8px; margin-bottom:12px; font-size:12px; color:#9ca3af;">
  <span style="display:flex; align-items:center; gap:7px;">
    <span style="display:inline-block; width:26px; height:2px;
                 background:repeating-linear-gradient(90deg,{AMBER} 0,{AMBER} 6px,transparent 6px,transparent 10px);"></span>
    P50 = {p50_mwh:,.1f} MWh &nbsp;·&nbsp; mediana histórica
  </span>
  <span style="display:flex; align-items:center; gap:7px;">
    <span style="display:inline-block; width:26px; height:2px;
                 background:repeating-linear-gradient(90deg,{ROSE} 0,{ROSE} 4px,transparent 4px,transparent 7px);"></span>
    P90 = {p90_mwh:,.1f} MWh &nbsp;·&nbsp; superado el 90% de los años
  </span>
  <span style="display:flex; align-items:center; gap:7px;">
    <span style="display:inline-block; width:12px; height:12px; border-radius:3px; background:{TEAL};"></span>
    Por encima del P50
  </span>
  <span style="display:flex; align-items:center; gap:7px;">
    <span style="display:inline-block; width:12px; height:12px; border-radius:3px; background:{AMBER};"></span>
    Entre P90 y P50
  </span>
  <span style="display:flex; align-items:center; gap:7px;">
    <span style="display:inline-block; width:12px; height:12px; border-radius:3px; background:{ROSE};"></span>
    Por debajo del P90
  </span>
</div>
""", unsafe_allow_html=True)

            # Métricas compactas P90
            mp1, mp2, mp3, mp4 = st.columns(4)
            mp1.metric("P50",        f"{p50_mwh:,.1f} MWh")
            mp2.metric("P90", f"{p90_mwh:,.1f} MWh",
                       f"{(p90_real/p50_real - 1)*100:+.1f}% vs P50")
            mp3.metric("Mejor año",  f"{max(gen_v):,.1f} MWh")
            mp4.metric("Peor año",   f"{min(gen_v):,.1f} MWh")
        else:
            st.info("ℹ️ Carga datos de NASA POWER desde el sidebar para calcular el P90 riguroso con variabilidad interanual.")


    # ── Modelo financiero — cacheado ──────────────────────────────────────
    # Si P90 está disponible se usa como base conservadora (recomendado);
    # si no, se cae a P50 (generación con irradiancia media).
    gen_para_fm = p90_real if has_p90 else annual_gen
    fm_base_label = "P90" if has_p90 else "P50"
    fm = calc_financial_model(
        gen_para_fm, kwp, float(inversion),
        tarifa_efectiva, inflation, discount_rate,
        panel_degradation, vida_util, usd_to_mxn,
        om_pct=om_pct_sidebar
    )
    years         = fm["years"]
    gen_proj      = fm["gen_proj"]
    tarifas_y     = fm["tarifas_y"]
    flujo_nominal = fm["flujo_nominal"]
    om_anual      = fm["om_anual"]
    flujo_neto    = fm["flujo_neto"]
    factor_desc   = fm["factor_desc"]
    flujo_desc    = fm["flujo_desc"]
    acum_nominal  = fm["acum_nominal"]
    acum_desc     = fm["acum_desc"]
    inversion_mxn = fm["inv_mxn"]
    vpn           = fm["vpn"]
    tir           = fm["tir"]
    tir_str       = f"{tir:.1f}%" if tir is not None else "N/A"
    pb_simple     = fm["pb_simple"]
    pb_simple_str = f"{pb_simple:.1f} años" if pb_simple else f">{vida_util} años"
    pb_disc       = fm["pb_disc"]
    pb_disc_str   = f"{pb_disc:.1f} años" if pb_disc else f">{vida_util} años"
    lcoe          = fm["lcoe"]

    # ── 2. KPIs principales ───────────────────────────────────────────────
    kc = "#4ade80" if vpn > 0 else "#f87171"

    st.markdown(f"""
<div style="margin-bottom:0.5rem">
  <div class="section-header">Modelo financiero · {vida_util} años · base {fm_base_label}</div>
</div>
<div style="display:grid; grid-template-columns:repeat(6,1fr); gap:10px; margin-bottom:12px;">

  <div class="snap-card">
    <div class="sc-label">VPN</div>
    <div class="sc-val" style="color:{kc};">${vpn:,.0f}</div>
    <div class="sc-sub">MXN</div>
  </div>

  <div class="snap-card">
    <div class="sc-label">TIR</div>
    <div class="sc-val" style="color:#22d3ee;">{tir_str}</div>
    <div class="sc-sub">vs {discount_rate}% WACC</div>
  </div>

  <div class="snap-card">
    <div class="sc-label">LCOE</div>
    <div class="sc-val" style="color:{VIOLET};">${lcoe:.2f}</div>
    <div class="sc-sub">MXN/kWh generado</div>
  </div>

  <div class="snap-card">
    <div class="sc-label">Payback simple</div>
    <div class="sc-val" style="color:#f9fafb;">{pb_simple_str}</div>
    <div class="sc-sub">flujos nominales</div>
  </div>

  <div class="snap-card">
    <div class="sc-label">Payback descontado</div>
    <div class="sc-val" style="color:#f9fafb;">{pb_disc_str}</div>
    <div class="sc-sub">flujos descontados</div>
  </div>

  <div class="snap-card">
    <div class="sc-label">O&amp;M año 1</div>
    <div class="sc-val" style="color:#6b7280;">${om_anual[0]:,.0f}</div>
    <div class="sc-sub">MXN (est. {om_pct_sidebar:.1f}% inv.)</div>
  </div>

</div>
""", unsafe_allow_html=True)

    # ── Gráficas en dos columnas — fila 1 ─────────────────────────────────
    fm_col1, fm_col2 = st.columns(2, gap="medium")

    with fm_col1:
        st.markdown('<div class="section-header">Flujos de efectivo anuales</div>',
                    unsafe_allow_html=True)
        fig_cf = go.Figure()
        fig_cf.add_trace(go.Bar(
            x=years, y=flujo_neto,
            name="Flujo neto nominal",
            marker_color=AMBER, opacity=0.9,
            hovertemplate="<b>Año %{x}</b><br>Flujo neto: $%{y:,.0f} MXN<extra></extra>",
        ))
        fig_cf.add_trace(go.Bar(
            x=years, y=flujo_desc,
            name="Flujo descontado",
            marker_color=TEAL, opacity=0.85,
            hovertemplate="<b>Año %{x}</b><br>Flujo desc.: $%{y:,.0f} MXN<extra></extra>",
        ))
        fig_cf.add_trace(go.Scatter(
            x=years, y=om_anual,
            name="O&M anual",
            mode="lines+markers",
            line=dict(color=ROSE, width=2, dash="dot"),
            marker=dict(size=5, color=ROSE),
            hovertemplate="<b>Año %{x}</b><br>O&M: $%{y:,.0f} MXN<extra></extra>",
        ))
        lay_cf = copy.deepcopy(PLOT_LAYOUT)
        lay_cf.update({
            "height": 360, "barmode": "group",
            "yaxis": dict(title="MXN", gridcolor="#1e2230", tickformat=","),
            "xaxis": dict(title="Año", tickmode="linear", dtick=max(1, vida_util // 10), gridcolor="#1e2230"),
            "legend": dict(orientation="h", y=1.12, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
            "margin": dict(l=10, r=10, t=50, b=40),
            "hovermode": "x unified",
        })
        fig_cf.update_layout(**lay_cf)
        st.plotly_chart(fig_cf, use_container_width=True)

    with fm_col2:
        st.markdown('<div class="section-header">VPN acumulado y payback</div>',
                    unsafe_allow_html=True)
        fig_vpn = go.Figure()
        fig_vpn.add_trace(go.Scatter(
            x=[0] + years, y=[-inversion_mxn] + acum_desc,
            name="VPN acumulado (desc.)",
            mode="lines+markers",
            line=dict(color=TEAL, width=3),
            marker=dict(size=6, color=TEAL),
            fill="tozeroy",
            fillcolor="rgba(20,184,166,0.08)",
            hovertemplate="<b>Año %{x}</b><br>VPN acum.: $%{y:,.0f} MXN<extra></extra>",
        ))
        fig_vpn.add_trace(go.Scatter(
            x=[0] + years, y=[-inversion_mxn] + acum_nominal,
            name="Acum. nominal",
            mode="lines",
            line=dict(color=AMBER, width=2, dash="dash"),
            hovertemplate="<b>Año %{x}</b><br>Acum. nominal: $%{y:,.0f} MXN<extra></extra>",
        ))
        fig_vpn.add_hline(y=0, line_color="#475569", line_dash="solid", line_width=1)
        if pb_disc:
            fig_vpn.add_vline(x=pb_disc, line_color=TEAL, line_dash="dot", line_width=1.5,
                              annotation_text=f"PB desc. {pb_disc:.1f}a",
                              annotation_font=dict(color=TEAL, size=10))
        if pb_simple and pb_simple != pb_disc:
            fig_vpn.add_vline(x=pb_simple, line_color=AMBER, line_dash="dot", line_width=1.5,
                              annotation_text=f"PB simple {pb_simple:.1f}a",
                              annotation_font=dict(color=AMBER, size=10),
                              annotation_position="bottom right")
        lay_vpn = copy.deepcopy(PLOT_LAYOUT)
        lay_vpn.update({
            "height": 320,
            "yaxis": dict(title="MXN acumulados", gridcolor="#1e2230", tickformat=","),
            "xaxis": dict(title="Año", tickmode="linear", dtick=max(1, vida_util // 10), gridcolor="#1e2230"),
            "legend": dict(orientation="h", y=1.12, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
            "margin": dict(l=10, r=10, t=50, b=40),
            "hovermode": "x unified",
        })
        fig_vpn.update_layout(**lay_vpn)
        st.plotly_chart(fig_vpn, use_container_width=True)

    # ── Gráfica sensibilidad + tabla — fila 2 ─────────────────────────────
    fm_col3, fm_col4 = st.columns([1, 1.4], gap="medium")

    with fm_col3:
        st.markdown('<div class="section-header">Sensibilidad VPN vs WACC</div>',
                    unsafe_allow_html=True)
        tasas_sens = [i * 0.5 for i in range(0, 61)]
        vpn_sens   = []
        for t_s in tasas_sens:
            r_s = t_s / 100
            fd_s = [flujo_neto[i] / (1 + r_s) ** years[i] for i in range(len(years))]
            vpn_sens.append(-inversion_mxn + sum(fd_s))

        fig_sens = go.Figure()
        fig_sens.add_trace(go.Scatter(
            x=tasas_sens, y=vpn_sens,
            mode="lines", name="VPN",
            line=dict(color=TEAL, width=3),
            fill="tozeroy",
            fillcolor="rgba(20,184,166,0.10)",
            hovertemplate="<b>WACC %{x:.1f}%</b><br>VPN: $%{y:,.0f} MXN<extra></extra>",
        ))
        fig_sens.add_vline(x=discount_rate, line_color=AMBER, line_dash="dash", line_width=2,
                           annotation_text=f"WACC {discount_rate}%",
                           annotation_font=dict(color=AMBER, size=10))
        if tir is not None:
            fig_sens.add_vline(x=tir, line_color=ROSE, line_dash="dot", line_width=2,
                               annotation_text=f"TIR {tir:.1f}%",
                               annotation_font=dict(color=ROSE, size=10),
                               annotation_position="top right")
        fig_sens.add_hline(y=0, line_color="#475569", line_width=1)
        lay_sens = copy.deepcopy(PLOT_LAYOUT)
        lay_sens.update({
            "height": 300,
            "yaxis": dict(title="VPN (MXN)", gridcolor="#1e2230", tickformat=","),
            "xaxis": dict(title="WACC (%)", gridcolor="#1e2230",
                          range=[0, max(32, (tir or 0) + 5)]),
            "legend": dict(orientation="h", y=1.12, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
            "margin": dict(l=10, r=10, t=50, b=40),
        })
        fig_sens.update_layout(**lay_sens)
        st.plotly_chart(fig_sens, use_container_width=True)

    with fm_col4:
        st.markdown('<div class="section-header">Tabla financiera año a año</div>',
                    unsafe_allow_html=True)
        tabla_fin = pd.DataFrame({
            "Año":               years,
            "Gen. (MWh)":        [f"{g/1000:,.2f}" for g in gen_proj],
            "Tarifa ($/kWh)":    [f"${t:.3f}" for t in tarifas_y],
            "Ingreso (MXN)":     [f"${v:,.0f}" for v in flujo_nominal],
            "O&M (MXN)":         [f"${v:,.0f}" for v in om_anual],
            "Flujo neto (MXN)":  [f"${v:,.0f}" for v in flujo_neto],
            "Flujo desc. (MXN)": [f"${v:,.0f}" for v in flujo_desc],
            "VPN acum. (MXN)":   [f"${v:,.0f}" for v in acum_desc],
        })
        st.dataframe(tabla_fin, use_container_width=True, hide_index=True)

    # ── Totales — ancho completo ───────────────────────────────────────────
    st.markdown(f"""
<div style="display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin-top:12px;">
  <div class="snap-card">
    <div class="sc-label">Ingreso bruto total</div>
    <div class="sc-val" style="color:#f9fafb;">${sum(flujo_nominal):,.0f}</div>
    <div class="sc-sub">MXN nominales</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">O&amp;M total (nominal)</div>
    <div class="sc-val" style="color:#6b7280;">${sum(om_anual):,.0f}</div>
    <div class="sc-sub">MXN nominales</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Flujo neto total</div>
    <div class="sc-val" style="color:#f9fafb;">${sum(flujo_neto):,.0f}</div>
    <div class="sc-sub">MXN nominales</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">VPN final</div>
    <div class="sc-val" style="color:{'#4ade80' if vpn>0 else '#f87171'};">${vpn:,.0f}</div>
    <div class="sc-sub">MXN</div>
  </div>
</div>
""", unsafe_allow_html=True)

    # Criterio de decisión
    if tir is None:
        st.warning("⚠️ No se pudo calcular la TIR (flujo sin solución en el rango analizado).")
    elif vpn > 0 and tir > discount_rate + 3:
        st.success("🟢 **Proyecto muy atractivo** — VPN positivo y TIR supera significativamente el costo de capital.")
    elif vpn > 0 and tir > discount_rate:
        st.success("🟡 **Proyecto atractivo** — VPN positivo y TIR supera el costo de capital.")
    else:
        st.error("🔴 **Proyecto poco atractivo** — VPN negativo o TIR por debajo del costo de capital.")

            # ── Exportar TOR ───────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Exportar TOR</div>', unsafe_allow_html=True)

    # Llamada corregida y simplificada a build_tor_text
    tor_text = build_tor_text(
        "Proyecto Solar",
        "—",
        proj_loc,
        "",
        panel_wp,
        panel_eff_declared,
        panel_largo_mm,
        panel_ancho_mm,
        panel_peso_kg,
        panel_area,
        n_panels,
        kwp,
        pr_pct,
        irr_vals,
        monthly_gen,
        annual_gen,
        p50,
        p90,
        co2_saved,
        inversion,
        ahorro1,
        payback,
        gen_por_anio,
    )

    ex1, ex2 = st.columns(2)
    with ex1:
        st.download_button(
            "⬇️ Descargar TOR (.txt)",
            data=tor_text.encode("utf-8"),
            file_name=f"TOR_Solar_{proj_loc[:20].replace(' ','_')}.txt",
            mime="text/plain",
            use_container_width=True)
    with ex2:
        pdf_sizing_bytes = build_pdf_sizing(
            proj_loc=proj_loc, lat=lat, lon=lon,
            panel_wp=panel_wp, panel_eff_declared=panel_eff_declared,
            panel_largo_mm=panel_largo_mm, panel_ancho_mm=panel_ancho_mm,
            panel_peso_kg=panel_peso_kg, panel_area=panel_area,
            n_panels=n_panels, kwp=kwp, pr_pct=pr_pct,
            irr_vals=irr_vals, monthly_gen=monthly_gen, annual_gen=annual_gen,
            p50=p50, p90=p90, co2_saved=co2_saved,
            inversion_usd=float(inversion), usd_to_mxn=usd_to_mxn,
            ahorro1=ahorro1, payback=payback,
            vpn=vpn, tir=tir, lcoe=lcoe, pb_disc=pb_disc,
            tarifa_efectiva=tarifa_efectiva, inflation=inflation,
            discount_rate=discount_rate, vida_util=vida_util,
            om_pct=om_pct_sidebar,
            sizing_mode_label="Por área" if uso_area else "Por recibo CFE",
        )
        st.download_button(
            "📄 Exportar PDF — Sizing",
            data=pdf_sizing_bytes,
            file_name=f"Sizing_Solar_{proj_loc[:20].replace(' ','_')}.pdf",
            mime="application/pdf",
            use_container_width=True,
            type="primary",
        )


# ═════════════════════════════════════════════════════════════════════════════
# TAB 3 — PPA · VENTA AL CLIENTE
# ═════════════════════════════════════════════════════════════════════════════
with tab3:

    st.markdown("""
    <div class="info-box">
    💡 <b>Power Purchase Agreement (PPA)</b> — El cliente paga por la energía generada a una
    tarifa fija acordada ($/kWh), en lugar de comprar la instalación. Tú (o el financiador)
    eres dueño del sistema durante el plazo del contrato. Evalúa distintos plazos y
    encuentra el precio PPA que hace viable el proyecto.
    </div>
    """, unsafe_allow_html=True)

    ppa_col1, ppa_col2, ppa_col3 = st.columns([1.1, 1.1, 1.8], gap="large")

    with ppa_col1:
        st.markdown('<div class="section-header">Sistema base</div>', unsafe_allow_html=True)

        # Valores heredados de Turnkey Solar (Tab 1)
        _kwp_turnkey = max(1.0, round(float(kwp), 1))
        _inv_turnkey = max(1000.0, round(float(kwp) * float(costo_kwp), 0))

        st.markdown(f"""
<div class="info-box" style="margin-bottom:0.5rem;">
  📐 <b>Valores importados de Turnkey Solar:</b>
  {_kwp_turnkey:.1f} kWp &nbsp;·&nbsp; ${_inv_turnkey:,.0f} USD inversión.
  Puedes editarlos abajo si deseas simular un escenario diferente.
</div>
""", unsafe_allow_html=True)

        ppa_kwp = st.number_input("Capacidad (kWp)", 1.0, 50000.0,
                                   _kwp_turnkey, 1.0, key="ppa_kwp")

        # ── Generación base: siempre P90 cuando está disponible ──────────────
        _p50_val = max(100.0, round(float(annual_gen), 0))
        _p90_val = max(100.0, round(float(p90), 0)) if p90 else None
        _has_p90 = _p90_val is not None

        if _has_p90:
            st.markdown(
                f'<div class="nasa-box">🔬 <b>Base de generación: P90</b> — '
                f'{_p90_val:,.0f} kWh/año · estándar de la industria para contratos PPA. '
                f'El sistema supera este valor el 90% de los años históricos (NASA POWER {NASA_START}–{NASA_END}). '
                f'P50 = {_p50_val:,.0f} kWh/año (mediana).</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div class="warn-box">⚠️ <b>P90 no disponible — usando P50 como base.</b> '
                'Carga datos NASA POWER desde el sidebar para usar el P90 (recomendado para PPA).</div>',
                unsafe_allow_html=True,
            )

        _gen_default = _p90_val if _has_p90 else _p50_val
        _gen_label   = f"Generación año 1 — {'P90' if _has_p90 else 'P50'} (kWh/año)"

        ppa_gen_anual = st.number_input(_gen_label, 100.0, 50_000_000.0,
                                         _gen_default, 100.0, key="ppa_gen")
        ppa_inversion_usd = st.number_input("Inversión total (USD)", 1000.0, 50_000_000.0,
                                             _inv_turnkey, 100.0, key="ppa_inv")
        st.caption(f"≈ ${ppa_inversion_usd * usd_to_mxn:,.0f} MXN al tipo de cambio configurado")

        st.markdown('<div class="section-header">Parámetros técnicos</div>', unsafe_allow_html=True)
        ppa_degradacion  = st.slider("Degradación anual (%)", 0.0, 1.5, 0.5, 0.05, key="ppa_deg")
        ppa_om_pct       = st.slider("O&M anual (% inv. MXN)", 0.3, 2.5, 1.0, 0.1, key="ppa_om")
        ppa_seguros_pct  = st.slider("Seguros / otros (% inv. MXN)", 0.0, 1.0, 0.3, 0.05, key="ppa_seg")

    with ppa_col2:
        st.markdown('<div class="section-header">Condiciones financieras</div>', unsafe_allow_html=True)
        ppa_wacc             = st.slider("WACC (%)", 5.0, 30.0, 15.0, 0.5, key="ppa_wacc")
        ppa_inflacion_tarifa = st.slider("Escalador PPA anual (%)", 0.0, 8.0, 3.5, 0.5, key="ppa_esc",
                                          help="Incremento anual pactado en el precio PPA")
        ppa_inflacion_om     = st.slider("Inflación O&M anual (%)", 0.0, 8.0, 4.0, 0.5, key="ppa_inf_om")

        ppa_financiamiento = st.checkbox("¿Incluir financiamiento?", value=False, key="ppa_fin_chk")
        if ppa_financiamiento:
            ppa_equity_pct  = st.slider("Capital propio (%)", 10, 100, 30, 5, key="ppa_eq")
            ppa_tasa_deuda  = st.slider("Tasa deuda anual (%)", 5.0, 25.0, 12.0, 0.5, key="ppa_debt_r")
            ppa_plazo_deuda = st.slider("Plazo deuda (años)", 3, 20, 10, 1, key="ppa_debt_p")
        else:
            ppa_equity_pct  = 100
            ppa_tasa_deuda  = 0.0
            ppa_plazo_deuda = 0

        st.markdown('<div class="section-header">Tarifa CFE del cliente</div>', unsafe_allow_html=True)
        ppa_tarifa_cliente = st.number_input("Tarifa actual (MXN/kWh)", 0.5, 15.0,
                                              max(0.5, round(float(tarifa_efectiva), 2)),
                                              0.05, key="ppa_tar")
        ppa_inflacion_cfe  = st.slider("Inflación CFE anual (%)", 0.0, 12.0, 6.0, 0.5, key="ppa_inf_cfe")

    with ppa_col3:
        st.markdown('<div class="section-header">Precio PPA a evaluar</div>', unsafe_allow_html=True)
        ppa_precio_manual = st.number_input(
            "Precio PPA año 1 (MXN/kWh)", 0.50, 10.0, 1.80, 0.05, key="ppa_price",
            help="Ajusta este valor hasta encontrar el precio óptimo para tu cliente")

        ppa_plazos = st.multiselect(
            "Plazos a comparar (años)",
            options=[3, 5, 10, 15, 20, 25],
            default=[10, 15, 20, 25],
            key="ppa_plazos")
        if not ppa_plazos:
            ppa_plazos = [10, 15, 20, 25]
        ppa_plazos = sorted(ppa_plazos)

        st.markdown('<div class="section-header">Plazo objetivo</div>', unsafe_allow_html=True)
        ppa_plazo_minimo = st.selectbox("Plazo para análisis detallado", ppa_plazos, key="ppa_pmin_plazo")

    # ── Calcular todos los plazos — usando funciones cacheadas globales ──────
    ppa_cache_kwargs = dict(
        gen1=ppa_gen_anual, inv_usd=ppa_inversion_usd,
        wacc_pct=ppa_wacc, esc_ppa=ppa_inflacion_tarifa,
        deg=ppa_degradacion, om_pct=ppa_om_pct,
        inf_om=ppa_inflacion_om, seg_pct=ppa_seguros_pct,
        usd_mx=usd_to_mxn, equity_pct=ppa_equity_pct,
        tasa_deuda=ppa_tasa_deuda, plazo_deuda=ppa_plazo_deuda,
        con_fin=ppa_financiamiento,
        vida_util_total=vida_util)

    resultados = {}
    for pl in ppa_plazos:
        res = dict(calc_ppa_result(precio_ppa=ppa_precio_manual, plazo=pl, **ppa_cache_kwargs))
        res["pm"] = calc_precio_minimo(plazo=pl, vida_util_total=vida_util, **{k: v for k, v in ppa_cache_kwargs.items() if k != "vida_util_total"})
        resultados[pl] = res

    descuento_vs_cfe = ((ppa_precio_manual / ppa_tarifa_cliente) - 1) * 100
    pm_obj = resultados[ppa_plazo_minimo]["pm"]
    viable = pm_obj is not None and ppa_precio_manual >= pm_obj
    color_viable = "#4ade80" if viable else "#f87171"
    pm_str = f"${pm_obj:.4f}/kWh" if pm_obj else "No viable en este plazo"

    # ── Hero PPA ─────────────────────────────────────────────────────────────
    ro_obj = resultados[ppa_plazo_minimo]
    val_res = ro_obj.get("valor_residual", 0.0)
    val_res_str = f"${val_res:,.0f} MXN" if val_res > 0 else ("Contrato = vida útil" if vida_util <= ppa_plazo_minimo else "—")
    _anios_rest = max(0, vida_util - ppa_plazo_minimo)
    _g_ing = round((ppa_inflacion_tarifa - ppa_degradacion), 2)
    _nota_rescate = (
        f"Gordon generalizado · suma finita exacta · {_anios_rest} años restantes · "
        f"g_ingreso = {_g_ing:+.2f}%/año (escalador {ppa_inflacion_tarifa:.1f}% − degradación {ppa_degradacion:.2f}%) · "
        f"descontado a WACC {ppa_wacc:.1f}% desde t={ppa_plazo_minimo}"
        if _anios_rest > 0 else "Contrato cubre toda la vida útil del sistema"
    )
    st.markdown(f"""
<div class="tor-hero" style="margin-top:1rem;">
  <div class="th-project">📄 ANÁLISIS PPA · Plazo objetivo {ppa_plazo_minimo} años</div>
  <div class="th-meta">
    Precio evaluado: <b style="color:#f59e0b">${ppa_precio_manual:.4f}/kWh</b>
    &nbsp;·&nbsp; Ahorro cliente vs CFE hoy: <b style="color:#14b8a6">{descuento_vs_cfe:+.1f}%</b>
    &nbsp;·&nbsp; <b style="color:{color_viable}">{'✅ Precio viable' if viable else '⚠️ Por debajo del mínimo'}</b>
  </div>
  <div class="th-grid" style="grid-template-columns:repeat(4,1fr);">
    <div class="th-item">
      <span class="th-label">Precio mínimo viable</span>
      <span class="th-val" style="color:{color_viable};font-size:15px;">{pm_str}</span>
      <span class="th-unit">a {ppa_plazo_minimo} años · VPN = 0</span>
    </div>
    <div class="th-item">
      <span class="th-label">Precio PPA evaluado</span>
      <span class="th-val">${ppa_precio_manual:.4f}</span>
      <span class="th-unit">MXN/kWh año 1</span>
    </div>
    <div class="th-item">
      <span class="th-label">Tarifa CFE cliente</span>
      <span class="th-val">${ppa_tarifa_cliente:.2f}</span>
      <span class="th-unit">MXN/kWh actual</span>
    </div>
    <div class="th-item">
      <span class="th-label">Inversión total</span>
      <span class="th-val">${ppa_inversion_usd:,.0f}</span>
      <span class="th-unit">USD · ${ppa_inversion_usd*usd_to_mxn:,.0f} MXN</span>
    </div>
    <div class="th-item" style="margin-top:10px;">
      <span class="th-label">VALOR DE RESCATE</span>
      <span class="th-val" style="color:#14b8a6;font-size:14px;">{val_res_str}</span>
      <span class="th-unit">{_nota_rescate}</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Tarjetas comparativas por plazo ───────────────────────────────────────
    st.markdown('<div class="section-header">Comparativo de plazos</div>', unsafe_allow_html=True)
    cols_pl = st.columns(len(ppa_plazos))
    for idx, pl in enumerate(ppa_plazos):
        r   = resultados[pl]
        vc  = "#4ade80" if r["vpn"]>0 else "#f87171"
        tis = f"{r['tir']:.1f}%" if r["tir"] is not None else "N/A"
        pbs      = f"{r['pb']} años" if r["pb"] else f">{pl}a"
        pbs_disc = f"{r['pb_disc']} años" if r.get("pb_disc") else f">{pl}a"
        pmc = "#4ade80" if r["pm"] and ppa_precio_manual>=r["pm"] else "#f87171"
        pms = f"${r['pm']:.4f}" if r["pm"] else "No viable"
        with cols_pl[idx]:
            st.markdown(f"""
<div class="snap-card" style="min-height:240px;padding:18px 12px;">
  <div class="sc-label" style="font-size:14px;font-weight:700;color:#f59e0b;margin-bottom:12px;">{pl} AÑOS</div>
  <div style="width:100%;text-align:left;display:flex;flex-direction:column;gap:8px;">
    <div><div class="sc-label">VPN</div>
         <div class="sc-val" style="color:{vc};font-size:13px;">${r['vpn']:,.0f}</div></div>
    <div><div class="sc-label">TIR equity</div>
         <div class="sc-val" style="color:#22d3ee;font-size:13px;">{tis}</div></div>
    <div><div class="sc-label">Payback simple</div>
         <div class="sc-val" style="color:#f9fafb;font-size:13px;">{pbs}</div></div>
    <div><div class="sc-label">Payback descontado</div>
         <div class="sc-val" style="color:#9ca3af;font-size:13px;">{pbs_disc}</div></div>
    <div><div class="sc-label">Precio mínimo viable</div>
         <div class="sc-val" style="color:{pmc};font-size:13px;">{pms}/kWh</div></div>
    <div><div class="sc-label">Valor de rescate</div>
         <div class="sc-val" style="color:#14b8a6;font-size:13px;">${r['valor_residual']:,.0f}</div></div>
  </div>
</div>""", unsafe_allow_html=True)

    # Tabla resumen
    st.markdown("<div style='margin-top:1.2rem'></div>", unsafe_allow_html=True)
    tabla_ppa = []
    for pl in ppa_plazos:
        r = resultados[pl]
        tabla_ppa.append({
            "Plazo":            f"{pl} años",
            "Precio evaluado":  f"${ppa_precio_manual:.4f}/kWh",
            "Precio mínimo":    f"${r['pm']:.4f}/kWh" if r["pm"] else "No viable",
            "VPN (MXN)":        f"${r['vpn']:,.0f}",
            "TIR equity":       f"{r['tir']:.1f}%" if r["tir"] else "N/A",
            "Payback simple":   f"{r['pb']} años" if r["pb"] else f">{pl}a",
            "Payback desc.":    f"{r['pb_disc']} años" if r.get("pb_disc") else f">{pl}a",
            "Ingreso total":    f"${r['ing_total']:,.0f}",
            "Valor de rescate": f"${r.get('valor_residual', 0):,.0f}",
        })
    st.dataframe(pd.DataFrame(tabla_ppa), use_container_width=True, hide_index=True)

    # ── Gráficas ──────────────────────────────────────────────────────────────
    gc1, gc2 = st.columns(2, gap="large")

    with gc1:
        st.markdown('<div class="section-header">VPN por plazo</div>', unsafe_allow_html=True)
        vpn_vals = [resultados[pl]["vpn"] for pl in ppa_plazos]
        fig_vp = go.Figure(go.Bar(
            x=[f"{pl}a" for pl in ppa_plazos], y=vpn_vals,
            marker_color=[TEAL if v>=0 else ROSE for v in vpn_vals],
            text=[f"${v/1e6:.2f}M" for v in vpn_vals],
            textposition="outside", textfont=dict(size=12, family="DM Mono"),
            hovertemplate="<b>%{x}</b><br>VPN: $%{y:,.0f} MXN<extra></extra>"))
        fig_vp.add_hline(y=0, line_color="#6b7280", line_width=1.5)
        lyt_vp = copy.deepcopy(PLOT_LAYOUT)
        lyt_vp.update({"height":300,
                       "yaxis": dict(title="VPN (MXN)", gridcolor="#2a2d3a", tickformat=","),
                       "xaxis": dict(title="Plazo"),
                       "margin": dict(l=20,r=20,t=30,b=40)})
        fig_vp.update_layout(**lyt_vp)
        st.plotly_chart(fig_vp, use_container_width=True)

    with gc2:
        st.markdown('<div class="section-header">Precio mínimo viable por plazo</div>', unsafe_allow_html=True)
        pm_vals = [resultados[pl]["pm"] for pl in ppa_plazos]
        fig_pm = go.Figure()
        fig_pm.add_trace(go.Scatter(
            x=[f"{pl}a" for pl in ppa_plazos],
            y=[v if v else None for v in pm_vals],
            mode="lines+markers+text",
            line=dict(color=AMBER, width=3),
            marker=dict(size=10, color=AMBER),
            text=[f"${v:.4f}" if v else "N/V" for v in pm_vals],
            textposition="top center", textfont=dict(size=11, family="DM Mono"),
            name="Precio mínimo",
            hovertemplate="<b>%{x}</b><br>Precio mín: $%{y:.4f}/kWh<extra></extra>"))
        fig_pm.add_hline(y=ppa_precio_manual, line_color=TEAL, line_dash="dash", line_width=2,
                         annotation_text=f"Precio evaluado ${ppa_precio_manual:.4f}",
                         annotation_font=dict(color=TEAL, size=11))
        fig_pm.add_hline(y=ppa_tarifa_cliente, line_color=ROSE, line_dash="dot", line_width=1.5,
                         annotation_text=f"CFE hoy ${ppa_tarifa_cliente:.2f}",
                         annotation_font=dict(color=ROSE, size=11), annotation_position="bottom right")
        lyt_pm = copy.deepcopy(PLOT_LAYOUT)
        lyt_pm.update({"height":300,
                       "yaxis": dict(title="MXN/kWh", gridcolor="#2a2d3a", tickformat=".4f"),
                       "xaxis": dict(title="Plazo"),
                       "margin": dict(l=20,r=20,t=30,b=40),
                       "legend": dict(orientation="h",y=1.1,x=0.5,xanchor="center",bgcolor="rgba(0,0,0,0)")})
        fig_pm.update_layout(**lyt_pm)
        st.plotly_chart(fig_pm, use_container_width=True)

    # ── Flujos anuales plazo objetivo ─────────────────────────────────────────
    st.markdown(f'<div class="section-header">Flujos anuales — {ppa_plazo_minimo} años</div>',
                unsafe_allow_html=True)
    ro = resultados[ppa_plazo_minimo]
    fig_fl = go.Figure()
    fig_fl.add_trace(go.Bar(x=ro["years"], y=ro["ing_y"], name="Ingreso PPA",
        marker_color=AMBER, opacity=0.9,
        hovertemplate="<b>Año %{x}</b><br>Ingreso: $%{y:,.0f} MXN<extra></extra>"))
    costos_y = [ro["om_y"][i]+ro["seg_y"][i] for i in range(ppa_plazo_minimo)]
    fig_fl.add_trace(go.Bar(x=ro["years"], y=costos_y, name="O&M + Seguros",
        marker_color="#374151",
        hovertemplate="<b>Año %{x}</b><br>Costos: $%{y:,.0f} MXN<extra></extra>"))
    if any(d>0 for d in ro["deu_y"]):
        fig_fl.add_trace(go.Bar(x=ro["years"], y=ro["deu_y"], name="Servicio deuda",
            marker_color=ROSE, opacity=0.8,
            hovertemplate="<b>Año %{x}</b><br>Deuda: $%{y:,.0f} MXN<extra></extra>"))
    fig_fl.add_trace(go.Scatter(x=ro["years"], y=ro["fn_y"], name="Flujo neto",
        mode="lines+markers", line=dict(color=TEAL,width=2.5), marker=dict(size=6,color=TEAL),
        hovertemplate="<b>Año %{x}</b><br>Flujo neto: $%{y:,.0f} MXN<extra></extra>"))
    lyt_fl = copy.deepcopy(PLOT_LAYOUT)
    lyt_fl.update({"height":330,"barmode":"stack",
                   "yaxis": dict(title="MXN",gridcolor="#2a2d3a",tickformat=","),
                   "xaxis": dict(title="Año",tickmode="linear",dtick=max(1,ppa_plazo_minimo//10)),
                   "legend": dict(orientation="h",y=1.12,x=0.5,xanchor="center",bgcolor="rgba(0,0,0,0)"),
                   "margin": dict(l=20,r=20,t=50,b=40),"hovermode":"x unified"})
    fig_fl.update_layout(**lyt_fl)
    st.plotly_chart(fig_fl, use_container_width=True)

    # ── Perspectiva del cliente ───────────────────────────────────────────────
    st.markdown('<div class="section-header">Perspectiva del cliente · Ahorro vs CFE</div>',
                unsafe_allow_html=True)
    gen_cl   = ro["gen_y"]
    prec_cl  = ro["prec_y"]
    cfe_y    = [ppa_tarifa_cliente*(1+ppa_inflacion_cfe/100)**i for i in range(ppa_plazo_minimo)]
    pago_ppa = [gen_cl[i]*prec_cl[i] for i in range(ppa_plazo_minimo)]
    pago_cfe = [gen_cl[i]*cfe_y[i]   for i in range(ppa_plazo_minimo)]
    ahorro_y = [pago_cfe[i]-pago_ppa[i] for i in range(ppa_plazo_minimo)]

    fig_cl = go.Figure()
    fig_cl.add_trace(go.Bar(x=ro["years"], y=pago_cfe, name="Lo que pagaría a CFE",
        marker_color="#374151",
        hovertemplate="<b>Año %{x}</b><br>CFE: $%{y:,.0f} MXN<extra></extra>"))
    fig_cl.add_trace(go.Bar(x=ro["years"], y=pago_ppa, name="Pago PPA",
        marker_color=AMBER, opacity=0.9,
        hovertemplate="<b>Año %{x}</b><br>PPA: $%{y:,.0f} MXN<extra></extra>"))
    fig_cl.add_trace(go.Scatter(x=ro["years"], y=ahorro_y, name="Ahorro anual",
        mode="lines+markers", line=dict(color=TEAL,width=2.5), marker=dict(size=6,color=TEAL),
        hovertemplate="<b>Año %{x}</b><br>Ahorro: $%{y:,.0f} MXN<extra></extra>"))
    lyt_cl = copy.deepcopy(PLOT_LAYOUT)
    lyt_cl.update({"height":310,"barmode":"overlay",
                   "yaxis": dict(title="MXN/año",gridcolor="#2a2d3a",tickformat=","),
                   "xaxis": dict(title="Año",tickmode="linear",dtick=max(1,ppa_plazo_minimo//10)),
                   "legend": dict(orientation="h",y=1.12,x=0.5,xanchor="center",bgcolor="rgba(0,0,0,0)"),
                   "margin": dict(l=20,r=20,t=50,b=40),"hovermode":"x unified"})
    fig_cl.update_layout(**lyt_cl)
    st.plotly_chart(fig_cl, use_container_width=True)

    # KPIs cliente
    ahorro_total = sum(ahorro_y)
    cfe_final    = cfe_y[-1]
    ppa_final    = prec_cl[-1]
    st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-top:8px;">
  <div class="snap-card">
    <div class="sc-label">Ahorro total cliente</div>
    <div class="sc-val" style="color:#4ade80;">${ahorro_total:,.0f}</div>
    <div class="sc-sub">MXN en {ppa_plazo_minimo} años</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Ahorro año 1</div>
    <div class="sc-val" style="color:#4ade80;">${ahorro_y[0]:,.0f}</div>
    <div class="sc-sub">MXN</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Descuento vs CFE hoy</div>
    <div class="sc-val" style="color:#f59e0b;">{descuento_vs_cfe:+.1f}%</div>
    <div class="sc-sub">precio PPA vs tarifa actual</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">CFE año {ppa_plazo_minimo} (proyectada)</div>
    <div class="sc-val" style="color:#f87171;">${cfe_final:.4f}</div>
    <div class="sc-sub">vs ${ppa_final:.4f} PPA ese año</div>
  </div>
</div>
""", unsafe_allow_html=True)

    # ── Tabla anual detallada — con servicio de deuda desglosado ─────────────
    st.markdown(f'<div class="section-header">Tabla año a año · {ppa_plazo_minimo} años</div>',
                unsafe_allow_html=True)

    # Desglose capital/interés por año (amortización francesa)
    _eq_mxn   = ro.get("equity_mxn", ro.get("inv_mxn", ppa_inversion_usd * usd_to_mxn))
    _inv_mxn  = ro.get("inv_mxn", ppa_inversion_usd * usd_to_mxn)
    _deu_ini  = _inv_mxn - _eq_mxn   # deuda inicial

    interes_y = []
    capital_y = []
    saldo_y   = []
    saldo     = _deu_ini

    for i, y in enumerate(ro["years"]):
        serv = ro["deu_y"][i]
        if serv > 0 and saldo > 0:
            int_y = saldo * (ppa_tasa_deuda / 100)
            cap_y = min(serv - int_y, saldo)
            cap_y = max(cap_y, 0.0)
            saldo = max(saldo - cap_y, 0.0)
        else:
            int_y = 0.0
            cap_y = 0.0
        interes_y.append(int_y)
        capital_y.append(cap_y)
        saldo_y.append(saldo)

    _tiene_deuda = any(d > 0 for d in ro["deu_y"])

    tabla_dict = {
        "Año":                ro["years"],
        "Generación (MWh)":  [f"{g/1000:.2f}" for g in gen_cl],
        "Precio PPA ($/kWh)":[f"${p:.4f}" for p in prec_cl],
        "Ingreso PPA (MXN)": [f"${v:,.0f}" for v in pago_ppa],
        "O&M + Seg. (MXN)":  [f"${ro['om_y'][i]+ro['seg_y'][i]:,.0f}" for i in range(ppa_plazo_minimo)],
    }
    if _tiene_deuda:
        tabla_dict["Interés (MXN)"]  = [f"${v:,.0f}" for v in interes_y]
        tabla_dict["Capital (MXN)"]  = [f"${v:,.0f}" for v in capital_y]
        tabla_dict["Serv. deuda (MXN)"] = [f"${v:,.0f}" for v in ro["deu_y"]]
        tabla_dict["Saldo deuda (MXN)"] = [f"${v:,.0f}" for v in saldo_y]
    tabla_dict["Flujo neto (MXN)"]   = [f"${v:,.0f}" for v in ro["fn_y"]]
    tabla_dict["CFE equiv. ($/kWh)"] = [f"${c:.4f}" for c in cfe_y]
    tabla_dict["Ahorro cliente"]     = [f"${v:,.0f}" for v in ahorro_y]

    st.dataframe(pd.DataFrame(tabla_dict), use_container_width=True, hide_index=True)

    if _tiene_deuda:
        st.caption(
            f"Servicio de deuda: anuidad fija calculada sobre ${_deu_ini:,.0f} MXN "
            f"a {ppa_tasa_deuda:.1f}% anual en {ppa_plazo_deuda} años. "
            f"Interés = saldo × tasa. Capital = servicio − interés. "
            f"Saldo = saldo anterior − capital amortizado."
        )

    # ── Exportar PDF PPA ──────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Exportar análisis PPA</div>',
                unsafe_allow_html=True)
    pdf_ppa_bytes = build_pdf_ppa(
        proj_loc=proj_loc,
        ppa_kwp=ppa_kwp, ppa_gen_anual=ppa_gen_anual,
        ppa_inversion_usd=ppa_inversion_usd, usd_to_mxn=usd_to_mxn,
        ppa_wacc=ppa_wacc, ppa_inflacion_tarifa=ppa_inflacion_tarifa,
        ppa_degradacion=ppa_degradacion, ppa_om_pct=ppa_om_pct,
        ppa_seguros_pct=ppa_seguros_pct, ppa_tarifa_cliente=ppa_tarifa_cliente,
        ppa_inflacion_cfe=ppa_inflacion_cfe,
        ppa_precio_manual=ppa_precio_manual,
        ppa_plazo_minimo=ppa_plazo_minimo, ppa_plazos=ppa_plazos,
        resultados=resultados, descuento_vs_cfe=descuento_vs_cfe,
        ro=ro,
        ahorro_total=ahorro_total, ahorro_y=ahorro_y,
        cfe_y=cfe_y, pago_ppa=pago_ppa, pago_cfe=pago_cfe,
        pm_obj=pm_obj, viable=viable,
    )
    st.download_button(
        "📄 Exportar PDF — Análisis PPA",
        data=pdf_ppa_bytes,
        file_name=f"PPA_Solar_{proj_loc[:20].replace(' ','_')}.pdf",
        mime="application/pdf",
        use_container_width=True,
        type="primary",
    )

# ═════════════════════════════════════════════════════════════════════════════
# TAB 4 — 🔋 BATERÍAS STANDALONE (Sizing + Modelación Financiera Rigurosa)
# ═════════════════════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════════════════════
# BESS — Constantes y motor de cálculo riguroso
# ═════════════════════════════════════════════════════════════════════════════

# Parámetros LFP de referencia (basado en datos de fabricantes: CATL, BYD, Pylontech)
LFP_PARAMS = {
    "dod_optimo":    0.90,    # DoD recomendado para maximizar vida útil
    "rte":           0.925,   # Round-trip efficiency AC-AC (incluye PCS)
    "c_rate_max":    1.0,     # C-rate máximo continuo (1C = carga/descarga en 1h)
    "eol_capacity":  0.80,    # Capacidad al final de vida (80% de nominal)
    # Curva ciclos-DoD (Wöhler simplificada, interpolación lineal entre puntos de datasheet)
    # Fuente: CATL LF280K datasheet + NREL Battery Lifetime Analysis Model
    "woehler": [
        (0.20, 12000),  # 20% DoD → 12,000 ciclos
        (0.40, 8000),
        (0.60, 6500),
        (0.80, 5000),
        (0.90, 4000),
        (1.00, 3000),   # 100% DoD → 3,000 ciclos
    ],
    # Costos de referencia 2024 (USD, instalado en México)
    "costo_energia_usd_kwh": 320,   # celdas + BMS + rack
    "costo_pcs_usd_kw":      120,   # Power Conversion System (inversor bidireccional)
    "costo_bos_pct":         0.20,  # BOS: cableado, protecciones, obra civil (% sobre celdas+PCS)
    "om_pct_anual":          0.015, # O&M anual (1.5% del CAPEX)
    "costo_reemplazo_pct":   0.65,  # Costo reemplazo de celdas (65% del CAPEX original de celdas)
    "degradacion_lineal_pct_año": 2.0,  # Pérdida de capacidad útil por año (para ahorro proyectado)
}

# Perfil horario típico por tipo de instalación (fracción de la demanda pico)
# Normalizado para que el área bajo la curva = consumo diario promedio
LOAD_PROFILES = {
    "Comercial (oficinas)": [
        0.20,0.18,0.17,0.17,0.18,0.22, 0.55,0.80,0.95,1.00,0.98,0.95,
        0.90,0.95,0.98,0.95,0.85,0.70, 0.55,0.45,0.38,0.32,0.27,0.22],
    "Industrial (2 turnos)": [
        0.25,0.22,0.20,0.20,0.22,0.30, 0.65,0.90,1.00,1.00,0.98,0.95,
        0.85,0.95,1.00,1.00,0.95,0.90, 0.80,0.70,0.50,0.40,0.32,0.28],
    "Industrial (3 turnos)": [
        0.85,0.82,0.80,0.80,0.82,0.85, 0.90,0.95,1.00,1.00,0.98,0.95,
        0.90,0.95,1.00,1.00,0.98,0.95, 0.92,0.90,0.88,0.87,0.86,0.85],
    "Residencial": [
        0.18,0.15,0.13,0.12,0.12,0.15, 0.35,0.55,0.45,0.38,0.35,0.40,
        0.50,0.45,0.40,0.42,0.55,0.85, 1.00,0.95,0.85,0.72,0.55,0.35],
    "Hotel / Hospitalario": [
        0.55,0.50,0.48,0.47,0.48,0.55, 0.70,0.88,0.95,1.00,0.98,0.95,
        0.92,0.90,0.88,0.88,0.90,0.95, 0.98,0.95,0.90,0.82,0.72,0.62],
}

# Horario tarifario CFE (DAC y MT): horas pico/intermedio/base típico zona norte
TARIFF_SCHEDULE = {
    "Punta":      list(range(18, 22)),   # 18:00–21:59 — Más caro
    "Intermedio": list(range(6, 18)),    # 06:00–17:59
    "Base":       list(range(0, 6)) + [22, 23],  # 00:00–05:59 y 22:00–23:59 — Más barato
}


@st.cache_data(show_spinner=False)
def woehler_cycles(dod: float, woehler_table: list) -> float:
    """
    Interpola linealmente la curva de Wöhler (ciclos vs DoD) del datasheet del fabricante.
    Devuelve los ciclos esperados hasta EoL (80% capacidad) al DoD dado.
    """
    pts = sorted(woehler_table, key=lambda x: x[0])
    if dod <= pts[0][0]:
        return float(pts[0][1])
    if dod >= pts[-1][0]:
        return float(pts[-1][1])
    for i in range(len(pts) - 1):
        d0, c0 = pts[i]
        d1, c1 = pts[i + 1]
        if d0 <= dod <= d1:
            t = (dod - d0) / (d1 - d0)
            return c0 + t * (c1 - c0)
    return float(pts[-1][1])


def simulate_bess_dispatch(
    load_profile_24h: tuple,   # kW hora a hora (24 valores, demanda real)
    peak_demand_kw: float,     # kW pico de la instalación
    daily_kwh: float,          # kWh consumo diario total
    bess_power_kw: float,      # kW potencia del PCS
    bess_energy_kwh: float,    # kWh capacidad útil (ya descontado DoD)
    rte: float,                # eficiencia round-trip
    caso_uso: str,             # "Peak shaving" | "Arbitraje" | "Respaldo" | "Solar+BESS"
    peak_shaving_target_kw: float = 0.0,  # kW objetivo de corte de punta
    solar_gen_24h: tuple = None,          # kW generación solar hora a hora (para híbrido)
    tariff_type: str = "Punta/Base",      # tipo de tarifa para arbitraje
) -> dict:
    """
    Simula el despacho horario de la batería en un día típico.
    Devuelve energía cargada, descargada, ciclos diarios, pico residual y ahorro diario.

    Estrategias:
    - Peak shaving: descarga cuando demanda > target, carga en horas valle
    - Arbitraje: carga en horas base, descarga en horas punta
    - Respaldo: reserva capacidad completa (no cicla en operación normal)
    - Solar+BESS: absorbe excedente solar, descarga cuando no hay solar
    """
    hours = list(range(24))
    # Escalar perfil normalizado a demanda real
    scale = peak_demand_kw
    demand = [p * scale for p in load_profile_24h]

    soc = bess_energy_kwh * 0.5   # Estado de carga inicial: 50%
    soc_min = 0.0
    soc_max = bess_energy_kwh

    charged_kwh   = 0.0
    discharged_kwh = 0.0
    peak_residual  = max(demand)
    soc_curve      = []
    dispatch_curve = []   # + = descarga, - = carga

    if caso_uso == "Respaldo / UPS":
        # En modo respaldo la batería NO cicla — está en reserva
        # Simulamos el escenario de outage: batería descarga a plena demanda
        autonomy_h = bess_energy_kwh / max(peak_demand_kw * rte, 0.1)
        return dict(
            charged_kwh=0.0, discharged_kwh=bess_energy_kwh,
            daily_cycles=1.0,  # 1 ciclo completo en evento de corte
            peak_residual=peak_demand_kw,
            autonomy_hours=round(autonomy_h, 2),
            soc_curve=[bess_energy_kwh] * 24,
            dispatch_curve=[0.0] * 24,
            demand_curve=demand,
        )

    for h in hours:
        d = demand[h]
        solar = float(solar_gen_24h[h]) if (solar_gen_24h and len(solar_gen_24h) > h) else 0.0
        net_load = d - solar   # carga neta después de solar (puede ser negativa = excedente)

        if caso_uso == "Peak shaving":
            target = peak_shaving_target_kw
            if d > target and soc > soc_min:
                # Descarga para cortar punta
                discharge = min(d - target, bess_power_kw, soc)
                soc -= discharge
                discharged_kwh += discharge * rte
                dispatch_curve.append(discharge)
                peak_residual = min(peak_residual, d - discharge) if h > 0 else d - discharge
            elif d < target * 0.6 and soc < soc_max:
                # Carga en horas valle (demanda < 60% del target)
                charge_room = min(bess_power_kw, soc_max - soc)
                charge = min(charge_room, (target * 0.6 - d))
                soc = min(soc + charge, soc_max)
                charged_kwh += charge
                dispatch_curve.append(-charge)
            else:
                dispatch_curve.append(0.0)

        elif caso_uso == "Arbitraje tarifario (valle→punta)":
            if h in TARIFF_SCHEDULE["Base"] and soc < soc_max:
                # Carga máxima en horario base (más barato)
                charge = min(bess_power_kw, soc_max - soc)
                soc = min(soc + charge, soc_max)
                charged_kwh += charge
                dispatch_curve.append(-charge)
            elif h in TARIFF_SCHEDULE["Punta"] and soc > soc_min:
                # Descarga máxima en horario punta (más caro)
                discharge = min(bess_power_kw, soc, d)
                soc = max(soc - discharge, soc_min)
                discharged_kwh += discharge * rte
                dispatch_curve.append(discharge)
            else:
                dispatch_curve.append(0.0)

        elif caso_uso == "Solar + BESS (autoconsumo)":
            if net_load < 0:
                # Excedente solar → cargar batería
                charge = min(-net_load, bess_power_kw, soc_max - soc)
                soc = min(soc + charge, soc_max)
                charged_kwh += charge
                dispatch_curve.append(-charge)
            elif net_load > 0 and soc > soc_min:
                # Déficit solar → descargar batería
                discharge = min(net_load, bess_power_kw, soc - soc_min)
                soc = max(soc - discharge, soc_min)
                discharged_kwh += discharge * rte
                dispatch_curve.append(discharge)
            else:
                dispatch_curve.append(0.0)

        soc_curve.append(round(soc, 3))

    # Ciclos diarios = energía descargada / capacidad útil
    daily_cycles = discharged_kwh / bess_energy_kwh if bess_energy_kwh > 0 else 0

    # Pad dispatch_curve to 24 if any path left it short (defensive)
    while len(dispatch_curve) < 24:
        dispatch_curve.append(0.0)
    while len(soc_curve) < 24:
        soc_curve.append(soc_curve[-1] if soc_curve else 0.0)

    peak_residual = max(
        demand[i] - max(dispatch_curve[i], 0.0) for i in range(24)
    )

    return dict(
        charged_kwh=round(charged_kwh, 2),
        discharged_kwh=round(discharged_kwh, 2),
        daily_cycles=round(daily_cycles, 4),
        peak_residual=round(peak_residual, 2),
        autonomy_hours=None,
        soc_curve=soc_curve,
        dispatch_curve=dispatch_curve,
        demand_curve=demand,
    )


@st.cache_data(show_spinner=False)
def calc_bess_sizing(
    caso_uso: str,
    daily_kwh: float,
    peak_demand_kw: float,
    autonomy_hours: float,
    peak_shaving_target_kw: float,
    arbitrage_daily_cycles: float,
    dod: float,
    rte: float,
    bess_power_override_kw: float,   # 0 = calcular automáticamente
) -> dict:
    """
    Dimensiona el BESS para el caso de uso dado.
    Retorna capacidad bruta, capacidad útil, potencia PCS y justificación.

    Metodología:
    - Respaldo/UPS: capacidad dimensionada por energía requerida en autonomía
    - Peak shaving: potencia dimensionada por kW a cortar, energía por duración del pico
    - Arbitraje: energía dimensionada por kWh a ciclar por día, potencia por C-rate
    - Solar+BESS: energía dimensionada por excedente solar esperado
    """
    if caso_uso == "Respaldo / UPS":
        # Energía útil requerida = consumo en horas de autonomía
        # Potencia = demanda pico (la batería debe soportar toda la carga)
        useful_kwh = peak_demand_kw * autonomy_hours
        pcs_kw     = peak_demand_kw * 1.0   # 1:1 con demanda pico (norma IEC 62477)
        reason_e   = f"Autonomía {autonomy_hours:.1f}h × {peak_demand_kw:.1f} kW pico"
        reason_p   = f"Igual a demanda pico ({peak_demand_kw:.1f} kW) — IEC 62477"

    elif caso_uso == "Peak shaving (cortar punta de demanda)":
        kw_a_cortar = max(0.0, peak_demand_kw - peak_shaving_target_kw)
        # Duración típica del pico: 4 horas (horario CFE punta 18-22h)
        pico_duration_h = 4.0
        useful_kwh = kw_a_cortar * pico_duration_h
        pcs_kw     = kw_a_cortar * 1.05   # 5% margen sobre corte requerido
        reason_e   = f"{kw_a_cortar:.1f} kW × {pico_duration_h:.0f}h punta CFE"
        reason_p   = f"{kw_a_cortar:.1f} kW a cortar + 5% margen"

    elif caso_uso == "Arbitraje tarifario (valle→punta)":
        # Energía a ciclar = fracción del consumo que ocurre en horario punta.
        # El horario punta CFE típico es 4h/día sobre 24h ≈ 16.7% del consumo diario.
        # Usando 20% como proxy conservador (margen para demanda no lineal).
        # FIX: NO usar daily_kwh completo — sobredimensionaría la batería al asumir que
        # todo el consumo pasa por ella. Solo la fracción desplazada de punta→valle cicla.
        punta_frac  = 4 / 24   # 4 h punta sobre 24 h ≈ 16.7 % del consumo diario
        useful_kwh  = daily_kwh * punta_frac * arbitrage_daily_cycles
        # C-rate de 0.5C para maximizar vida útil (carga/descarga en 2h)
        pcs_kw      = useful_kwh / 2.0
        reason_e    = (f"{daily_kwh:.0f} kWh × {punta_frac:.3f} frac. punta "
                       f"× {arbitrage_daily_cycles:.1f} ciclos/día")
        reason_p    = f"C-rate 0.5C sobre capacidad útil ({useful_kwh/2:.1f} kW)"

    else:  # Solar + BESS
        # Energía = excedente solar estimado (~30% de consumo diario como proxy)
        useful_kwh = daily_kwh * 0.30
        pcs_kw     = useful_kwh / 2.0
        reason_e   = f"30% de consumo diario como proxy de excedente solar"
        reason_p   = f"C-rate 0.5C ({useful_kwh/2:.1f} kW)"

    if bess_power_override_kw > 0:
        pcs_kw   = bess_power_override_kw
        reason_p = f"Definido por usuario ({bess_power_override_kw:.1f} kW)"

    # Capacidad bruta: corrige por DoD y pérdidas round-trip en carga
    # Energía bruta = útil / DoD  (recuperamos la capacidad reservada)
    # El RTE afecta la energía entregada, no la capacidad nominal
    capacity_kwh_bruto = useful_kwh / dod

    return dict(
        useful_kwh=round(useful_kwh, 2),
        capacity_kwh_bruto=round(capacity_kwh_bruto, 2),
        pcs_kw=round(pcs_kw, 2),
        c_rate=round(pcs_kw / capacity_kwh_bruto, 3) if capacity_kwh_bruto > 0 else 0,
        reason_e=reason_e,
        reason_p=reason_p,
    )


@st.cache_data(show_spinner=False)
def calc_bess_financial(
    capacity_kwh_bruto: float,
    pcs_kw: float,
    daily_discharge_kwh: float,   # kWh descargados útiles por día
    daily_cycles: float,          # ciclos/día al DoD de operación
    dod_op: float,                # DoD de operación real
    rte: float,
    vida_util_years: int,
    discount_rate_pct: float,
    inflation_tarifa_pct: float,
    tarifa_desplazada: float,     # $/kWh del bloque tarifario que desplaza
    tarifa_carga: float,          # $/kWh del bloque en que se carga (para arbitraje)
    usd_to_mxn: float,
    om_pct: float = 1.5,          # % CAPEX/año
) -> dict:
    """
    Modelo financiero riguroso para BESS.

    CAPEX:
      - Costo energía: capacity_kwh_bruto × $/kWh
      - Costo PCS: pcs_kw × $/kW
      - BOS: 20% sobre (energía + PCS)
      Total en USD, convertido a MXN

    OPEX anual:
      - O&M = om_pct% del CAPEX MXN, crece con inflación

    Reemplazo de celdas:
      - En el año en que se alcanza EoL (ciclos agotados), se incurre en costo
        de reemplazo = 65% del costo de energía original
      - Si EoL > vida_util, no hay reemplazo en el horizonte

    Ahorro anual:
      - Energía descargada útil × (tarifa_desplazada - tarifa_carga)
      - Para respaldo: valor del kWh no interrumpido (paramétrico, usuario define)
      - Crece con inflación_tarifa anualmente
      - Se degrada con la capacidad de la batería (pérdida lineal de throughput)

    LCOS (Levelized Cost of Storage):
      LCOS = (CAPEX + PV_OPEX + PV_REEMPLAZO) / PV_kWh_ciclados
      Unidad: MXN/kWh ciclado (comparable con spread tarifario)
    """
    p = LFP_PARAMS

    # ── CAPEX ──────────────────────────────────────────────────────────────────
    capex_energia_usd = capacity_kwh_bruto * p["costo_energia_usd_kwh"]
    capex_pcs_usd     = pcs_kw * p["costo_pcs_usd_kw"]
    capex_bos_usd     = (capex_energia_usd + capex_pcs_usd) * p["costo_bos_pct"]
    capex_total_usd   = capex_energia_usd + capex_pcs_usd + capex_bos_usd
    capex_total_mxn   = capex_total_usd * usd_to_mxn

    # ── Vida útil por ciclos (Wöhler) ─────────────────────────────────────────
    ciclos_eol    = woehler_cycles(dod_op, p["woehler"])
    vida_ciclos_y = ciclos_eol / (daily_cycles * 365) if daily_cycles > 0 else vida_util_years
    año_reemplazo = int(math.floor(vida_ciclos_y))  # año exacto en que se reemplaza

    # ── Flujos anuales ─────────────────────────────────────────────────────────
    r     = discount_rate_pct / 100
    inf   = inflation_tarifa_pct / 100
    years = list(range(1, vida_util_years + 1))

    # Energía anual descargada con degradación de capacidad
    kwh_año1 = daily_discharge_kwh * 365
    # Degradación: pérdida lineal de throughput por año (conservador)
    degrad_anual = p["degradacion_lineal_pct_año"] / 100

    ahorro_y  = []
    om_y      = []
    fn_y      = []
    fd_y      = []
    kwh_desc_y = []

    for i, y in enumerate(years):
        # Degradación compuesta (exponencial): cada año pierde degrad_anual sobre el anterior.
        # Modelo lineal (1 - degrad_anual*(y-1)) sobreestima generación tardía.
        factor_deg  = max(0.0, (1 - degrad_anual) ** (y - 1))
        kwh_y       = kwh_año1 * factor_deg
        kwh_desc_y.append(kwh_y)

        # Ahorro neto = ingreso por descarga − costo de carga (incluyendo pérdidas RTE).
        # Para cargar kwh_y kWh útiles se necesita kwh_y/rte kWh de la red.
        # FIX: antes se usaba (tarifa_desplazada - tarifa_carga) × kwh_y, lo que ignoraba
        # que la energía comprada en valle es kwh_y/rte, no kwh_y. Con rte=0.92 la
        # diferencia es ~8.7% en el costo de carga, sobreestimando el ahorro.
        # Fórmula correcta:
        #   ahorro = kwh_y × tarifa_desplazada − (kwh_y / rte) × tarifa_carga
        #          = kwh_y × (tarifa_desplazada − tarifa_carga / rte)
        factor_inf  = (1 + inf) ** (y - 1)
        tar_desp_y  = tarifa_desplazada * factor_inf
        tar_carga_y = tarifa_carga      * factor_inf
        ahorro      = kwh_y * (tar_desp_y - tar_carga_y / rte)
        ahorro_y.append(ahorro)

        # O&M crece con inflación
        om_y.append(capex_total_mxn * (om_pct / 100) * (1 + inf) ** (y - 1))

        # Costo de reemplazo en el año EoL (si cae dentro del horizonte)
        reemplazo = 0.0
        if y == año_reemplazo and año_reemplazo <= vida_util_years:
            reemplazo = capex_energia_usd * p["costo_reemplazo_pct"] * usd_to_mxn

        fn  = ahorro - om_y[-1] - reemplazo
        fd  = fn / (1 + r) ** y
        fn_y.append(fn)
        fd_y.append(fd)

    vpn = -capex_total_mxn + sum(fd_y)

    # TIR — función compartida a nivel de módulo (_bisection_irr)
    tir = _bisection_irr([-capex_total_mxn] + fn_y)

    # Payback con interpolación lineal
    acum = -capex_total_mxn
    pb_simple = None
    for i, fn in enumerate(fn_y):
        prev = acum
        acum += fn
        if acum >= 0 and pb_simple is None:
            pb_simple = round(years[i] - 1 + (-prev) / (acum - prev), 1)

    acum_desc = -capex_total_mxn
    pb_disc = None
    for i, fd in enumerate(fd_y):
        prev = acum_desc
        acum_desc += fd
        if acum_desc >= 0 and pb_disc is None:
            pb_disc = round(years[i] - 1 + (-prev) / (acum_desc - prev), 1)

    # LCOS riguroso: (CAPEX + PV_OPEX + PV_REEMPLAZO) / PV_kWh_ciclados
    # Todos los términos descontados al mismo t=0
    pv_costos = capex_total_mxn + sum(om_y[i] / (1 + r) ** years[i] for i in range(len(years)))
    # Añadir reemplazo descontado si aplica
    if año_reemplazo <= vida_util_years:
        pv_costos += (capex_energia_usd * p["costo_reemplazo_pct"] * usd_to_mxn
                      / (1 + r) ** año_reemplazo)
    pv_kwh = sum(kwh_desc_y[i] / (1 + r) ** years[i] for i in range(len(years)))
    lcos_mxn_kwh = pv_costos / pv_kwh if pv_kwh > 0 else 0.0

    # Acumulados para gráfica
    acum_n, acum_d = [], -capex_total_mxn
    acum_nom, run  = [], -capex_total_mxn
    for fn, fd in zip(fn_y, fd_y):
        run     += fn; acum_nom.append(run)
        acum_d  += fd; acum_d_val = acum_d
        acum_n.append(round(acum_d_val, 0))

    return dict(
        # CAPEX desglosado
        capex_energia_usd=capex_energia_usd,
        capex_pcs_usd=capex_pcs_usd,
        capex_bos_usd=capex_bos_usd,
        capex_total_usd=capex_total_usd,
        capex_total_mxn=capex_total_mxn,
        # Vida útil
        ciclos_eol=round(ciclos_eol, 0),
        vida_ciclos_y=round(vida_ciclos_y, 1),
        año_reemplazo=año_reemplazo,
        # Financiero
        vpn=round(vpn, 0),
        tir=tir,
        pb_simple=pb_simple,
        pb_disc=pb_disc,
        lcos_mxn_kwh=round(lcos_mxn_kwh, 4),
        # Series anuales
        years=years,
        ahorro_y=ahorro_y,
        om_y=om_y,
        fn_y=fn_y,
        fd_y=fd_y,
        kwh_desc_y=kwh_desc_y,
        acum_nom=acum_nom,
        acum_desc=acum_n,
    )


with tab4:
    # ══════════════════════════════════════════════════════════════════════════
    # TAB 4 — BESS STANDALONE  ·  Motor riguroso con Wöhler + LCOS + Despacho
    # ══════════════════════════════════════════════════════════════════════════
    bess_col_l, bess_col_r = st.columns([1, 2.2], gap="large")

    with bess_col_l:
        st.markdown('<div class="app-title">🔋 BESS Standalone</div>', unsafe_allow_html=True)
        st.markdown('<div class="app-sub">Dimensionamiento con curva de Wöhler · LCOS riguroso · Despacho horario</div>', unsafe_allow_html=True)

        # ── Caso de uso ─────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Caso de uso</div>', unsafe_allow_html=True)
        caso_uso = st.selectbox(
            "Selecciona el caso de uso",
            ["Peak shaving (cortar punta de demanda)",
             "Respaldo / UPS (autonomía ante cortes)",
             "Arbitraje tarifario (valle→punta)",
             "Solar + BESS (autoconsumo)"],
            help="El caso de uso determina cómo se dimensiona la potencia y la energía del sistema.",
            key="bess_caso_uso",
        )

        # ── Datos de demanda ────────────────────────────────────────────────
        st.markdown('<div class="section-header">Demanda del cliente</div>', unsafe_allow_html=True)

        perfil_tipo = st.selectbox(
            "Perfil de carga",
            list(LOAD_PROFILES.keys()),
            help="Perfil horario típico. Puedes ajustar los parámetros de demanda abajo.",
            key="bess_perfil",
        )

        bess_input_mode = st.radio(
            "Modo de captura de demanda",
            ["Parámetros simplificados", "Cargar CSV horario (8,760 h)"],
            horizontal=True,
            key="bess_input_mode",
        )

        if bess_input_mode == "Parámetros simplificados":
            peak_demand_kw = st.number_input(
                "Demanda pico / Potencia contratada (kW)", 5.0, 50000.0, 100.0, 5.0,
                help="Pico máximo de demanda medido o contratado con CFE.",
                key="bess_peak_kw")
            daily_kwh = st.number_input(
                "Consumo diario promedio (kWh/día)", 10.0, 500000.0, 300.0, 10.0,
                key="bess_daily_kwh")
            load_profile_24h = tuple(LOAD_PROFILES[perfil_tipo])
            csv_loaded = False
        else:
            uploaded_csv = st.file_uploader(
                "CSV con 8,760 filas · columna 'kW' (una por hora del año)",
                type=["csv"], key="bess_csv")
            if uploaded_csv:
                try:
                    df_csv = pd.read_csv(uploaded_csv)
                    col_kw = [c for c in df_csv.columns if "kw" in c.lower() or "potencia" in c.lower() or "demand" in c.lower()]
                    if col_kw and len(df_csv) >= 8760:
                        series_kw = df_csv[col_kw[0]].values[:8760]
                        peak_demand_kw = float(np.max(series_kw))
                        daily_kwh      = float(np.mean(series_kw) * 24)
                        # Perfil promedio de 24h (colapsar los 365 días)
                        load_profile_24h = tuple(
                            float(np.mean(series_kw[h::24])) / peak_demand_kw
                            for h in range(24)
                        )
                        st.success(f"✅ CSV cargado · {len(df_csv):,} filas · Pico: {peak_demand_kw:.1f} kW · Promedio: {daily_kwh/24:.1f} kW")
                        csv_loaded = True
                    else:
                        st.error("El CSV debe tener ≥ 8,760 filas y una columna con kW.")
                        load_profile_24h = tuple(LOAD_PROFILES[perfil_tipo])
                        peak_demand_kw = 100.0; daily_kwh = 300.0; csv_loaded = False
                except Exception as e:
                    st.error(f"Error al leer CSV: {e}")
                    load_profile_24h = tuple(LOAD_PROFILES[perfil_tipo])
                    peak_demand_kw = 100.0; daily_kwh = 300.0; csv_loaded = False
            else:
                st.info("Carga un CSV o cambia a modo simplificado.")
                load_profile_24h = tuple(LOAD_PROFILES[perfil_tipo])
                peak_demand_kw = 100.0; daily_kwh = 300.0; csv_loaded = False

        # ── Parámetros específicos del caso de uso ──────────────────────────
        st.markdown('<div class="section-header">Parámetros del caso de uso</div>', unsafe_allow_html=True)

        if caso_uso == "Peak shaving (cortar punta de demanda)":
            peak_shaving_target_kw = st.number_input(
                "Demanda objetivo tras peak shaving (kW)",
                0.0, peak_demand_kw * 0.99, peak_demand_kw * 0.75, 5.0,
                help="kW al que se quiere limitar la demanda máxima.",
                key="bess_ps_target")
            autonomy_hours = 4.0
            arb_cycles = 1.0
            st.markdown(f'<div class="info-box">Corte requerido: <b>{peak_demand_kw - peak_shaving_target_kw:.1f} kW</b> · Duración punta CFE: <b>4 h</b></div>', unsafe_allow_html=True)

        elif caso_uso == "Respaldo / UPS (autonomía ante cortes)":
            autonomy_hours = st.slider(
                "Autonomía requerida (horas)", 0.5, 24.0, 4.0, 0.5,
                help="Horas que el BESS debe cubrir la demanda pico completa sin red.",
                key="bess_aut_h")
            peak_shaving_target_kw = 0.0
            arb_cycles = 0.0

        elif caso_uso == "Arbitraje tarifario (valle→punta)":
            arb_cycles = st.slider(
                "Ciclos diarios de arbitraje", 0.5, 2.0, 1.0, 0.25,
                help="Número de veces que se carga/descarga la batería por día.",
                key="bess_arb_cycles")
            autonomy_hours = 0.0
            peak_shaving_target_kw = 0.0

        else:  # Solar + BESS
            autonomy_hours = 0.0
            peak_shaving_target_kw = 0.0
            arb_cycles = 1.0

        # ── Tecnología y parámetros de operación ───────────────────────────
        st.markdown('<div class="section-header">Parámetros LFP</div>', unsafe_allow_html=True)
        st.markdown('<div class="info-box">Solo LFP (LiFePO₄) — tecnología estándar BESS comercial/industrial 2024. Parámetros de referencia CATL / BYD.</div>', unsafe_allow_html=True)

        dod_op = st.slider(
            "DoD de operación (%)", 50, 95, 90, 5,
            help="A mayor DoD más energía útil pero menos ciclos de vida. 90% es el estándar LFP.",
            key="bess_dod") / 100
        rte_op = st.slider(
            "Eficiencia round-trip AC-AC (%)", 85, 96, 92, 1,
            help="Incluye pérdidas del PCS. LFP típico: 92-93%.",
            key="bess_rte") / 100

        bess_power_override = st.number_input(
            "Potencia PCS manual (kW) — 0 = calcular automáticamente",
            0.0, 10000.0, 0.0, 5.0,
            help="Si sabes la potencia del inversor bidireccional, ingrésala aquí.",
            key="bess_pcs_override")

        # ── Parámetros financieros ──────────────────────────────────────────
        st.markdown('<div class="section-header">Parámetros financieros</div>', unsafe_allow_html=True)

        bess_vida_util = st.slider("Horizonte de análisis (años)", 5, 20, 15, 1, key="bess_vida")
        bess_discount  = st.slider("Tasa de descuento WACC (%)", 8.0, 25.0, 15.0, 0.5, key="bess_wacc")
        bess_inf_tar   = st.slider("Inflación anual de tarifa (%)", 0.0, 12.0, 6.0, 0.5, key="bess_inf")

        st.markdown('<div class="section-header">Tarifas eléctricas</div>', unsafe_allow_html=True)
        tarifa_punta_bess = st.number_input(
            "Tarifa bloque punta (MXN/kWh)", 1.0, 30.0,
            st.session_state.get('tarifa', 5.50), 0.1,
            help="Costo del kWh en horario punta (lo que la batería ahorra al descargar).",
            key="bess_tar_punta")
        tarifa_valle_bess = st.number_input(
            "Tarifa bloque valle / base (MXN/kWh)", 0.5, 15.0, 2.80, 0.1,
            help="Costo del kWh en horario base (lo que cuesta cargar la batería).",
            key="bess_tar_valle")

    # ── Panel derecho: resultados ────────────────────────────────────────────
    with bess_col_r:

        # ── 1. DIMENSIONAMIENTO ─────────────────────────────────────────────
        sizing = calc_bess_sizing(
            caso_uso=caso_uso,
            daily_kwh=daily_kwh,
            peak_demand_kw=peak_demand_kw,
            autonomy_hours=autonomy_hours,
            peak_shaving_target_kw=peak_shaving_target_kw,
            arbitrage_daily_cycles=arb_cycles,
            dod=dod_op,
            rte=rte_op,
            bess_power_override_kw=bess_power_override,
        )

        useful_kwh      = sizing["useful_kwh"]
        capacity_bruto  = sizing["capacity_kwh_bruto"]
        pcs_kw          = sizing["pcs_kw"]
        c_rate          = sizing["c_rate"]

        # Ciclos LFP al DoD de operación (Wöhler)
        ciclos_eol = woehler_cycles(dod_op, LFP_PARAMS["woehler"])

        # ── Validar inputs BESS antes de calcular ─────────────────────────
        _ok_b4, _msg_b4 = _validate_bess_inputs(peak_demand_kw, daily_kwh, useful_kwh, pcs_kw)
        if not _ok_b4:
            st.error(f"⚠️ {_msg_b4}")
            st.stop()

        # ── Validar inputs BESS ────────────────────────────────────────────
        _ok_b4, _msg_b4 = _validate_bess_inputs(peak_demand_kw, daily_kwh, useful_kwh, pcs_kw)
        if not _ok_b4:
            st.error(f"⚠️ {_msg_b4}")
            st.stop()

        # ── 2. DESPACHO HORARIO ─────────────────────────────────────────────
        dispatch = simulate_bess_dispatch(
            load_profile_24h=load_profile_24h,
            peak_demand_kw=peak_demand_kw,
            daily_kwh=daily_kwh,
            bess_power_kw=pcs_kw,
            bess_energy_kwh=useful_kwh,
            rte=rte_op,
            caso_uso=caso_uso,
            peak_shaving_target_kw=peak_shaving_target_kw
                if caso_uso == "Peak shaving (cortar punta de demanda)" else 0.0,
        )

        daily_discharge = dispatch["discharged_kwh"]
        daily_cycles    = dispatch["daily_cycles"]

        # ── 3. MODELO FINANCIERO RIGUROSO ───────────────────────────────────
        fm_bess = calc_bess_financial(
            capacity_kwh_bruto=capacity_bruto,
            pcs_kw=pcs_kw,
            daily_discharge_kwh=daily_discharge,
            daily_cycles=daily_cycles,
            dod_op=dod_op,
            rte=rte_op,
            vida_util_years=bess_vida_util,
            discount_rate_pct=bess_discount,
            inflation_tarifa_pct=bess_inf_tar,
            tarifa_desplazada=tarifa_punta_bess,
            tarifa_carga=tarifa_valle_bess,
            usd_to_mxn=usd_to_mxn,
        )

        vida_ciclos_y = fm_bess["vida_ciclos_y"]
        año_reemplazo = fm_bess["año_reemplazo"]

        # ── HERO: resultados de dimensionamiento ───────────────────────────
        capex_usd = fm_bess["capex_total_usd"]
        capex_mxn = fm_bess["capex_total_mxn"]
        c_rate_badge_color = "#4ade80" if c_rate <= 1.0 else "#facc15" if c_rate <= 2.0 else "#f87171"
        c_rate_note = "✅ Conservador" if c_rate <= 0.5 else ("✅ Normal" if c_rate <= 1.0 else "⚠️ Alto — revisar vida útil")

        st.markdown(f"""
<div class="tor-hero">
  <div class="th-project">🔋 PRE-SIZING BESS · LFP · {caso_uso.upper()}</div>
  <div class="th-meta">
    Tecnología LFP (LiFePO₄) &nbsp;·&nbsp; DoD {dod_op*100:.0f}% &nbsp;·&nbsp;
    RTE {rte_op*100:.0f}% &nbsp;·&nbsp;
    <span style="color:{c_rate_badge_color};">C-rate {c_rate:.2f}C — {c_rate_note}</span>
  </div>
  <div class="th-grid" style="grid-template-columns:repeat(4,1fr);">
    <div class="th-item">
      <span class="th-label">CAPACIDAD BRUTA</span>
      <span class="th-val">{capacity_bruto:.1f}</span>
      <span class="th-unit">kWh nominales</span>
    </div>
    <div class="th-item">
      <span class="th-label">ENERGÍA ÚTIL</span>
      <span class="th-val">{useful_kwh:.1f}</span>
      <span class="th-unit">kWh (DoD {dod_op*100:.0f}%)</span>
    </div>
    <div class="th-item">
      <span class="th-label">POTENCIA PCS</span>
      <span class="th-val">{pcs_kw:.1f}</span>
      <span class="th-unit">kW inversor bidireccional</span>
    </div>
    <div class="th-item">
      <span class="th-label">INVERSIÓN REF.</span>
      <span class="th-val">${capex_usd:,.0f}</span>
      <span class="th-unit">USD · ≈ ${capex_mxn:,.0f} MXN</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── CAPEX desglosado ───────────────────────────────────────────────
        st.markdown('<div class="section-header">Desglose CAPEX</div>', unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Celdas + BMS + Rack",
                  f"${fm_bess['capex_energia_usd']:,.0f} USD",
                  f"${fm_bess['capex_energia_usd']*usd_to_mxn:,.0f} MXN")
        c2.metric("PCS (inversor bidir.)",
                  f"${fm_bess['capex_pcs_usd']:,.0f} USD",
                  f"${fm_bess['capex_pcs_usd']*usd_to_mxn:,.0f} MXN")
        c3.metric("BOS (cableado, obra civil)",
                  f"${fm_bess['capex_bos_usd']:,.0f} USD",
                  f"20% sobre celdas+PCS")
        c4.metric("Total instalado",
                  f"${capex_usd:,.0f} USD",
                  f"${LFP_PARAMS['costo_energia_usd_kwh']:.0f}/kWh + ${LFP_PARAMS['costo_pcs_usd_kw']:.0f}/kW")

        # ── Vida útil y ciclos (Wöhler) ────────────────────────────────────
        st.markdown('<div class="section-header">Vida útil por ciclos — Curva de Wöhler LFP</div>', unsafe_allow_html=True)

        # Graficar curva Wöhler
        dod_pts = [d for d, _ in LFP_PARAMS["woehler"]]
        cyc_pts = [c for _, c in LFP_PARAMS["woehler"]]
        dod_fine = [i/100 for i in range(20, 101)]
        cyc_fine = [woehler_cycles(d, LFP_PARAMS["woehler"]) for d in dod_fine]

        fig_woe = go.Figure()
        fig_woe.add_trace(go.Scatter(
            x=[d*100 for d in dod_fine], y=cyc_fine,
            mode="lines", name="Ciclos hasta EoL (80%)",
            line=dict(color=TEAL, width=3),
            fill="tozeroy", fillcolor="rgba(20,184,166,0.08)",
        ))
        fig_woe.add_trace(go.Scatter(
            x=[d*100 for d in dod_pts], y=cyc_pts,
            mode="markers", name="Puntos datasheet",
            marker=dict(color=AMBER, size=8, symbol="diamond"),
        ))
        # Marcar punto de operación actual
        fig_woe.add_trace(go.Scatter(
            x=[dod_op*100], y=[ciclos_eol],
            mode="markers+text",
            name=f"Op. actual: {ciclos_eol:,.0f} ciclos",
            marker=dict(color=ROSE, size=14, symbol="star"),
            text=[f"  {ciclos_eol:,.0f} ciclos · {vida_ciclos_y:.1f} años"],
            textposition="middle right",
            textfont=dict(color=ROSE, size=10),
        ))
        lay_woe = copy.deepcopy(PLOT_LAYOUT)
        lay_woe.update({
            "height": 280,
            "xaxis": dict(title="DoD (%)", gridcolor="#1e2230", range=[15, 105]),
            "yaxis": dict(title="Ciclos hasta EoL", gridcolor="#1e2230", tickformat=","),
            "legend": dict(orientation="h", y=1.12, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
            "margin": dict(l=10, r=20, t=40, b=40),
        })
        fig_woe.update_layout(**lay_woe)
        st.plotly_chart(fig_woe, use_container_width=True)

        w1, w2, w3 = st.columns(3)
        w1.metric("Ciclos EoL (Wöhler)", f"{ciclos_eol:,.0f}", f"DoD {dod_op*100:.0f}%")
        w2.metric("Vida útil por ciclos", f"{vida_ciclos_y:.1f} años",
                  f"{daily_cycles:.3f} ciclos/día")
        w3.metric("Año de reemplazo celdas",
                  f"Año {año_reemplazo}" if año_reemplazo <= bess_vida_util else "Sin reemplazo",
                  f"en horizonte de {bess_vida_util} años")

        # ── Despacho horario ────────────────────────────────────────────────
        st.markdown('<div class="section-header">Despacho horario — día típico</div>', unsafe_allow_html=True)

        hours_24 = list(range(24))
        demand_curve = dispatch["demand_curve"]
        disp_curve   = dispatch["dispatch_curve"]
        soc_curve    = dispatch["soc_curve"]

        # Demanda residual (después del BESS)
        demand_res = [max(0.0, demand_curve[h] - max(disp_curve[h], 0)) for h in hours_24]

        fig_disp = make_subplots(
            rows=2, cols=1, shared_xaxes=True,
            row_heights=[0.6, 0.4],
            subplot_titles=["Demanda bruta vs residual y despacho BESS", "Estado de carga (SoC)"],
            vertical_spacing=0.12,
        )
        fig_disp.add_trace(go.Scatter(
            x=hours_24, y=demand_curve, name="Demanda bruta",
            line=dict(color=ROSE, width=2, dash="dot"),
            fill="tozeroy", fillcolor="rgba(244,63,94,0.05)",
        ), row=1, col=1)
        fig_disp.add_trace(go.Scatter(
            x=hours_24, y=demand_res, name="Demanda residual",
            line=dict(color=AMBER, width=2),
            fill="tozeroy", fillcolor="rgba(245,158,11,0.08)",
        ), row=1, col=1)
        # Despacho: positivo=descarga (verde), negativo=carga (azul)
        fig_disp.add_trace(go.Bar(
            x=hours_24,
            y=[max(d, 0) for d in disp_curve],
            name="Descarga BESS",
            marker_color=TEAL, opacity=0.8,
        ), row=1, col=1)
        fig_disp.add_trace(go.Bar(
            x=hours_24,
            y=[min(d, 0) for d in disp_curve],
            name="Carga BESS",
            marker_color=BLUE, opacity=0.7,
        ), row=1, col=1)
        if caso_uso == "Peak shaving (cortar punta de demanda)":
            fig_disp.add_hline(
                y=peak_shaving_target_kw, row=1, col=1,
                line_color=ROSE, line_dash="dash", line_width=1.5,
                annotation_text=f"Target {peak_shaving_target_kw:.0f} kW",
                annotation_font=dict(color=ROSE, size=10),
            )
        fig_disp.add_trace(go.Scatter(
            x=hours_24, y=soc_curve, name="SoC (kWh)",
            line=dict(color=VIOLET, width=2),
            fill="tozeroy", fillcolor="rgba(139,92,246,0.10)",
        ), row=2, col=1)
        fig_disp.add_hline(y=useful_kwh, row=2, col=1,
                           line_color="#475569", line_dash="dot", line_width=1)
        lay_disp = copy.deepcopy(PLOT_LAYOUT)
        lay_disp.update({
            "height": 440, "barmode": "relative",
            "legend": dict(orientation="h", y=1.08, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
            "margin": dict(l=10, r=10, t=60, b=40),
            "xaxis2": dict(title="Hora del día", gridcolor="#1e2230"),
            "yaxis":  dict(title="kW", gridcolor="#1e2230"),
            "yaxis2": dict(title="kWh", gridcolor="#1e2230"),
        })
        fig_disp.update_layout(**lay_disp)
        st.plotly_chart(fig_disp, use_container_width=True)

        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Descarga diaria útil", f"{daily_discharge:.1f} kWh/día")
        d2.metric("Ciclos/día", f"{daily_cycles:.3f}", "DoD operación")
        d3.metric("Pico residual",
                  f"{dispatch['peak_residual']:.1f} kW",
                  f"↓ {peak_demand_kw - dispatch['peak_residual']:.1f} kW vs sin BESS")
        if dispatch["autonomy_hours"] is not None:
            d4.metric("Autonomía real", f"{dispatch['autonomy_hours']:.1f} h",
                      "a demanda pico constante")
        else:
            d4.metric("kWh anuales ciclados", f"{daily_discharge*365:,.0f}", "año 1")

        # ── Modelo financiero ───────────────────────────────────────────────
        vpn    = fm_bess["vpn"]
        tir    = fm_bess["tir"]
        lcos   = fm_bess["lcos_mxn_kwh"]
        pb_s   = fm_bess["pb_simple"]
        pb_d   = fm_bess["pb_disc"]
        spread = tarifa_punta_bess - tarifa_valle_bess

        kc = "#4ade80" if vpn > 0 else "#f87171"
        tir_str  = f"{tir:.1f}%" if tir else "N/D"
        pb_s_str = f"{pb_s:.1f} a" if pb_s else f">{bess_vida_util} a"
        pb_d_str = f"{pb_d:.1f} a" if pb_d else f">{bess_vida_util} a"
        lcos_color = "#4ade80" if lcos < spread else "#f87171"
        lcos_note  = "✅ LCOS < spread tarifario" if lcos < spread else "🔴 LCOS > spread — revisar"

        st.markdown(f"""
<div style="margin-top:8px">
  <div class="section-header">Modelo financiero · {bess_vida_util} años · WACC {bess_discount:.1f}%
  {f"· ⚠️ Reemplazo de celdas en Año {año_reemplazo}" if año_reemplazo <= bess_vida_util else "· Sin reemplazo en horizonte"}</div>
</div>
<div style="display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:12px;">
  <div class="snap-card">
    <div class="sc-label">VPN</div>
    <div class="sc-val" style="color:{kc};">${vpn:,.0f}</div>
    <div class="sc-sub">MXN</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">TIR</div>
    <div class="sc-val" style="color:#22d3ee;">{tir_str}</div>
    <div class="sc-sub">vs {bess_discount:.1f}% WACC</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">LCOS</div>
    <div class="sc-val" style="color:{lcos_color};">${lcos:.3f}</div>
    <div class="sc-sub">MXN/kWh ciclado · {lcos_note}</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Spread tarifario</div>
    <div class="sc-val" style="color:#f1f5f9;">${spread:.3f}</div>
    <div class="sc-sub">MXN/kWh punta-valle</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Payback simple</div>
    <div class="sc-val">{pb_s_str}</div>
    <div class="sc-sub">nominal</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Payback desc.</div>
    <div class="sc-val">{pb_d_str}</div>
    <div class="sc-sub">WACC {bess_discount:.1f}%</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Gráficas financieras ────────────────────────────────────────────
        fg1, fg2 = st.columns(2, gap="medium")

        with fg1:
            st.markdown('<div class="section-header">Flujos anuales con reemplazo</div>', unsafe_allow_html=True)
            colors_fn = [AMBER if v >= 0 else ROSE for v in fm_bess["fn_y"]]
            fig_fn = go.Figure()
            fig_fn.add_trace(go.Bar(
                x=fm_bess["years"], y=fm_bess["fn_y"],
                name="Flujo neto", marker_color=colors_fn, opacity=0.9,
                hovertemplate="<b>Año %{x}</b><br>Flujo neto: $%{y:,.0f} MXN<extra></extra>",
            ))
            fig_fn.add_trace(go.Scatter(
                x=fm_bess["years"], y=fm_bess["ahorro_y"],
                name="Ahorro bruto", mode="lines",
                line=dict(color=TEAL, width=2, dash="dot"),
            ))
            if año_reemplazo <= bess_vida_util:
                fig_fn.add_vline(x=año_reemplazo, line_color=ROSE, line_dash="dash",
                                 annotation_text=f"Reemplazo celdas",
                                 annotation_font=dict(color=ROSE, size=10))
            lay_fn = copy.deepcopy(PLOT_LAYOUT)
            lay_fn.update({
                "height": 320,
                "yaxis": dict(title="MXN", gridcolor="#1e2230", tickformat=","),
                "xaxis": dict(title="Año", gridcolor="#1e2230", tickmode="linear"),
                "legend": dict(orientation="h", y=1.12, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=10, r=10, t=50, b=40),
            })
            fig_fn.update_layout(**lay_fn)
            st.plotly_chart(fig_fn, use_container_width=True)

        with fg2:
            st.markdown('<div class="section-header">VPN acumulado y payback</div>', unsafe_allow_html=True)
            fig_vpn_b = go.Figure()
            fig_vpn_b.add_trace(go.Scatter(
                x=[0] + fm_bess["years"],
                y=[-capex_mxn] + fm_bess["acum_desc"],
                name="VPN acumulado (desc.)",
                mode="lines+markers",
                line=dict(color=TEAL, width=3),
                marker=dict(size=5),
                fill="tozeroy", fillcolor="rgba(20,184,166,0.07)",
            ))
            fig_vpn_b.add_trace(go.Scatter(
                x=[0] + fm_bess["years"],
                y=[-capex_mxn] + fm_bess["acum_nom"],
                name="Acum. nominal",
                mode="lines", line=dict(color=AMBER, width=2, dash="dash"),
            ))
            fig_vpn_b.add_hline(y=0, line_color="#475569", line_width=1)
            if pb_s:
                fig_vpn_b.add_vline(x=pb_s, line_color=AMBER, line_dash="dot",
                                    annotation_text=f"PB {pb_s:.1f}a",
                                    annotation_font=dict(color=AMBER, size=10))
            if año_reemplazo <= bess_vida_util:
                fig_vpn_b.add_vline(x=año_reemplazo, line_color=ROSE, line_dash="dash",
                                    annotation_text="Reemplazo",
                                    annotation_font=dict(color=ROSE, size=10))
            lay_vpn_b = copy.deepcopy(PLOT_LAYOUT)
            lay_vpn_b.update({
                "height": 320,
                "yaxis": dict(title="MXN acumulados", gridcolor="#1e2230", tickformat=","),
                "xaxis": dict(title="Año", gridcolor="#1e2230", tickmode="linear"),
                "legend": dict(orientation="h", y=1.12, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=10, r=10, t=50, b=40),
            })
            fig_vpn_b.update_layout(**lay_vpn_b)
            st.plotly_chart(fig_vpn_b, use_container_width=True)

        # ── Tabla año a año ─────────────────────────────────────────────────
        st.markdown('<div class="section-header">Tabla financiera año a año</div>', unsafe_allow_html=True)
        df_fin_bess = pd.DataFrame({
            "Año":                fm_bess["years"],
            "kWh ciclados":       [f"{v:,.0f}" for v in fm_bess["kwh_desc_y"]],
            "Ahorro bruto (MXN)": [f"${v:,.0f}" for v in fm_bess["ahorro_y"]],
            "O&M (MXN)":          [f"${v:,.0f}" for v in fm_bess["om_y"]],
            "Flujo neto (MXN)":   [f"${v:,.0f}" for v in fm_bess["fn_y"]],
            "Flujo desc. (MXN)":  [f"${v:,.0f}" for v in fm_bess["fd_y"]],
            "VPN acum. (MXN)":    [f"${v:,.0f}" for v in fm_bess["acum_desc"]],
        })
        st.dataframe(df_fin_bess, use_container_width=True, hide_index=True)

        # ── Criterio de decisión ─────────────────────────────────────────────
        if lcos < spread * 0.85 and vpn > 0 and (tir or 0) > bess_discount:
            st.success(f"🟢 **Proyecto muy atractivo** — LCOS ${lcos:.3f}/kWh vs spread ${spread:.3f}/kWh · VPN positivo · TIR {tir_str}")
        elif lcos < spread and vpn > 0:
            st.success(f"🟡 **Proyecto viable** — LCOS menor al spread tarifario pero margen ajustado")
        elif lcos < spread:
            st.warning(f"⚠️ **LCOS factible pero VPN negativo** — revisar CAPEX, tarifas o vida útil")
        else:
            st.error(f"🔴 **No viable financieramente** — LCOS ${lcos:.3f}/kWh supera spread ${spread:.3f}/kWh")

        st.markdown(f"""
<div class="info-box">
  <b>Metodología:</b> Capacidad dimensionada según caso de uso ·
  Vida útil por curva de Wöhler LFP (CATL datasheet) ·
  LCOS = (CAPEX + PV_OPEX + PV_Reemplazo) / PV_kWh_ciclados ·
  Reemplazo de celdas en año {año_reemplazo} ({LFP_PARAMS["costo_reemplazo_pct"]*100:.0f}% del CAPEX de celdas) ·
  O&M = {LFP_PARAMS["om_pct_anual"]*100:.1f}% CAPEX/año ·
  Despacho horario con estrategia {caso_uso.split("(")[0].strip()}
</div>
""", unsafe_allow_html=True)


with tab5:
    # ══════════════════════════════════════════════════════════════════════════
    # TAB 5 — BESS PPA  ·  Motor riguroso coherente con Módulo Standalone
    # El desarrollador instala y opera el BESS; el cliente paga $/kWh descargado.
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="app-title">🔋 BESS · PPA</div>', unsafe_allow_html=True)
    st.markdown('<div class="app-sub">El desarrollador financia y opera el BESS · El cliente paga por kWh descargado · Motor coherente con Módulo Standalone</div>', unsafe_allow_html=True)

    ppa5_l, ppa5_r = st.columns([1, 2.2], gap="large")

    with ppa5_l:
        # ── Configuración técnica del sistema ──────────────────────────────
        st.markdown('<div class="section-header">Sistema BESS</div>', unsafe_allow_html=True)
        st.markdown('<div class="info-box">Los parámetros técnicos reutilizan la física del Módulo Standalone (Wöhler, CAPEX desglosado, LCOS).</div>', unsafe_allow_html=True)

        p5_caso = st.selectbox(
            "Caso de uso del BESS",
            ["Peak shaving (cortar punta de demanda)",
             "Arbitraje tarifario (valle→punta)",
             "Respaldo / UPS (autonomía ante cortes)"],
            help="Determina cómo opera la batería y cuántos kWh descarga por día.",
            key="p5_caso")

        p5_perfil = st.selectbox("Perfil de carga cliente", list(LOAD_PROFILES.keys()), key="p5_perfil")
        p5_peak_kw   = st.number_input("Demanda pico cliente (kW)", 5.0, 50000.0, 200.0, 10.0, key="p5_peak")
        p5_daily_kwh = st.number_input("Consumo diario (kWh/día)", 10.0, 500000.0, 600.0, 10.0, key="p5_daily")

        if p5_caso == "Peak shaving (cortar punta de demanda)":
            p5_ps_target = st.number_input(
                "Demanda objetivo peak shaving (kW)",
                0.0, p5_peak_kw * 0.99, p5_peak_kw * 0.75, 5.0, key="p5_ps")
            p5_aut_h = 4.0; p5_arb = 1.0
        elif p5_caso == "Respaldo / UPS (autonomía ante cortes)":
            p5_aut_h = st.slider("Autonomía (horas)", 0.5, 12.0, 4.0, 0.5, key="p5_aut")
            p5_ps_target = 0.0; p5_arb = 0.0
        else:
            p5_arb = st.slider("Ciclos diarios arbitraje", 0.5, 2.0, 1.0, 0.25, key="p5_arb")
            p5_ps_target = 0.0; p5_aut_h = 0.0

        p5_dod = st.slider("DoD operación (%)", 50, 95, 90, 5, key="p5_dod") / 100
        p5_rte = st.slider("Eficiencia round-trip (%)", 85, 96, 92, 1, key="p5_rte") / 100
        p5_pcs_override = st.number_input("Potencia PCS manual (kW, 0=auto)", 0.0, 10000.0, 0.0, 5.0, key="p5_pcs")

        # ── Estructura financiera PPA ───────────────────────────────────────
        st.markdown('<div class="section-header">Contrato PPA</div>', unsafe_allow_html=True)
        p5_precio_ppa  = st.number_input(
            "Precio PPA año 1 (MXN/kWh descargado)", 0.50, 15.0, 3.50, 0.10,
            help="Lo que el cliente paga por cada kWh que la batería le entrega.",
            key="p5_precio")
        p5_esc_ppa     = st.slider("Escalador PPA anual (%)", 0.0, 8.0, 4.0, 0.5, key="p5_esc")
        p5_plazo       = st.selectbox("Plazo del contrato (años)", [8, 10, 12, 15, 20], index=2, key="p5_plazo")
        p5_vida_util   = st.slider("Vida útil total del sistema (años)", p5_plazo, 25, max(p5_plazo, 15), 1, key="p5_vida")

        # ── Parámetros financieros desarrollador ───────────────────────────
        st.markdown('<div class="section-header">Parámetros financieros</div>', unsafe_allow_html=True)
        p5_wacc        = st.slider("WACC desarrollador (%)", 8.0, 25.0, 15.0, 0.5, key="p5_wacc")
        p5_equity_pct  = st.slider("Equity del desarrollador (%)", 20, 100, 40, 5, key="p5_eq")
        p5_con_fin     = p5_equity_pct < 100
        if p5_con_fin:
            p5_tasa_deuda  = st.slider("Tasa de deuda anual (%)", 5.0, 20.0, 10.0, 0.5, key="p5_td")
            p5_plazo_deuda = st.slider("Plazo deuda (años)", 3, 15, 7, 1, key="p5_pd")
        else:
            p5_tasa_deuda = 0.0; p5_plazo_deuda = 0

        p5_inf_om      = st.slider("Inflación O&M (%)", 0.0, 10.0, 5.0, 0.5, key="p5_inf_om")

        # Tarifa que el cliente paga a CFE (para calcular ahorro del cliente)
        st.markdown('<div class="section-header">Tarifa CFE del cliente</div>', unsafe_allow_html=True)
        p5_tarifa_cfe  = st.number_input("Tarifa CFE actual (MXN/kWh)", 1.0, 20.0, 5.50, 0.10, key="p5_tcfe")
        p5_inf_cfe     = st.slider("Inflación tarifa CFE (%/año)", 0.0, 12.0, 7.0, 0.5, key="p5_icfe")

    with ppa5_r:

        # ── 1. DIMENSIONAMIENTO (mismo motor que Tab4) ──────────────────────
        p5_sizing = calc_bess_sizing(
            caso_uso=p5_caso,
            daily_kwh=p5_daily_kwh,
            peak_demand_kw=p5_peak_kw,
            autonomy_hours=p5_aut_h,
            peak_shaving_target_kw=p5_ps_target,
            arbitrage_daily_cycles=p5_arb,
            dod=p5_dod,
            rte=p5_rte,
            bess_power_override_kw=p5_pcs_override,
        )
        p5_useful_kwh = p5_sizing["useful_kwh"]
        p5_cap_bruto  = p5_sizing["capacity_kwh_bruto"]
        p5_pcs_kw     = p5_sizing["pcs_kw"]
        p5_c_rate     = p5_sizing["c_rate"]

        # ── Validar inputs BESS antes de calcular ─────────────────────────
        _ok_b5, _msg_b5 = _validate_bess_inputs(p5_peak_kw, p5_daily_kwh, p5_useful_kwh, p5_pcs_kw)
        if not _ok_b5:
            st.error(f"⚠️ {_msg_b5}")
            st.stop()

        # ── Validar inputs BESS tab5 ──────────────────────────────────────
        _ok_b5, _msg_b5 = _validate_bess_inputs(p5_peak_kw, p5_daily_kwh, p5_useful_kwh, p5_pcs_kw)
        if not _ok_b5:
            st.error(f"⚠️ {_msg_b5}")
            st.stop()

        # ── 2. DESPACHO HORARIO (mismo motor que Tab4) ──────────────────────
        p5_dispatch = simulate_bess_dispatch(
            load_profile_24h=tuple(LOAD_PROFILES[p5_perfil]),
            peak_demand_kw=p5_peak_kw,
            daily_kwh=p5_daily_kwh,
            bess_power_kw=p5_pcs_kw,
            bess_energy_kwh=p5_useful_kwh,
            rte=p5_rte,
            caso_uso=p5_caso,
            peak_shaving_target_kw=p5_ps_target if p5_caso == "Peak shaving (cortar punta de demanda)" else 0.0,
        )
        p5_daily_disc  = p5_dispatch["discharged_kwh"]
        p5_daily_cyc   = p5_dispatch["daily_cycles"]
        p5_peak_res    = p5_dispatch["peak_residual"]
        p5_ann_disc    = p5_daily_disc * 365   # kWh descargados año 1

        # ── 3. CAPEX y ciclos de vida (coherente con Tab4) ──────────────────
        p5_capex_en_usd  = p5_cap_bruto * LFP_PARAMS["costo_energia_usd_kwh"]
        p5_capex_pcs_usd = p5_pcs_kw   * LFP_PARAMS["costo_pcs_usd_kw"]
        p5_capex_bos_usd = (p5_capex_en_usd + p5_capex_pcs_usd) * LFP_PARAMS["costo_bos_pct"]
        p5_capex_usd     = p5_capex_en_usd + p5_capex_pcs_usd + p5_capex_bos_usd
        p5_capex_mxn     = p5_capex_usd * usd_to_mxn

        p5_ciclos_eol    = woehler_cycles(p5_dod, LFP_PARAMS["woehler"])
        p5_vida_ciclos_y = p5_ciclos_eol / (p5_daily_cyc * 365) if p5_daily_cyc > 0 else p5_vida_util
        p5_año_reemplazo = int(math.floor(p5_vida_ciclos_y))

        # ── 4. MODELO FINANCIERO PPA ──────────────────────────────────────────
        # Estructura PPA:
        #   Ingresos: kWh descargados × precio PPA (escala con esc_ppa)
        #   Costos:   O&M + Seguro + Servicio deuda + Reemplazo en año EoL
        #   Flujo sobre equity del desarrollador
        #   Perspectiva cliente: compara pago PPA vs tarifa CFE
        r5     = p5_wacc / 100
        inf5   = p5_inf_om / 100

        # Servicio de deuda (si hay financiamiento)
        p5_equity_mxn = p5_capex_mxn * (p5_equity_pct / 100)
        p5_deuda_mxn  = p5_capex_mxn - p5_equity_mxn
        if p5_con_fin and p5_tasa_deuda > 0 and p5_plazo_deuda > 0 and p5_deuda_mxn > 0:
            rd5 = p5_tasa_deuda / 100
            p5_serv_deuda = p5_deuda_mxn * rd5 / (1 - (1 + rd5) ** (-p5_plazo_deuda))
        else:
            p5_serv_deuda = 0.0; p5_deuda_mxn = 0.0; p5_equity_mxn = p5_capex_mxn

        years5     = list(range(1, p5_plazo + 1))
        degrad5    = LFP_PARAMS["degradacion_lineal_pct_año"] / 100

        # Arrays anuales
        gen5_y   = [p5_ann_disc * max(0.0, 1 - degrad5 * i) for i in range(p5_plazo)]
        prec5_y  = [p5_precio_ppa * (1 + p5_esc_ppa / 100) ** i for i in range(p5_plazo)]
        ing5_y   = [gen5_y[i] * prec5_y[i] for i in range(p5_plazo)]
        om5_y    = [p5_capex_mxn * LFP_PARAMS["om_pct_anual"] * (1 + inf5) ** i for i in range(p5_plazo)]
        deu5_y   = [p5_serv_deuda if y <= p5_plazo_deuda else 0.0 for y in years5]
        # Reemplazo de celdas si EoL cae dentro del plazo PPA
        rep5_y   = [
            (p5_capex_en_usd * LFP_PARAMS["costo_reemplazo_pct"] * usd_to_mxn
             if y == p5_año_reemplazo and p5_año_reemplazo <= p5_plazo else 0.0)
            for y in years5
        ]
        fn5_y    = [ing5_y[i] - om5_y[i] - deu5_y[i] - rep5_y[i] for i in range(p5_plazo)]
        fd5_y    = [fn5_y[i] / (1 + r5) ** years5[i] for i in range(p5_plazo)]

        # Valor residual (Gordon suma finita) si plazo < vida_util
        anios_rest5 = max(0, p5_vida_util - p5_plazo)
        if anios_rest5 > 0 and r5 > 0 and p5_año_reemplazo > p5_plazo:
            # Solo hay valor residual si no hay reemplazo pendiente dentro del plazo
            g5 = (p5_esc_ppa / 100) - degrad5
            ratio5 = (1 + g5) / (1 + r5)
            fn_post5 = fn5_y[-1]  # proxy del flujo post-contrato
            if abs(ratio5 - 1.0) < 1e-9:
                vr5 = fn_post5 * anios_rest5 / (1 + r5)
            else:
                vr5 = fn_post5 / (1 + r5) * (1 - ratio5 ** anios_rest5) / (1 - ratio5)
            valor_residual5 = vr5 / (1 + r5) ** p5_plazo
        else:
            valor_residual5 = 0.0

        vpn5 = -p5_equity_mxn + sum(fd5_y) + valor_residual5

        # TIR (bisección)
        def _irr5(cf):
            def npv(rr): return sum(c / (1 + rr) ** t for t, c in enumerate(cf))
            try:
                lo, hi = -0.99, 5.0
                if npv(lo) * npv(hi) > 0: return None
                for _ in range(200):
                    mid = (lo + hi) / 2
                    fm  = npv(mid)
                    if (hi - lo) / 2 < 1e-10: return mid * 100
                    if abs(fm) < 1e-9:        return mid * 100
                    if npv(lo) * fm < 0: hi = mid
                    else: lo = mid
                return ((lo + hi) / 2) * 100
            except Exception: return None

        tir5 = _irr5([-p5_equity_mxn] + fn5_y)

        # Payback sobre equity
        acum5 = -p5_equity_mxn
        pb5_s = None
        for i, fn in enumerate(fn5_y):
            prev5 = acum5; acum5 += fn
            if acum5 >= 0 and pb5_s is None:
                pb5_s = round(years5[i] - 1 + (-prev5) / (acum5 - prev5), 1)

        acum5d = -p5_equity_mxn
        pb5_d  = None
        for i, fd in enumerate(fd5_y):
            prev5d = acum5d; acum5d += fd
            if acum5d >= 0 and pb5_d is None:
                pb5_d = round(years5[i] - 1 + (-prev5d) / (acum5d - prev5d), 1)

        # LCOS del PPA (misma fórmula que Tab4)
        pv_cost5 = p5_capex_mxn + sum(om5_y[i] / (1 + r5) ** years5[i] for i in range(p5_plazo))
        if p5_año_reemplazo <= p5_plazo:
            pv_cost5 += (p5_capex_en_usd * LFP_PARAMS["costo_reemplazo_pct"] * usd_to_mxn
                         / (1 + r5) ** p5_año_reemplazo)
        pv_kwh5  = sum(gen5_y[i] / (1 + r5) ** years5[i] for i in range(p5_plazo))
        lcos5    = pv_cost5 / pv_kwh5 if pv_kwh5 > 0 else 0.0

        # Precio mínimo viable (VPN=0) por bisección
        def _pmin5():
            def vpn_at(p):
                ing_t  = [gen5_y[i] * p * (1 + p5_esc_ppa / 100) ** i for i in range(p5_plazo)]
                fn_t   = [ing_t[i] - om5_y[i] - deu5_y[i] - rep5_y[i] for i in range(p5_plazo)]
                fd_t   = [fn_t[i] / (1 + r5) ** years5[i] for i in range(p5_plazo)]
                return -p5_equity_mxn + sum(fd_t) + valor_residual5
            lo, hi = 0.01, 30.0
            if vpn_at(hi) < 0: return None
            for _ in range(80):
                mid = (lo + hi) / 2
                if vpn_at(mid) >= 0: hi = mid
                else: lo = mid
            return round((lo + hi) / 2, 4)

        precio_min5 = _pmin5()

        # Perspectiva del cliente: ahorro vs CFE
        cfe5_y   = [p5_ann_disc * (p5_tarifa_cfe * (1 + p5_inf_cfe / 100) ** i) for i in range(p5_plazo)]
        pago5_y  = [ing5_y[i] for i in range(p5_plazo)]   # lo que paga al desarrollador
        ahorro5_y = [cfe5_y[i] - pago5_y[i] for i in range(p5_plazo)]
        ahorro5_tot = sum(ahorro5_y)
        desc_vs_cfe = (p5_precio_ppa / p5_tarifa_cfe - 1) * 100  # % diferencia año 1

        # ── HERO ───────────────────────────────────────────────────────────
        kc5  = "#4ade80" if vpn5 > 0 else "#f87171"
        cr5c = "#4ade80" if p5_c_rate <= 1.0 else "#facc15"
        tir5_str  = f"{tir5:.1f}%" if tir5 else "N/D"
        pb5_s_str = f"{pb5_s:.1f} a" if pb5_s else f">{p5_plazo} a"
        pm5_str   = f"${precio_min5:.4f}/kWh" if precio_min5 else "No viable"
        viable5   = precio_min5 is not None and p5_precio_ppa >= precio_min5

        st.markdown(f"""
<div class="tor-hero">
  <div class="th-project">🔋 BESS PPA · LFP · {p5_caso.split("(")[0].strip().upper()}</div>
  <div class="th-meta">
    Capacidad {p5_cap_bruto:.1f} kWh brutos &nbsp;·&nbsp; {p5_useful_kwh:.1f} kWh útiles &nbsp;·&nbsp;
    PCS {p5_pcs_kw:.1f} kW &nbsp;·&nbsp;
    <span style="color:{cr5c};">C-rate {p5_c_rate:.2f}C</span> &nbsp;·&nbsp;
    Plazo {p5_plazo} años &nbsp;·&nbsp;
    {"✅ Precio PPA viable" if viable5 else "⚠️ Precio PPA por debajo del mínimo"}
  </div>
  <div class="th-grid" style="grid-template-columns:repeat(4,1fr);">
    <div class="th-item">
      <span class="th-label">INVERSIÓN TOTAL</span>
      <span class="th-val">${p5_capex_usd:,.0f}</span>
      <span class="th-unit">USD · ≈ ${p5_capex_mxn:,.0f} MXN</span>
    </div>
    <div class="th-item">
      <span class="th-label">kWh DESCARGADOS AÑO 1</span>
      <span class="th-val">{p5_ann_disc:,.0f}</span>
      <span class="th-unit">kWh/año · {p5_daily_disc:.1f} kWh/día</span>
    </div>
    <div class="th-item">
      <span class="th-label">INGRESO AÑO 1</span>
      <span class="th-val">${ing5_y[0]:,.0f}</span>
      <span class="th-unit">MXN · ${p5_precio_ppa:.4f}/kWh</span>
    </div>
    <div class="th-item">
      <span class="th-label">PRECIO MÍNIMO VIABLE</span>
      <span class="th-val" style="color:{'#4ade80' if viable5 else '#f43f5e'};">{pm5_str}</span>
      <span class="th-unit">VPN = 0 a WACC {p5_wacc:.1f}%</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── KPIs financieros ───────────────────────────────────────────────
        st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:14px;">
  <div class="snap-card">
    <div class="sc-label">VPN</div>
    <div class="sc-val" style="color:{kc5};">${vpn5:,.0f}</div>
    <div class="sc-sub">MXN sobre equity</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">TIR equity</div>
    <div class="sc-val" style="color:#22d3ee;">{tir5_str}</div>
    <div class="sc-sub">vs {p5_wacc:.1f}% WACC</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">LCOS</div>
    <div class="sc-val" style="color:{'#4ade80' if lcos5 < p5_precio_ppa else '#f87171'};">${lcos5:.3f}</div>
    <div class="sc-sub">MXN/kWh ciclado</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Payback equity</div>
    <div class="sc-val">{pb5_s_str}</div>
    <div class="sc-sub">nominal</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Valor residual</div>
    <div class="sc-val">${valor_residual5:,.0f}</div>
    <div class="sc-sub">MXN post-contrato</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Vida por ciclos</div>
    <div class="sc-val">{p5_vida_ciclos_y:.1f} a</div>
    <div class="sc-sub">{"Reemplazo en año " + str(p5_año_reemplazo) if p5_año_reemplazo <= p5_plazo else "Sin reemplazo en plazo"}</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Gráficas ───────────────────────────────────────────────────────
        g5c1, g5c2 = st.columns(2, gap="medium")

        with g5c1:
            st.markdown('<div class="section-header">Flujos anuales desarrollador</div>', unsafe_allow_html=True)
            colors5 = [AMBER if v >= 0 else ROSE for v in fn5_y]
            fig5_fn = go.Figure()
            fig5_fn.add_trace(go.Bar(
                x=years5, y=fn5_y, name="Flujo neto equity",
                marker_color=colors5, opacity=0.9,
                hovertemplate="<b>Año %{x}</b><br>Flujo neto: $%{y:,.0f} MXN<extra></extra>",
            ))
            fig5_fn.add_trace(go.Scatter(
                x=years5, y=ing5_y, name="Ingresos PPA",
                mode="lines", line=dict(color=TEAL, width=2, dash="dot"),
            ))
            if p5_año_reemplazo <= p5_plazo:
                fig5_fn.add_vline(x=p5_año_reemplazo, line_color=ROSE, line_dash="dash",
                                  annotation_text="Reemplazo celdas",
                                  annotation_font=dict(color=ROSE, size=10))
            lay5fn = copy.deepcopy(PLOT_LAYOUT)
            lay5fn.update({"height": 300,
                "yaxis": dict(title="MXN", gridcolor="#1e2230", tickformat=","),
                "xaxis": dict(title="Año", gridcolor="#1e2230", tickmode="linear"),
                "legend": dict(orientation="h", y=1.14, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=10, r=10, t=50, b=40)})
            fig5_fn.update_layout(**lay5fn)
            st.plotly_chart(fig5_fn, use_container_width=True)

        with g5c2:
            st.markdown('<div class="section-header">Perspectiva cliente: PPA vs CFE</div>', unsafe_allow_html=True)
            fig5_cl = go.Figure()
            fig5_cl.add_trace(go.Bar(
                x=years5, y=cfe5_y, name="Costo equivalente CFE",
                marker_color=ROSE, opacity=0.7,
                hovertemplate="<b>Año %{x}</b><br>CFE: $%{y:,.0f} MXN<extra></extra>",
            ))
            fig5_cl.add_trace(go.Bar(
                x=years5, y=pago5_y, name="Pago PPA",
                marker_color=TEAL, opacity=0.85,
                hovertemplate="<b>Año %{x}</b><br>PPA: $%{y:,.0f} MXN<extra></extra>",
            ))
            fig5_cl.add_trace(go.Scatter(
                x=years5, y=ahorro5_y, name="Ahorro cliente",
                mode="lines+markers", line=dict(color=AMBER, width=2),
                hovertemplate="<b>Año %{x}</b><br>Ahorro: $%{y:,.0f} MXN<extra></extra>",
            ))
            lay5cl = copy.deepcopy(PLOT_LAYOUT)
            lay5cl.update({"height": 300, "barmode": "group",
                "yaxis": dict(title="MXN", gridcolor="#1e2230", tickformat=","),
                "xaxis": dict(title="Año", gridcolor="#1e2230", tickmode="linear"),
                "legend": dict(orientation="h", y=1.14, x=0.5, xanchor="center", bgcolor="rgba(0,0,0,0)"),
                "margin": dict(l=10, r=10, t=50, b=40)})
            fig5_cl.update_layout(**lay5cl)
            st.plotly_chart(fig5_cl, use_container_width=True)

        # ── Tabla año a año ────────────────────────────────────────────────
        st.markdown('<div class="section-header">Tabla financiera año a año</div>', unsafe_allow_html=True)
        df5 = pd.DataFrame({
            "Año":                 years5,
            "kWh desc.":           [f"{v:,.0f}" for v in gen5_y],
            "Precio PPA ($/kWh)":  [f"${v:.4f}" for v in prec5_y],
            "Ingreso PPA (MXN)":   [f"${v:,.0f}" for v in ing5_y],
            "O&M (MXN)":           [f"${v:,.0f}" for v in om5_y],
            "Deuda (MXN)":         [f"${v:,.0f}" for v in deu5_y],
            "Reemplazo (MXN)":     [f"${v:,.0f}" for v in rep5_y],
            "Flujo neto (MXN)":    [f"${v:,.0f}" for v in fn5_y],
            "Flujo desc. (MXN)":   [f"${v:,.0f}" for v in fd5_y],
            "CFE equiv. (MXN)":    [f"${v:,.0f}" for v in cfe5_y],
            "Ahorro cliente (MXN)":[f"${v:,.0f}" for v in ahorro5_y],
        })
        st.dataframe(df5, use_container_width=True, hide_index=True)

        # ── Totales y criterio de decisión ─────────────────────────────────
        st.markdown(f"""
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-top:12px;">
  <div class="snap-card">
    <div class="sc-label">Ingreso total desarrollador</div>
    <div class="sc-val" style="color:#f1f5f9;">${sum(ing5_y):,.0f}</div>
    <div class="sc-sub">MXN en {p5_plazo} años</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Ahorro total cliente</div>
    <div class="sc-val" style="color:#4ade80;">${ahorro5_tot:,.0f}</div>
    <div class="sc-sub">MXN en {p5_plazo} años vs CFE</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">Descuento PPA vs CFE hoy</div>
    <div class="sc-val" style="color:{'#4ade80' if desc_vs_cfe < 0 else '#f87171'};">{desc_vs_cfe:+.1f}%</div>
    <div class="sc-sub">año 1 · ${p5_precio_ppa:.4f} vs ${p5_tarifa_cfe:.4f}/kWh</div>
  </div>
  <div class="snap-card">
    <div class="sc-label">VPN final</div>
    <div class="sc-val" style="color:{kc5};">${vpn5:,.0f}</div>
    <div class="sc-sub">MXN sobre equity</div>
  </div>
</div>
""", unsafe_allow_html=True)

        if viable5 and vpn5 > 0 and (tir5 or 0) > p5_wacc + 3:
            st.success(f"🟢 **Proyecto muy atractivo** — Precio PPA ${p5_precio_ppa:.4f} > mínimo ${precio_min5:.4f} · VPN positivo · TIR {tir5_str}")
        elif viable5 and vpn5 > 0:
            st.success(f"🟡 **Proyecto viable** — margen ajustado sobre precio mínimo")
        elif not viable5:
            st.error(f"🔴 **Precio PPA insuficiente** — mínimo requerido: {pm5_str}. Ajusta precio, plazo o estructura financiera.")
        else:
            st.warning("⚠️ Precio viable pero VPN negativo — revisar CAPEX o plazo.")

        st.markdown(f"""
<div class="info-box">
  <b>Metodología BESS PPA:</b>
  Dimensionamiento con curva de Wöhler LFP · CAPEX desglosado (celdas + PCS + BOS) ·
  Ingresos = kWh_descargados × precio_PPA (con degradación y escalador) ·
  Reemplazo de celdas en año {p5_año_reemplazo} ({LFP_PARAMS["costo_reemplazo_pct"]*100:.0f}% del CAPEX de celdas) si cae en el plazo ·
  LCOS = (CAPEX + PV_OPEX + PV_Reemplazo) / PV_kWh_ciclados ·
  Valor residual por suma finita de Gordon post-contrato ·
  Precio mínimo viable por bisección (VPN = 0)
</div>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# ═════════════════════════════════════════════════════════════════════════════
# TAB 6 — SISTEMA HÍBRIDO  (reemplaza Optimización 8760h)
# Motor: CHP + PV + BESS · Tarifas GDMTH reales · Despacho hora a hora
# ═════════════════════════════════════════════════════════════════════════════

# ── Helpers internos del tab ──────────────────────────────────────────────────

def _sh(txt):
    """Section header reutilizable."""
    return st.markdown(f'<div class="section-header">{txt}</div>', unsafe_allow_html=True)


def _info(txt):
    return st.markdown(f'<div class="info-box">{txt}</div>', unsafe_allow_html=True)


def _nasa_box(txt):
    return st.markdown(f'<div class="nasa-box">{txt}</div>', unsafe_allow_html=True)


def _warn(txt):
    return st.markdown(f'<div class="warn-box">{txt}</div>', unsafe_allow_html=True)


# ── Conversión gas USD/MMBTU → MXN/kWh_th ────────────────────────────────────
def gas_usdmmbtu_to_mxn_kwh(usd_mmbtu: float, tc: float) -> float:
    """1 MMBTU = 293.07 kWh_th"""
    return usd_mmbtu * tc / 293.07


# ── Construcción de máscara de horas críticas ─────────────────────────────────
def build_critical_mask(demand_df: pd.DataFrame, bloques: list) -> np.ndarray:
    """
    demand_df: DataFrame con columnas Fecha(datetime) y Hora(1-24).
    bloques: list of dict con keys dias(set 0-6), meses(set 1-12), h_ini(1-24), h_fin(1-24), etiqueta.
    Retorna array booleano len(df) y array de etiquetas (str o '').
    """
    n = len(demand_df)
    mask   = np.zeros(n, dtype=bool)
    labels = np.full(n, '', dtype=object)

    for blq in bloques:
        dias   = blq['dias']    # set de weekday 0=lun
        meses  = blq['meses']   # set 1-12
        h_ini  = blq['h_ini']   # 1-24
        h_fin  = blq['h_fin']   # 1-24
        etq    = blq['etiqueta']

        for i, row in demand_df.iterrows():
            fecha = pd.Timestamp(row['Fecha'])
            hora  = int(row['Hora'])
            if (fecha.weekday() in dias and
                    fecha.month in meses and
                    h_ini <= hora <= h_fin):
                mask[i]   = True
                labels[i] = etq

    return mask, labels


# ── Motor principal de simulación GDMTH ──────────────────────────────────────
def simulate_gdmth(
    demand_arr:   np.ndarray,    # kW por hora
    pv_arr:       np.ndarray,    # kW por hora (ya calculado)
    periodo_arr:  np.ndarray,    # 'B','I','P' por hora
    mes_arr:      np.ndarray,    # mes 1-12 por hora
    key_arr:      np.ndarray,    # 'YYYY-M' por hora para cargo demanda
    tarifas_dict: dict,          # {key: {B,I,P,dist,cap,trans}}
    # CHP
    chp_kw:       float,
    chp_disp:     float,         # factor disponibilidad 0-1
    chp_var_cost: float,         # MXN/kWh_e (gas/eff + O&M)
    chp_fijo:     bool,          # True=fijo, False=sigue carga
    include_chp:  bool,
    # PV
    include_pv:   bool,
    # BESS
    include_bess: bool,
    bess_kwh:     float,
    bess_kw:      float,
    soc_reserva:  float,         # kWh intocables en modo normal
    bess_eff:     float,         # one-way
    bess_dod_min: float,         # SOC mínimo absoluto kWh
    discharge_periods: set,      # {'P','I','B'}
    charge_from_grid:  bool,
    lookahead_h:       int,
    chp_excess_charge: bool,     # cargar BESS con excedente CHP
    # Horas críticas
    critical_mask:  np.ndarray,  # bool array
    critical_soc_min: float,     # SOC mínimo en horas críticas (kWh)
    # Sustitución
    sust_arr:  np.ndarray,       # bool — horas sustituidas
    orig_arr:  np.ndarray,       # demanda original
) -> dict:
    """
    Motor de despacho hora a hora con lógica completa:
    - CHP sigue carga o fijo según configuración
    - PV solar-first
    - BESS con reserva de contingencia, prioridad P→I, lookahead
    - Horas críticas con descarga máxima
    - Tres fuentes de carga BESS: solar, CHP excedente, red Base
    - Cargo de demanda calculado por mes sobre pico de CFE
    """
    n = len(demand_arr)
    chp_despacho = chp_kw * chp_disp if include_chp else 0.0

    # ── Pre-calcular lookahead de déficit en P e I ────────────────────────────
    future_pi_deficit = np.zeros(n)
    if include_bess and charge_from_grid and lookahead_h > 0:
        for t in range(n):
            end   = min(t + 1 + lookahead_h, n)
            dw    = demand_arr[t+1:end]
            pv_w  = pv_arr[t+1:end]
            per_w = periodo_arr[t+1:end]
            if len(dw) == 0:
                continue
            chp_w = np.minimum(chp_despacho, dw) if include_chp else np.zeros(len(dw))
            net_w = dw - chp_w - pv_w
            mask  = np.array([p in ('P', 'I') for p in per_w])
            if mask.sum() > 0:
                future_pi_deficit[t] = np.sum(np.maximum(0, net_w[mask]))

    # ── Estado inicial BESS ───────────────────────────────────────────────────
    soc     = min(bess_kwh, soc_reserva + (bess_kwh - soc_reserva) * 0.5) if include_bess else 0.0
    soc_max = bess_kwh
    soc_abs_min = bess_dod_min   # límite absoluto (DOD)

    # ── Acumuladores ─────────────────────────────────────────────────────────
    cfe_e = cfe_t = chp_f = 0.0
    cfe_kwh = chp_kwh = pv_kwh = curtail = 0.0
    bess_ch_solar = bess_ch_chp = bess_ch_red = 0.0
    bess_dis_P = bess_dis_I = bess_dis_B = bess_dis_crit = 0.0
    peak_by_month = {}

    rows = []

    for t in range(n):
        demand  = demand_arr[t]
        pv_gen  = pv_arr[t] if include_pv else 0.0
        periodo = periodo_arr[t]
        key     = key_arr[t]
        is_crit = bool(critical_mask[t])

        # ── Generación CHP ────────────────────────────────────────────────────
        if include_chp:
            if chp_fijo:
                chp_gen = chp_despacho if demand > 0 else 0.0
            else:
                chp_gen = min(chp_despacho, demand)
        else:
            chp_gen = 0.0

        net = demand - chp_gen - pv_gen
        bc_solar = bc_chp = bc_red = bd = curtail_h = 0.0
        fuente_carga = ''

        soc_min_efectivo = critical_soc_min if is_crit else soc_reserva

        if net < 0:
            # ── Excedente de generación ───────────────────────────────────────
            excedente = -net

            if include_bess:
                # 1. Cargar con excedente solar (PV > 0 y hay excedente)
                if pv_gen > 0 and soc < soc_max:
                    # Fracción del excedente atribuible a PV
                    pv_exc = max(0.0, pv_gen - max(0.0, demand - chp_gen))
                    solar_charge = min(pv_exc, bess_kw, (soc_max - soc) / bess_eff)
                    solar_charge = max(0.0, solar_charge)
                    soc         += solar_charge * bess_eff
                    bc_solar    += solar_charge
                    excedente   -= solar_charge
                    fuente_carga = 'Solar'

                # 2. Cargar con excedente CHP si está habilitado
                if chp_excess_charge and excedente > 0 and soc < soc_max:
                    chp_charge = min(excedente, bess_kw, (soc_max - soc) / bess_eff)
                    chp_charge = max(0.0, chp_charge)
                    soc       += chp_charge * bess_eff
                    bc_chp    += chp_charge
                    excedente -= chp_charge
                    fuente_carga = 'CHP' if not bc_solar else 'Solar+CHP'

            curtail_h = max(0.0, excedente)
            curtail  += curtail_h
            cfe_purchase = 0.0

        else:
            # ── Déficit: primero intentar cargar desde red si aplica ──────────
            if include_bess and charge_from_grid and periodo == 'B':
                if soc < soc_reserva and future_pi_deficit[t] > 0:
                    red_charge = min(bess_kw * 0.5,
                                     (soc_reserva - soc) / bess_eff,
                                     net * 0.5)
                    red_charge = max(0.0, red_charge)
                    if red_charge > 0:
                        soc       += red_charge * bess_eff
                        bc_red    += red_charge
                        net       += red_charge  # aumenta compra a CFE
                        fuente_carga = 'Red-B'

            # ── Descarga BESS ─────────────────────────────────────────────────
            if include_bess:
                can_discharge = (is_crit or periodo in discharge_periods)
                if can_discharge and soc > soc_min_efectivo:
                    soc_disp  = max(0.0, soc - soc_min_efectivo)
                    discharge = min(net, bess_kw, soc_disp * bess_eff)
                    discharge = max(0.0, discharge)
                    if discharge > 0:
                        soc -= discharge / bess_eff
                        bd   = discharge
                        if is_crit:
                            bess_dis_crit += discharge
                        if periodo == 'P':
                            bess_dis_P += discharge
                        elif periodo == 'I':
                            bess_dis_I += discharge
                        else:
                            bess_dis_B += discharge

            cfe_purchase = max(0.0, net - bd)

        # ── Cargo de demanda (pico mensual CFE) ───────────────────────────────
        if key not in peak_by_month:
            try:
                tar = tarifas_dict.get(key, {})
                peak_by_month[key] = {
                    'peak': 0.0,
                    'dist': tar.get('dist', 0.0),
                    'cap':  tar.get('cap', 0.0),
                }
            except Exception:
                peak_by_month[key] = {'peak': 0.0, 'dist': 0.0, 'cap': 0.0}

        peak_by_month[key]['peak'] = max(peak_by_month[key]['peak'], cfe_purchase)

        # ── Tarifas de energía ────────────────────────────────────────────────
        try:
            tar       = tarifas_dict.get(key, {})
            t_energia = tar.get(periodo, tar.get('I', 0.0))
            t_trans   = tar.get('trans', 0.0)
        except Exception:
            t_energia = t_trans = 0.0

        cfe_e  += cfe_purchase * t_energia
        cfe_t  += cfe_purchase * t_trans
        chp_f  += chp_gen * chp_var_cost
        cfe_kwh+= cfe_purchase
        chp_kwh+= chp_gen
        pv_kwh += pv_gen
        bess_ch_solar += bc_solar
        bess_ch_chp   += bc_chp
        bess_ch_red   += bc_red

        rows.append({
            'Fecha':              str(demand_arr[t]) if False else '',  # placeholder
            '_t':                 t,
            'Periodo':            periodo,
            'Dato_Sustituido':    'SÍ — paro total' if sust_arr[t] else '',
            'Demanda_Original_kW': round(float(orig_arr[t]), 2),
            'GHI_Wm2':            0.0,   # se rellena después
            'Temp_C':             0.0,
            'Demanda_kW':         round(demand, 2),
            'CHP_kW':             round(chp_gen, 2),
            'CHP_Modo':           'Fijo' if (include_chp and chp_fijo) else 'Sigue carga',
            'PV_kW':              round(pv_gen, 2),
            'BESS_Carga_kW':      round(bc_solar + bc_chp + bc_red, 2),
            'BESS_Fuente_Carga':  fuente_carga,
            'BESS_Descarga_kW':   round(bd, 2),
            'BESS_SOC_kWh':       round(soc, 2),
            'BESS_SOC_pct':       round(soc / bess_kwh * 100, 1) if bess_kwh > 0 else 0.0,
            'Reserva_Contg_kWh':  round(min(soc, soc_reserva), 2),
            'Curtailment_kW':     round(curtail_h, 2),
            'Hora_Critica':       'SÍ' if is_crit else '',
            'CFE_kW':             round(cfe_purchase, 2),
            'Tarifa_Energia_kWh': round(t_energia, 4),
            'Tarifa_Trans_kWh':   round(t_trans, 4),
            'Costo_CFE_E':        round(cfe_purchase * t_energia, 2),
            'Costo_CFE_T':        round(cfe_purchase * t_trans, 2),
            'Costo_CHP':          round(chp_gen * chp_var_cost, 2),
            'Costo_Total_Hora':   round(cfe_purchase * (t_energia + t_trans) + chp_gen * chp_var_cost, 2),
        })

    # ── Cargo de demanda mensual ──────────────────────────────────────────────
    demand_cost = sum(v['peak'] * (v['dist'] + v['cap']) for v in peak_by_month.values())

    total_cost  = cfe_e + cfe_t + demand_cost + chp_f
    fp_energia  = (chp_kwh / chp_kw) / n if (include_chp and chp_kw > 0) else 0.0
    bess_ch_tot = bess_ch_solar + bess_ch_chp + bess_ch_red
    autogen_pct = (chp_kwh + pv_kwh - curtail) / max(demand_arr.sum(), 1) * 100

    return dict(
        # Financiero
        cfe_e=cfe_e, cfe_t=cfe_t, demand_cost=demand_cost,
        chp_f=chp_f, total_cost=total_cost,
        # Energía
        chp_kwh=chp_kwh, pv_kwh=pv_kwh, cfe_kwh=cfe_kwh,
        curtail=curtail, autogen_pct=autogen_pct,
        # BESS
        bess_ch_solar=bess_ch_solar, bess_ch_chp=bess_ch_chp,
        bess_ch_red=bess_ch_red, bess_ch_tot=bess_ch_tot,
        bess_dis_P=bess_dis_P, bess_dis_I=bess_dis_I,
        bess_dis_B=bess_dis_B, bess_dis_crit=bess_dis_crit,
        # CHP
        fp_energia=fp_energia,
        # Despacho horario
        rows=rows,
        peak_by_month=peak_by_month,
    )


# ── Generador de Excel de despacho ───────────────────────────────────────────
def build_dispatch_excel(
    rows:         list,
    demand_df:    pd.DataFrame,
    pv_kw_arr:    np.ndarray,
    ghi_arr:      np.ndarray,
    temp_arr:     np.ndarray,
    include_chp:  bool,
    include_pv:   bool,
    include_bess: bool,
    has_critical: bool,
    has_sust:     bool,
    proj_nombre:  str,
    chp_kw:       float,
    pv_kwp:       float,
    bess_kwh:     float,
    soc_reserva:  float,
) -> bytes:
    """Genera el Excel de despacho horario con formato completo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _fill(h):
        return PatternFill('solid', fgColor=h)

    thin = Side(style='thin',   color='2A2D3A')
    med  = Side(style='medium', color='475569')
    B_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    B_MED = Border(left=med,  right=med,  top=med,  bottom=med)

    def _fnt(size=9, bold=False, color='E2E8F0'):
        return Font(name='JetBrains Mono', size=size, bold=bold, color=color)

    AL_C = Alignment(horizontal='center', vertical='center')
    AL_R = Alignment(horizontal='right',  vertical='center')
    AL_L = Alignment(horizontal='left',   vertical='center')
    AL_W = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Colores dark mode
    BG_PAGE  = '0A0C10'
    BG_HDR1  = '1F3864'   # id
    BG_HDR_S = 'B8860B'   # solar
    BG_HDR_C = '1A5C20'   # CHP
    BG_HDR_P = 'C44A00'   # PV
    BG_HDR_B = '4A148C'   # BESS
    BG_HDR_D = '2C3E6B'   # demanda/red
    BG_HDR_T = '37474F'   # costos
    BG_CRIT  = 'FF6B00'   # horas críticas

    PER_BG   = {'B': '0B1C0B', 'I': '0B1323', 'P': '1C1600'}
    SUST_BG  = '1C0E00'

    def _soc_bg(soc, cap):
        if cap <= 0:
            return '111318'
        p = soc / cap
        if p >= 0.75: return '052E16'
        if p >= 0.55: return '0F2A10'
        if p >= 0.40: return '1C1A04'
        return '1F0A0A'

    # ── Definir columnas ──────────────────────────────────────────────────────
    COLS = [
        # (campo, h1, h2, ancho, fmt, grupo_bg)
        ('Fecha',               'Fecha',          '',                  12, 'YYYY-MM-DD',            BG_HDR1),
        ('Hora',                'Hora',            '',                   5, '0',                     BG_HDR1),
        ('Periodo',             'Periodo',         '',                   8, '@',                     BG_HDR1),
    ]

    if has_sust:
        COLS += [
            ('Dato_Sustituido',     'Dato',           'Sustituido',        14, '@',                     BG_HDR1),
            ('Demanda_Original_kW', 'Demanda',        'Original kW',       14, '#,##0.0',               BG_HDR1),
        ]

    COLS += [
        ('GHI_Wm2',             'GHI',            'W/m²',               9, '0.0',                   BG_HDR_S),
        ('Temp_C',              'Temp',            '°C',                  8, '0.0',                   BG_HDR_S),
        ('Demanda_kW',          'Demanda',         'kW',                 11, '#,##0.0',               BG_HDR_D),
    ]

    if include_chp:
        COLS += [
            ('CHP_kW',          'CHP',             'kW',                 10, '#,##0.0',               BG_HDR_C),
            ('CHP_Modo',        'CHP',             'Modo',                9, '@',                     BG_HDR_C),
        ]

    if include_pv:
        COLS += [
            ('PV_kW',           'FV Solar',        'kW',                 10, '#,##0.0',               BG_HDR_P),
        ]

    if include_bess:
        COLS += [
            ('BESS_Carga_kW',    'BESS',           'Carga kW',           11, '#,##0.0',               BG_HDR_B),
            ('BESS_Fuente_Carga','Fuente',          'Carga',             10, '@',                     BG_HDR_B),
            ('BESS_Descarga_kW', 'BESS',           'Descarga kW',        12, '#,##0.0',               BG_HDR_B),
            ('BESS_SOC_kWh',     'SOC',            'kWh',                10, '#,##0.0',               BG_HDR_B),
            ('BESS_SOC_pct',     'SOC',            '%',                   8, '0.0"%"',                BG_HDR_B),
            ('Reserva_Contg_kWh','Reserva',        'Contg. kWh',         12, '#,##0.0',               BG_HDR_B),
        ]

    COLS += [
        ('Curtailment_kW',      'Curtail',         'kW',                 10, '#,##0.0',               BG_HDR_P if include_pv else BG_HDR_D),
    ]

    if has_critical:
        COLS += [
            ('Hora_Critica',    'Hora',            'Crítica',             9, '@',                     BG_HDR1),
        ]

    COLS += [
        ('CFE_kW',              'CFE',             'Compra kW',           11, '#,##0.0',               BG_HDR_D),
        ('Tarifa_Energia_kWh',  'Tarifa',          'Energía $/kWh',       15, '"$"#,##0.0000',         BG_HDR_T),
        ('Tarifa_Trans_kWh',    'Tarifa',          'Trans $/kWh',         13, '"$"#,##0.0000',         BG_HDR_T),
        ('Costo_CFE_E',         'Costo CFE',       'Energía $',           13, '"$"#,##0;-;-',          BG_HDR_T),
        ('Costo_CFE_T',         'Costo CFE',       'Transporte $',        14, '"$"#,##0;-;-',          BG_HDR_T),
        ('Costo_CHP',           'Costo',           'CHP $',               11, '"$"#,##0;-;-',          BG_HDR_T),
        ('Costo_Total_Hora',    'Costo',           'Total $',             12, '"$"#,##0;-;-',          BG_HDR_T),
    ]

    campo_list = [c[0] for c in COLS]

    # ── Enriquecer rows con GHI, temp, Fecha, Hora ───────────────────────────
    n = len(rows)
    for i, r in enumerate(rows):
        r['GHI_Wm2']  = round(float(ghi_arr[i]),  1) if i < len(ghi_arr)  else 0.0
        r['Temp_C']   = round(float(temp_arr[i]),  1) if i < len(temp_arr) else 20.0
        r['Fecha']    = str(demand_df['Fecha'].iloc[i].date()) if i < len(demand_df) else ''
        r['Hora']     = int(demand_df['Hora'].iloc[i]) if i < len(demand_df) else i % 24 + 1
        r['PV_kW']    = round(float(pv_kw_arr[i]), 2) if (include_pv and i < len(pv_kw_arr)) else 0.0

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Despacho_Horario"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.freeze_panes = 'A3'
    ws.sheet_properties.tabColor = 'F59E0B'

    # Fondo global de la hoja
    for row in ws.iter_rows(min_row=1, max_row=n + 5, min_col=1, max_col=len(COLS)):
        for cell in row:
            cell.fill = _fill(BG_PAGE)

    # ── Fila 1: Título ────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(len(COLS))}1')
    c = ws['A1']
    titulo = (f"Despacho Hora a Hora  ·  {proj_nombre}  ·  "
              f"{'CHP ' + str(int(chp_kw)) + 'kW ' if include_chp else ''}"
              f"{'+ FV ' + str(int(pv_kwp)) + 'kWp ' if include_pv else ''}"
              f"{'+ BESS ' + str(int(bess_kwh)) + 'kWh' if include_bess else ''}")
    c.value     = titulo
    c.font      = Font(name='Inter', bold=True, size=11, color='F1F5F9')
    c.alignment = AL_L
    c.fill      = _fill('111318')
    ws.row_dimensions[1].height = 24

    # ── Fila 2: Encabezados ───────────────────────────────────────────────────
    ws.row_dimensions[2].height = 42
    for j, (campo, h1, h2, ancho, fmt, bg) in enumerate(COLS):
        col = j + 1
        ws.column_dimensions[get_column_letter(col)].width = ancho
        c = ws.cell(row=2, column=col, value=f"{h1}\n{h2}".strip())
        c.fill      = _fill(bg)
        c.font      = Font(name='Inter', bold=True, size=9, color='F1F5F9')
        c.alignment = AL_W
        c.border    = B_ALL

    # ── Filas de datos ────────────────────────────────────────────────────────
    for i, r in enumerate(rows):
        row_num  = i + 3
        periodo  = r.get('Periodo', 'B')
        is_sust  = bool(r.get('Dato_Sustituido', ''))
        is_crit  = bool(r.get('Hora_Critica', ''))
        soc_val  = r.get('BESS_SOC_kWh', 0.0)

        if is_sust:
            row_bg = SUST_BG
        elif is_crit:
            row_bg = '1A0E00'
        else:
            row_bg = PER_BG.get(periodo, '111318')

        ws.row_dimensions[row_num].height = 13

        for j, (campo, h1, h2, ancho, fmt, bg) in enumerate(COLS):
            col = j + 1
            val = r.get(campo, None)
            c   = ws.cell(row=row_num, column=col, value=val)
            c.number_format = fmt
            c.border = Border(left=thin, right=thin)
            c.alignment = AL_C if campo in ('Fecha', 'Hora', 'Periodo',
                                              'BESS_Fuente_Carga', 'CHP_Modo',
                                              'Dato_Sustituido', 'Hora_Critica') else AL_R

            # ── Color de fondo por columna ────────────────────────────────────
            if campo == 'BESS_SOC_kWh':
                c.fill = _fill(_soc_bg(soc_val, bess_kwh))
                c.font = _fnt(size=9, bold=(soc_val <= soc_reserva * 1.05))
            elif campo == 'BESS_SOC_pct':
                c.fill = _fill(_soc_bg(soc_val, bess_kwh))
                c.font = _fnt(size=9)
            elif campo == 'Reserva_Contg_kWh':
                ok = val is not None and val >= soc_reserva * 0.99
                c.fill = _fill('052E16' if ok else '1F0A0A')
                c.font = _fnt(size=9, color='4ADE80' if ok else 'F87171')
            elif campo == 'BESS_Descarga_kW':
                p = r.get('Periodo', 'B')
                v = val or 0
                if r.get('Hora_Critica') and v > 0:
                    c.fill = _fill('FF6B00'); c.font = _fnt(size=9, bold=True, color='FFFFFF')
                elif p == 'P' and v > 0:
                    c.fill = _fill('2D0047'); c.font = _fnt(size=9, bold=True, color='CE93D8')
                elif p == 'I' and v > 0:
                    c.fill = _fill('1A0A2E'); c.font = _fnt(size=9, color='9C5FD6')
                else:
                    c.fill = _fill('111318'); c.font = _fnt(size=9, color='475569')
            elif campo == 'BESS_Fuente_Carga':
                v = val or ''
                colors_fuente = {
                    'Solar': ('1C1000', 'F59E0B'),
                    'CHP':   ('0B1C0B', '4ADE80'),
                    'Red-B': ('0B1323', '60A5FA'),
                    'Solar+CHP': ('121000', 'FBBF24'),
                }
                bg_f, fg_f = colors_fuente.get(v, ('111318', '475569'))
                c.fill = _fill(bg_f); c.font = _fnt(size=9, bold=bool(v), color=fg_f)
            elif campo == 'BESS_Carga_kW':
                v = val or 0
                c.fill = _fill('1C1000' if v > 0 else '111318')
                c.font = _fnt(size=9, color='F59E0B' if v > 0 else '475569')
            elif campo == 'Dato_Sustituido':
                c.fill = _fill('1C0A00' if is_sust else '111318')
                c.font = _fnt(size=9, bold=is_sust,
                               color='F97316' if is_sust else '1E293B')
            elif campo == 'Hora_Critica':
                c.fill = _fill('1C0E00' if is_crit else '111318')
                c.font = _fnt(size=9, bold=is_crit,
                               color='FF6B00' if is_crit else '1E293B')
            elif campo in ('Costo_CFE_E', 'Costo_CFE_T', 'Costo_CHP', 'Costo_Total_Hora',
                           'Tarifa_Energia_kWh', 'Tarifa_Trans_kWh'):
                c.fill = _fill('0D1117'); c.font = _fnt(size=9, color='94A3B8')
            else:
                c.fill = _fill(row_bg); c.font = _fnt(size=9)

    # ── Fila de totales ───────────────────────────────────────────────────────
    r_tot = n + 3
    ws.row_dimensions[r_tot].height = 20
    for j in range(len(COLS)):
        c = ws.cell(row=r_tot, column=j + 1)
        c.fill   = _fill('111318')
        c.font   = Font(name='Inter', bold=True, size=9, color='F59E0B')
        c.border = B_MED
        c.alignment = AL_R

    ws.cell(row=r_tot, column=1, value='TOTALES').alignment = AL_C

    SUM_COLS = {'Demanda_kW', 'Demanda_Original_kW', 'GHI_Wm2',
                'CHP_kW', 'PV_kW', 'BESS_Carga_kW', 'BESS_Descarga_kW',
                'Curtailment_kW', 'CFE_kW',
                'Costo_CFE_E', 'Costo_CFE_T', 'Costo_CHP', 'Costo_Total_Hora'}
    AVG_COLS = {'BESS_SOC_pct', 'Tarifa_Energia_kWh', 'Tarifa_Trans_kWh', 'Temp_C'}
    MIN_COLS = {'Reserva_Contg_kWh', 'BESS_SOC_kWh'}

    for j, (campo, *_) in enumerate(COLS):
        col = j + 1
        cl  = get_column_letter(col)
        r3  = 3
        if campo in SUM_COLS:
            c = ws.cell(row=r_tot, column=col,
                        value=f'=SUM({cl}{r3}:{cl}{r_tot - 1})')
            c.number_format = COLS[j][4]
        elif campo in AVG_COLS:
            c = ws.cell(row=r_tot, column=col,
                        value=f'=AVERAGE({cl}{r3}:{cl}{r_tot - 1})')
            c.number_format = COLS[j][4]
        elif campo in MIN_COLS:
            c = ws.cell(row=r_tot, column=col,
                        value=f'=MIN({cl}{r3}:{cl}{r_tot - 1})')
            c.number_format = '"Mín: "#,##0.0'
        else:
            continue

        c.fill   = _fill('111318')
        c.font   = Font(name='Inter', bold=True, size=9, color='F59E0B')
        c.border = B_MED
        c.alignment = AL_R

    # FP CHP en celda Hora
    if include_chp:
        chp_col_idx = next(j + 1 for j, (c, *_) in enumerate(COLS) if c == 'CHP_kW')
        cl_chp = get_column_letter(chp_col_idx)
        fp_cell = ws.cell(row=r_tot, column=2,
                          value=f'=TEXT({cl_chp}{r_tot}/({chp_kw}*{n}),"0.0%")&" FP CHP"')
        fp_cell.font      = Font(name='Inter', bold=True, size=9, color='F59E0B')
        fp_cell.alignment = AL_C

    # SOC promedio si hay BESS
    if include_bess:
        soc_col_idx = next(j + 1 for j, (c, *_) in enumerate(COLS) if c == 'BESS_SOC_kWh')
        cl_soc = get_column_letter(soc_col_idx)
        sc = ws.cell(row=r_tot, column=soc_col_idx,
                     value=f'=TEXT(AVERAGE({cl_soc}3:{cl_soc}{r_tot - 1})/({bess_kwh}),"0%")&" SOC prom"')
        sc.fill   = _fill('111318')
        sc.font   = Font(name='Inter', bold=True, size=9, color='F59E0B')
        sc.border = B_MED
        sc.alignment = AL_C

    # ── Leyenda ───────────────────────────────────────────────────────────────
    r_leg = r_tot + 2
    leyendas = [
        [('Base',  PER_BG['B']), ('Intermedio', PER_BG['I']), ('Punta', PER_BG['P']),
         ('Sustituido', SUST_BG), ('Hora crítica', '1A0E00')],
        [('BESS desc. Punta', '2D0047'), ('BESS desc. Interm.', '1A0A2E'),
         ('Carga solar', '1C1000'), ('Carga CHP', '0B1C0B'), ('Carga Red-B', '0B1323')],
        [('SOC >75%', '052E16'), ('SOC 55-75%', '0F2A10'),
         ('SOC 40-55%', '1C1A04'), ('SOC <40% (reserva)', '1F0A0A')],
    ]
    labels_fila = ['Periodos:', 'BESS:', 'SOC:']
    for r_off, items in enumerate(leyendas):
        ws.cell(row=r_leg + r_off, column=1,
                value=labels_fila[r_off]).font = Font(name='Inter', size=8, color='475569')
        ws.cell(row=r_leg + r_off, column=1).fill = _fill(BG_PAGE)
        for k, (txt, bg) in enumerate(items):
            c = ws.cell(row=r_leg + r_off, column=2 + k, value=txt)
            c.fill      = _fill(bg)
            c.font      = Font(name='Inter', size=8, color='94A3B8')
            c.alignment = AL_C
            c.border    = B_ALL
        ws.row_dimensions[r_leg + r_off].height = 15

    # Autofilter
    ws.auto_filter.ref = f'A2:{get_column_letter(len(COLS))}2'

    # ── Guardar ───────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# TAB 6 UI
# ═════════════════════════════════════════════════════════════════════════════
with tab6:
    st.markdown('<div class="app-title">⚡ Sistema Híbrido</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="app-sub">CHP + FV + BESS · Tarifas GDMTH reales · '
        'Despacho hora a hora · Multi-escenario</div>',
        unsafe_allow_html=True,
    )

    # ── Session state ─────────────────────────────────────────────────────────
    for _key in ['sh_escenarios', 'sh_demand_df', 'sh_tmy_df', 'sh_tmy_meta',
                 'sh_bloques_crit']:
        if _key not in st.session_state:
            st.session_state[_key] = {} if _key == 'sh_escenarios' else (
                [] if _key == 'sh_bloques_crit' else None)

    # ── Layout principal: inputs izq | outputs der ────────────────────────────
    col_inp, col_out = st.columns([1, 1.9], gap='large')

    # ══════════════════════════════════════════════════════════════════════════
    # PANEL IZQUIERDO — INPUTS
    # ══════════════════════════════════════════════════════════════════════════
    with col_inp:

        # ── Bloque 1: Proyecto ────────────────────────────────────────────────
        _sh('1 · Proyecto')
        sh_nombre  = st.text_input('Nombre del proyecto', value='Proyecto Híbrido', key='sh_nombre')
        sh_cliente = st.text_input('Cliente',             value='',                 key='sh_cliente')

        # ── Bloque 2: Archivo de demanda ──────────────────────────────────────
        _sh('2 · Archivo de demanda')
        _info('Columnas requeridas: <b>Fecha · Hora · PeriodoTarifario · Consumo (kWh)</b>'
              '<br>El periodo tarifario debe ser B, I o P por hora.')

        sh_file = st.file_uploader(
            'Cargar Excel de demanda', type=['xlsx', 'xls'],
            key='sh_demand_file',
        )

        if sh_file is not None:
            try:
                # Leer solo primeras 4 columnas — ignora columnas extra con formulas
                _df_raw = pd.read_excel(sh_file, usecols=range(4))
                _df_raw.columns = [str(c).strip() for c in _df_raw.columns]

                # Mapeo robusto: elimina espacios, parentesis y guiones
                _col_map = {}
                for col in _df_raw.columns:
                    cl = col.lower().replace(' ','').replace('(','').replace(')','').replace('_','').replace('-','')
                    if 'fecha' in cl:
                        _col_map[col] = 'Fecha'
                    elif cl == 'hora':
                        _col_map[col] = 'Hora'
                    elif 'periodo' in cl:
                        _col_map[col] = 'PeriodoTarifario'
                    elif 'demanda' in cl or 'consumo' in cl or cl.endswith('kwh'):
                        _col_map[col] = 'Consumo_kWh'
                _df_raw = _df_raw.rename(columns=_col_map)

                _required = {'Fecha', 'Hora', 'PeriodoTarifario', 'Consumo_kWh'}
                _missing  = _required - set(_df_raw.columns)
                if _missing:
                    st.error(
                        f'❌ Faltan columnas: {", ".join(_missing)} · '
                        f'Columnas detectadas: {list(_df_raw.columns)}'
                    )
                else:
                    _df_raw = _df_raw.dropna(subset=['Fecha']).copy()
                    _df_raw['Fecha'] = pd.to_datetime(_df_raw['Fecha'], errors='coerce')
                    _df_raw = _df_raw.dropna(subset=['Fecha']).copy()
                    _df_raw['Hora'] = pd.to_numeric(
                        _df_raw['Hora'], errors='coerce').fillna(1).astype(int)
                    _df_raw['Consumo_kWh'] = pd.to_numeric(
                        _df_raw['Consumo_kWh'], errors='coerce').fillna(0).clip(lower=0)
                    _df_raw['PeriodoTarifario'] = (
                        _df_raw['PeriodoTarifario'].astype(str).str.strip().str.upper()
                    )
                    _df_raw['PeriodoTarifario'] = _df_raw['PeriodoTarifario'].where(
                        _df_raw['PeriodoTarifario'].isin(['B','I','P']), 'B'
                    )
                    _df_raw = _df_raw.drop_duplicates(
                        subset=['Fecha','Hora'], keep='last').reset_index(drop=True)
                    st.session_state['sh_demand_df'] = _df_raw
                    n_rows = len(_df_raw)
                    d_pico = _df_raw['Consumo_kWh'].max()
                    d_avg  = _df_raw['Consumo_kWh'].mean()
                    d_sum  = _df_raw['Consumo_kWh'].sum()
                    _per   = _df_raw['PeriodoTarifario'].value_counts().to_dict()
                    st.success(
                        f'✅ {n_rows:,} horas · '
                        f'{_df_raw["Fecha"].min().date()} → {_df_raw["Fecha"].max().date()} · '
                        f'Pico: {d_pico:,.0f} kW · Media: {d_avg:,.0f} kW · '
                        f'Total: {d_sum/1e6:.2f} GWh · '
                        f'B:{_per.get("B",0)} I:{_per.get("I",0)} P:{_per.get("P",0)} h'
                    )
            except Exception as _e:
                st.error(f'Error al leer el archivo: {_e}')

        _demand_ok = st.session_state.get('sh_demand_df') is not None

        # ── Bloque 3: Datos solares NASA ──────────────────────────────────────
        _sh('3 · Datos solares NASA POWER')
        _nasa_box(
            f'🌍 Coordenadas activas: <b>Lat {lat:.4f} · Lon {lon:.4f}</b> · '
            f'Climatología NASA POWER 2005–{NASA_END}'
        )

        sh_year_ref = st.selectbox(
            'Año de referencia para perfil intradiario',
            [2023, 2022, 2021, 2020, 2019], index=0, key='sh_year_ref',
        )
        sh_noct   = st.number_input('NOCT del panel (°C)', 35.0, 55.0, 45.0, 0.5,
                                     key='sh_noct')
        sh_t_coef = st.number_input('Coef. térmico (%/°C)',
                                     -0.60, -0.20, -0.40, 0.01,
                                     format='%.2f', key='sh_tcoef')

        col_tmy_btn, _ = st.columns([1, 2])
        with col_tmy_btn:
            sh_tmy_btn = st.button('🌍 Construir TMY 8760h',
                                   type='primary', use_container_width=True,
                                   key='sh_tmy_btn')

        if sh_tmy_btn:
            _irr_ss = st.session_state.get('nasa_irradiance')
            if not _irr_ss or sum(_irr_ss) == 0:
                st.warning('⚠️ Carga primero la irradiancia NASA desde el sidebar.')
            else:
                with st.spinner('Construyendo TMY…'):
                    _temp_m, _ = get_nasa_monthly_temp(lat, lon)
                    _tmy_df, _tmy_meta = build_tmy_8760(
                        lat=lat, lon=lon,
                        irr_media_mensual=tuple(_irr_ss),
                        temp_media_mensual=tuple(_temp_m),
                        year_ref=sh_year_ref,
                    )
                if _tmy_df is None:
                    st.error(f'❌ {_tmy_meta}')
                else:
                    st.session_state['sh_tmy_df']   = _tmy_df
                    st.session_state['sh_tmy_meta'] = _tmy_meta
                    st.success(f'✅ {_tmy_meta}')

        _tmy_ok = st.session_state.get('sh_tmy_df') is not None
        if _tmy_ok:
            _tmy_meta = st.session_state.get('sh_tmy_meta', '')
            st.caption(_tmy_meta)

        # ── Bloque 4: CHP ─────────────────────────────────────────────────────
        _sh('4 · CHP — Motor de Combustión Interna')
        sh_use_chp = st.checkbox('Incluir CHP en el sistema', value=True, key='sh_use_chp')

        if sh_use_chp:
            sh_chp_kw   = st.number_input('Capacidad instalada (kW)', 100, 10000, 2500, 100, key='sh_chp_kw')
            sh_chp_disp = st.slider('Factor de disponibilidad (%)', 75, 99, 90, 1, key='sh_chp_disp')
            sh_chp_mode = st.radio('Modo de operación', ['Sigue carga', 'Fijo'], horizontal=True, key='sh_chp_mode')
            sh_chp_fijo = sh_chp_mode == 'Fijo'
            st.caption(f'Despacho efectivo: **{sh_chp_kw * sh_chp_disp / 100:,.0f} kW**')

            c4a, c4b = st.columns(2)
            with c4a:
                sh_chp_eff_e = st.number_input('Eficiencia eléctrica (%)', 30.0, 50.0, 40.0, 0.5, key='sh_chp_eff_e')
                sh_chp_eff_t = st.number_input('Eficiencia térmica (%)',   35.0, 55.0, 45.0, 0.5, key='sh_chp_eff_t')
            with c4b:
                sh_gas_price = st.number_input('Precio gas (USD/MMBTU)', 1.0, 15.0, 3.50, 0.10,
                                               format='%.2f', key='sh_gas_price')
                sh_tc        = st.number_input('Tipo de cambio (MXN/USD)', 10.0, 25.0, 17.50, 0.10,
                                               format='%.2f', key='sh_tc')
                sh_chp_om    = st.number_input('O&M variable (MXN/kWh_e)', 0.05, 0.30, 0.12, 0.01,
                                               format='%.3f', key='sh_chp_om')

            _gas_mxn_kwh = gas_usdmmbtu_to_mxn_kwh(sh_gas_price, sh_tc)
            _chp_var_cost = _gas_mxn_kwh / (sh_chp_eff_e / 100) + sh_chp_om
            st.caption(
                f'Gas: **${_gas_mxn_kwh:.4f}/kWh_th** → '
                f'Costo variable CHP: **${_chp_var_cost:.4f}/kWh_e** · '
                f'Eficiencia global: {sh_chp_eff_e + sh_chp_eff_t:.0f}%'
            )
            sh_chp_capex_kw = st.number_input('CAPEX (MXN/kW)', 10000, 30000, 18000, 500, key='sh_chp_capex')
        else:
            sh_chp_kw = sh_chp_disp = sh_chp_eff_e = sh_chp_eff_t = 0
            sh_chp_fijo = False; sh_gas_price = 0; sh_tc = 17.5; sh_chp_om = 0
            _chp_var_cost = 0.0; sh_chp_capex_kw = 0

        # ── Bloque 5: PV ──────────────────────────────────────────────────────
        _sh('5 · Sistema Fotovoltaico')
        sh_use_pv = st.checkbox('Incluir PV en el sistema', value=True, key='sh_use_pv')

        if sh_use_pv:
            sh_pv_kwp    = st.number_input('Capacidad pico (kWp)', 50, 20000, 500, 50, key='sh_pv_kwp')
            sh_pv_pr     = st.slider('Performance Ratio (PR)', 0.60, 0.95, float(effective_pr),
                                     0.01, format='%.2f', key='sh_pv_pr')
            sh_pv_capex  = st.number_input('CAPEX (MXN/kWp)', 5000, 20000, 12000, 500, key='sh_pv_capex')
            if _tmy_ok:
                _ghi_anual = st.session_state['sh_tmy_df']['irradiance_Wm2'].sum() / 1000
                _yield_est = _ghi_anual * sh_pv_pr
                st.caption(
                    f'GHI anual sitio: **{_ghi_anual:.0f} kWh/m²** · '
                    f'Rendimiento estimado: **{_yield_est:.0f} kWh/kWp/año** · '
                    f'Generación: **{sh_pv_kwp * _yield_est / 1e6:.2f} GWh/año**'
                )
        else:
            sh_pv_kwp = 0; sh_pv_pr = 0.80; sh_pv_capex = 0

        # ── Bloque 6: BESS ────────────────────────────────────────────────────
        _sh('6 · Sistema de Almacenamiento (BESS)')
        sh_use_bess = st.checkbox('Incluir BESS en el sistema', value=True, key='sh_use_bess')

        if sh_use_bess:
            c6a, c6b = st.columns(2)
            with c6a:
                sh_bess_kwh  = st.number_input('Capacidad energética (kWh)', 50, 20000, 500, 50, key='sh_bess_kwh')
                sh_bess_ratio = st.radio('Ratio de potencia', ['C/4', 'C/2', 'C/1'],
                                          index=1, horizontal=True, key='sh_bess_ratio')
                _ratio_map = {'C/4': 4, 'C/2': 2, 'C/1': 1}
                sh_bess_kw = sh_bess_kwh / _ratio_map[sh_bess_ratio]
                st.caption(f'Potencia BESS: **{sh_bess_kw:.0f} kW**')

            with c6b:
                sh_bess_res_pct = st.slider('Reserva contingencia (%)', 10, 60, 25, 5, key='sh_bess_res')
                sh_bess_res_kwh = sh_bess_kwh * sh_bess_res_pct / 100
                sh_bess_eff     = st.slider('Eficiencia one-way (%)', 85, 97, 92, 1, key='sh_bess_eff') / 100
                sh_bess_dod     = st.slider('DOD máximo (%)', 70, 95, 80, 5, key='sh_bess_dod')
                sh_bess_dod_min = sh_bess_kwh * (1 - sh_bess_dod / 100)
                sh_bess_capex   = st.number_input('CAPEX (MXN/kWh)', 4000, 15000, 8000, 500, key='sh_bess_capex')

            # Autonomía estimada (usa valor previo del session_state si aún no se renderizó Bloque 7)
            if sh_use_chp:
                _carga_esencial = st.session_state.get('sh_carga_esencial', 2750)
                _diferencial    = max(0, _carga_esencial - sh_chp_kw * sh_chp_disp / 100)
                if _diferencial > 0:
                    _autonomia = sh_bess_res_kwh / _diferencial * 60
                    st.caption(f'Reserva: **{sh_bess_res_kwh:.0f} kWh** · '
                               f'Autonomía estimada en contingencia: **{_autonomia:.0f} min**')
                else:
                    st.caption(f'Reserva: **{sh_bess_res_kwh:.0f} kWh** · '
                               'CHP cubre la carga esencial completa')
            else:
                st.caption(f'Reserva de contingencia: **{sh_bess_res_kwh:.0f} kWh**')

            st.markdown('**Estrategia de descarga**')
            sh_dis_per = st.multiselect(
                'Descargar en periodos', ['P', 'I', 'B'],
                default=['P', 'I'], key='sh_dis_per',
            )
            sh_lookahead = st.slider('Lookahead (horas)', 0, 48, 24, 1, key='sh_lookahead')

            st.markdown('**Fuentes de carga BESS**')
            st.caption('Prioridad 1 — Excedente PV (siempre activo)')
            sh_chp_excess = st.checkbox('Prioridad 2 — Excedente CHP',
                                         value=True, disabled=not sh_use_chp,
                                         key='sh_chp_excess')
            sh_grid_charge = st.checkbox('Prioridad 3 — Red CFE en Base (si SOC < reserva)',
                                          value=False, key='sh_grid_charge')
        else:
            sh_bess_kwh = 0; sh_bess_kw = 0; sh_bess_res_kwh = 0
            sh_bess_eff = 0.92; sh_bess_dod_min = 0; sh_bess_capex = 0
            sh_dis_per = []; sh_lookahead = 0
            sh_chp_excess = False; sh_grid_charge = False

        # ── Bloque 7: Resiliencia ─────────────────────────────────────────────
        _sh('7 · Carga esencial y resiliencia')
        sh_carga_esencial = st.number_input(
            'Carga esencial (kW)', 0, 10000, 2750, 50, key='sh_carga_esencial',
        )
        sh_margen_seg = st.slider('Margen de seguridad (%)', 0, 25, 10, 5, key='sh_margen')
        sh_carga_diseno = sh_carga_esencial * (1 + sh_margen_seg / 100)
        st.caption(f'Carga de diseño: **{sh_carga_diseno:,.0f} kW**')

        c7a, c7b = st.columns(2)
        with c7a:
            sh_perd_dist = st.number_input('Pérdidas distribución (MXN/año)',
                                            0, 50_000_000, 5_000_000, 500_000,
                                            format='%d', key='sh_perd_dist')
        with c7b:
            sh_perd_prod = st.number_input('Pérdidas prod. no prod. (MXN/año)',
                                            0, 100_000_000, 8_000_000, 500_000,
                                            format='%d', key='sh_perd_prod')
        sh_total_res = sh_perd_dist + sh_perd_prod
        st.caption(f'Pérdidas totales por apagones: **${sh_total_res:,.0f} MXN/año**')

        # ── Bloque 8: Datos atípicos ──────────────────────────────────────────
        _sh('8 · Datos atípicos')
        sh_sust = st.checkbox('Detectar y sustituir días de paro total', value=True, key='sh_sust')

        if sh_sust:
            sh_umbral = st.slider('Umbral de detección (% de la media)', 5, 35, 20, 5, key='sh_umbral')
            sh_metodo = st.radio(
                'Método de sustitución',
                ['Promedio Lun–Sáb', 'Mismo día de semana', 'Semana anterior + siguiente'],
                key='sh_metodo',
            )
            if _demand_ok:
                _df_tmp = st.session_state['sh_demand_df']
                _media  = _df_tmp['Consumo_kWh'].mean()
                _thresh = _media * sh_umbral / 100
                _dias_anom = _df_tmp.groupby(
                    _df_tmp['Fecha'].dt.date
                )['Consumo_kWh'].mean()
                _dias_excl = _dias_anom[_dias_anom < _thresh]
                st.caption(
                    f'Media global: **{_media:,.0f} kW** · Umbral: **{_thresh:,.0f} kW** · '
                    f'Días detectados: **{len(_dias_excl)}**'
                )

        # ── Bloque 9: Tarifas ─────────────────────────────────────────────────
        _sh('9 · Tarifas')
        sh_modalidad = st.radio(
            'Modalidad del contrato',
            ['Tarifa regulada GDMTH', 'Suministro calificado'],
            horizontal=True, key='sh_modalidad',
        )

        if sh_modalidad == 'Tarifa regulada GDMTH':
            _info('Ingresa las tarifas mensuales. '
                  '<b>B</b>=Base · <b>I</b>=Intermedio · <b>P</b>=Punta · '
                  '<b>Dist/Cap</b>=cargo demanda · <b>Trans</b>=transporte+CENACE+SCnMEM.')

            _tar_defaults = pd.DataFrame({
                'Mes':   MONTHS,
                'B ($/kWh)':    [0.9795]*12,
                'I ($/kWh)':    [1.7486]*12,
                'P ($/kWh)':    [2.0549]*12,
                'Dist ($/kW)':  [38.50]*12,
                'Cap ($/kW)':   [58.20]*12,
                'Trans ($/kWh)':[0.2063]*12,
            })
            sh_tar_df = st.data_editor(
                _tar_defaults,
                column_config={
                    'Mes': st.column_config.TextColumn(disabled=True),
                    'B ($/kWh)':     st.column_config.NumberColumn(format='$%.4f', step=0.0001),
                    'I ($/kWh)':     st.column_config.NumberColumn(format='$%.4f', step=0.0001),
                    'P ($/kWh)':     st.column_config.NumberColumn(format='$%.4f', step=0.0001),
                    'Dist ($/kW)':   st.column_config.NumberColumn(format='$%.2f', step=0.10),
                    'Cap ($/kW)':    st.column_config.NumberColumn(format='$%.2f', step=0.10),
                    'Trans ($/kWh)': st.column_config.NumberColumn(format='$%.4f', step=0.0001),
                },
                hide_index=True, use_container_width=True, key='sh_tar_editor',
            )
            # Construir tarifas_dict por mes del año
            sh_tarifas_dict = {}
            if _demand_ok:
                _df_d = st.session_state['sh_demand_df']
                for _, row in _df_d.drop_duplicates(subset=['Fecha']).iterrows():
                    m_idx = row['Fecha'].month - 1
                    key   = f"{row['Fecha'].year}-{row['Fecha'].month}"
                    if key not in sh_tarifas_dict:
                        sh_tarifas_dict[key] = {
                            'B':    sh_tar_df['B ($/kWh)'].iloc[m_idx],
                            'I':    sh_tar_df['I ($/kWh)'].iloc[m_idx],
                            'P':    sh_tar_df['P ($/kWh)'].iloc[m_idx],
                            'dist': sh_tar_df['Dist ($/kW)'].iloc[m_idx],
                            'cap':  sh_tar_df['Cap ($/kW)'].iloc[m_idx],
                            'trans':sh_tar_df['Trans ($/kWh)'].iloc[m_idx],
                        }
            else:
                sh_tarifas_dict = {}

        else:  # Suministro calificado
            _info('Ingresa los componentes del suministro calificado.')
            c9a, c9b = st.columns(2)
            with c9a:
                sh_sc_B = st.number_input('Precio energía Base ($/kWh)',   0.01, 5.0, 0.98, 0.01, format='%.4f', key='sh_sc_B')
                sh_sc_I = st.number_input('Precio energía Interm. ($/kWh)',0.01, 5.0, 1.75, 0.01, format='%.4f', key='sh_sc_I')
                sh_sc_P = st.number_input('Precio energía Punta ($/kWh)', 0.01, 5.0, 2.05, 0.01, format='%.4f', key='sh_sc_P')
            with c9b:
                sh_sc_dist  = st.number_input('Cargo distribución ($/kW/mes)', 0.0, 200.0, 38.5, 0.5,  format='%.2f', key='sh_sc_dist')
                sh_sc_cap   = st.number_input('Cargo capacidad ($/kW/mes)',     0.0, 200.0, 58.2, 0.5,  format='%.2f', key='sh_sc_cap')
                sh_sc_trans = st.number_input('Cargo transporte ($/kWh)',       0.0,   1.0, 0.21, 0.01, format='%.4f', key='sh_sc_trans')
                sh_sc_pot   = st.number_input('Cargo potencia horas críticas ($/kW)',
                                               0.0, 500.0, 0.0, 1.0, format='%.2f', key='sh_sc_pot',
                                               help='Se aplica sobre la demanda promedio en horas críticas definidas')

            sh_tarifas_dict = {}
            if _demand_ok:
                _df_d = st.session_state['sh_demand_df']
                for _, row in _df_d.drop_duplicates(subset=['Fecha']).iterrows():
                    key = f"{row['Fecha'].year}-{row['Fecha'].month}"
                    if key not in sh_tarifas_dict:
                        sh_tarifas_dict[key] = {
                            'B':    sh_sc_B,   'I': sh_sc_I,
                            'P':    sh_sc_P,   'dist': sh_sc_dist,
                            'cap':  sh_sc_cap, 'trans': sh_sc_trans,
                        }

        # ── Bloque 10: Horas críticas (opcional) ─────────────────────────────
        _sh('10 · Horas críticas de potencia (opcional)')
        sh_use_critical = st.checkbox(
            'Definir horas críticas de potencia', value=False, key='sh_use_crit',
        )

        if sh_use_critical:
            _info(
                'En horas críticas el BESS descarga al máximo posible '
                'independientemente del periodo tarifario, para minimizar '
                'la demanda CFE registrada.'
            )
            sh_crit_soc_min = st.slider(
                'SOC mínimo en horas críticas (kWh)',
                0, int(sh_bess_kwh) if sh_use_bess else 0,
                int(sh_bess_dod_min) if sh_use_bess else 0,
                10, key='sh_crit_soc_min',
            )

            # Gestión de bloques
            st.markdown('**Bloques de horas críticas**')
            _DIAS_MAP = {'Lun': 0, 'Mar': 1, 'Mié': 2, 'Jue': 3,
                         'Vie': 4, 'Sáb': 5, 'Dom': 6}
            _MESES_MAP = {m: i + 1 for i, m in enumerate(MONTHS)}

            bloques_crit = st.session_state['sh_bloques_crit']

            with st.expander('➕ Agregar bloque de horas críticas', expanded=len(bloques_crit) == 0):
                bc_dias    = st.multiselect('Días de la semana', list(_DIAS_MAP.keys()),
                                             default=['Lun', 'Mar', 'Mié', 'Jue', 'Vie'],
                                             key='sh_bc_dias')
                bc_meses   = st.multiselect('Meses', MONTHS, default=['Jun', 'Jul'], key='sh_bc_meses')
                bc_h_ini, bc_h_fin = st.slider('Rango de horas', 1, 24, (19, 22), key='sh_bc_horas')
                bc_etq     = st.text_input('Etiqueta (opcional)', value='Pico crítico', key='sh_bc_etq')

                if st.button('✅ Agregar bloque', key='sh_add_bloque'):
                    bloques_crit.append({
                        'dias':     {_DIAS_MAP[d] for d in bc_dias},
                        'meses':    {_MESES_MAP[m] for m in bc_meses},
                        'h_ini':    bc_h_ini,
                        'h_fin':    bc_h_fin,
                        'etiqueta': bc_etq,
                    })
                    st.session_state['sh_bloques_crit'] = bloques_crit
                    st.rerun()

            if bloques_crit:
                for i_b, blq in enumerate(bloques_crit):
                    dias_names = [k for k, v in _DIAS_MAP.items() if v in blq['dias']]
                    mes_names  = [k for k, v in _MESES_MAP.items() if v in blq['meses']]
                    st.markdown(
                        f'**Bloque {i_b+1}:** {", ".join(dias_names)} · '
                        f'{", ".join(mes_names)} · Horas {blq["h_ini"]}–{blq["h_fin"]} · '
                        f'_{blq["etiqueta"]}_'
                    )
                if st.button('🗑 Limpiar todos los bloques', key='sh_clear_bloques'):
                    st.session_state['sh_bloques_crit'] = []
                    st.rerun()
        else:
            sh_crit_soc_min = sh_bess_dod_min if sh_use_bess else 0.0
            bloques_crit = []

        # ── Bloque 11: Financiero ─────────────────────────────────────────────
        _sh('11 · Parámetros financieros')
        c11a, c11b = st.columns(2)
        with c11a:
            sh_horizonte  = st.radio('Horizonte (años)', [10, 15, 20, 25],
                                      index=1, horizontal=True, key='sh_horizonte')
            sh_tasa       = st.number_input('Tasa de descuento (%)', 5.0, 25.0, 10.0, 0.5,
                                            format='%.1f', key='sh_tasa')
            sh_esc_cfe    = st.number_input('Escalación tarifas CFE (%/año)', 0.0, 15.0, 0.0, 0.5,
                                            format='%.1f', key='sh_esc_cfe')
        with c11b:
            sh_esc_gas    = st.number_input('Escalación precio gas (%/año)', 0.0, 15.0, 0.0, 0.5,
                                            format='%.1f', key='sh_esc_gas')
            sh_om_pct     = st.number_input('O&M (% CAPEX/año)', 0.5, 5.0, 1.5, 0.25,
                                            format='%.2f', key='sh_om_pct')

        # ── Bloque 12: Modo de análisis ───────────────────────────────────────
        _sh('12 · Modo de análisis')
        sh_modo = st.radio(
            'Modo', ['Diseño manual', 'Optimización (grid search)'],
            key='sh_modo',
        )

        if sh_modo == 'Optimización (grid search)':
            _info('Define rangos. El sistema evaluará todas las combinaciones '
                  'y presentará el <b>Top 5 por VPN</b>.')
            if sh_use_chp:
                c_gs1, c_gs2, c_gs3 = st.columns(3)
                sh_gs_chp_min = c_gs1.number_input('CHP mín (kW)', 500, 5000, 1000, 500, key='sh_gs_chp_min')
                sh_gs_chp_max = c_gs2.number_input('CHP máx (kW)', 500, 10000, 3000, 500, key='sh_gs_chp_max')
                sh_gs_chp_step= c_gs3.number_input('Paso (kW)',    100, 2000, 500, 100, key='sh_gs_chp_step')
            if sh_use_pv:
                c_gs4, c_gs5, c_gs6 = st.columns(3)
                sh_gs_pv_min  = c_gs4.number_input('PV mín (kWp)', 0, 5000, 0, 250, key='sh_gs_pv_min')
                sh_gs_pv_max  = c_gs5.number_input('PV máx (kWp)', 0, 20000, 2000, 250, key='sh_gs_pv_max')
                sh_gs_pv_step = c_gs6.number_input('Paso (kWp)',   50, 2000, 500, 50, key='sh_gs_pv_step')
            if sh_use_bess:
                c_gs7, c_gs8, c_gs9 = st.columns(3)
                sh_gs_bess_min = c_gs7.number_input('BESS mín (kWh)', 0, 5000, 0, 250, key='sh_gs_bess_min')
                sh_gs_bess_max = c_gs8.number_input('BESS máx (kWh)', 0, 20000, 2000, 250, key='sh_gs_bess_max')
                sh_gs_bess_step= c_gs9.number_input('Paso (kWh)',     50, 2000, 500, 50, key='sh_gs_bess_step')

            _n_chp  = max(1, int((sh_gs_chp_max  - sh_gs_chp_min)  / max(sh_gs_chp_step,  1) + 1)) if sh_use_chp  else 1
            _n_pv   = max(1, int((sh_gs_pv_max   - sh_gs_pv_min)   / max(sh_gs_pv_step,   1) + 1)) if sh_use_pv   else 1
            _n_bess = max(1, int((sh_gs_bess_max - sh_gs_bess_min) / max(sh_gs_bess_step, 1) + 1)) if sh_use_bess else 1
            _n_comb = _n_chp * _n_pv * _n_bess
            st.caption(f'**{_n_comb:,}** combinaciones a evaluar · ~{_n_comb * 0.8:.0f}s estimado')

        # ── Botones de acción ─────────────────────────────────────────────────
        st.markdown('---')
        _can_run = _demand_ok and _tmy_ok

        if not _can_run:
            _warn('⚠️ Carga el archivo de demanda (Bloque 2) y construye el TMY '
                  '(Bloque 3) para habilitar la simulación.')

        btn_col1, btn_col2, btn_col3 = st.columns([2, 2, 1])
        with btn_col1:
            sh_btn_sim = st.button(
                '▶ Simular escenario',
                type='primary', use_container_width=True,
                disabled=not _can_run, key='sh_btn_sim',
            )
        with btn_col2:
            sh_esc_nombre = st.text_input(
                'Nombre del escenario', value='Escenario 1',
                label_visibility='collapsed', key='sh_esc_nombre',
            )
            sh_btn_add = st.button(
                '+ Guardar como escenario',
                use_container_width=True,
                disabled=not _can_run, key='sh_btn_add',
            )
        with btn_col3:
            if st.button('🗑 Limpiar', use_container_width=True, key='sh_btn_clear'):
                st.session_state['sh_escenarios'] = {}
                st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA DE SIMULACIÓN (se ejecuta al pulsar Simular o Agregar)
    # ══════════════════════════════════════════════════════════════════════════

    def _run_simulation(
        label:    str,
        chp_kw:   float,
        pv_kwp:   float,
        bess_kwh: float,
    ) -> dict | None:
        """Corre una simulación completa y devuelve el resultado."""

        df_d   = st.session_state['sh_demand_df'].copy()
        tmy_df = st.session_state['sh_tmy_df']

        n_rows = len(df_d)

        # ── Merge demanda con TMY por mes y hora ──────────────────────────────
        df_d['_mes']  = df_d['Fecha'].dt.month
        df_d['_hora'] = df_d['Hora'].astype(int)

        # Promedio GHI y temp del TMY por (mes, hora)
        tmy_df2 = tmy_df.copy()
        tmy_df2['_mes']  = tmy_df2['datetime'].dt.month
        tmy_df2['_hora'] = tmy_df2['datetime'].dt.hour + 1  # 1-24
        tmy_avg = tmy_df2.groupby(['_mes', '_hora'])[['irradiance_Wm2', 'temp_C']].mean()

        ghi_arr  = np.zeros(n_rows)
        temp_arr = np.full(n_rows, 20.0)
        for i, row in df_d.iterrows():
            try:
                ghi_arr[i]  = tmy_avg.loc[(row['_mes'], row['_hora']), 'irradiance_Wm2']
                temp_arr[i] = tmy_avg.loc[(row['_mes'], row['_hora']), 'temp_C']
            except KeyError:
                pass

        # ── PV generación hora a hora ─────────────────────────────────────────
        if sh_use_pv and pv_kwp > 0:
            pv_kw_arr = np.array(simulate_pv_8760(
                irr_8760  = tuple(ghi_arr),
                temp_8760 = tuple(temp_arr),
                kwp       = pv_kwp,
                pr_base   = sh_pv_pr,
                panel_temp_coeff = sh_t_coef / 100,
                t_noct    = sh_noct,
            ))
        else:
            pv_kw_arr = np.zeros(n_rows)

        # ── Sustitución de atípicos ───────────────────────────────────────────
        demand_orig = df_d['Consumo_kWh'].values.copy()
        demand_arr  = demand_orig.copy()
        sust_arr    = np.zeros(n_rows, dtype=bool)

        if sh_sust:
            _media   = demand_orig.mean()
            _thresh  = _media * sh_umbral / 100
            _df_dias = df_d.groupby(df_d['Fecha'].dt.date)['Consumo_kWh'].mean()
            _dias_ex = set(_df_dias[_df_dias < _thresh].index)

            # Calcular promedio de sustitución
            _df_norm = df_d[~df_d['Fecha'].dt.date.isin(_dias_ex)].copy()
            _df_norm['_weekday'] = _df_norm['Fecha'].dt.dayofweek
            _df_norm['_hora2']   = _df_norm['Hora'].astype(int)

            if sh_metodo == 'Promedio Lun–Sáb':
                _avg_ref = (_df_norm[_df_norm['_weekday'] < 6]
                            .groupby('_hora2')['Consumo_kWh'].mean())
            elif sh_metodo == 'Mismo día de semana':
                _avg_ref = (_df_norm
                            .groupby(['_weekday', '_hora2'])['Consumo_kWh'].mean())
            else:
                _avg_ref = _df_norm.groupby('_hora2')['Consumo_kWh'].mean()

            for i, row in df_d.iterrows():
                if row['Fecha'].date() in _dias_ex:
                    hora_i = int(row['Hora'])
                    wd     = row['Fecha'].weekday()
                    try:
                        if sh_metodo == 'Mismo día de semana':
                            val = _avg_ref.loc[(wd, hora_i)]
                        else:
                            val = _avg_ref.loc[hora_i]
                    except KeyError:
                        val = _media
                    demand_arr[i] = round(float(val), 2)
                    sust_arr[i]   = True

        # ── Arrays auxiliares ─────────────────────────────────────────────────
        periodo_arr = df_d['PeriodoTarifario'].values
        mes_arr     = df_d['Fecha'].dt.month.values
        key_arr     = np.array([
            f"{df_d['Fecha'].iloc[i].year}-{df_d['Fecha'].iloc[i].month}"
            for i in range(n_rows)
        ])

        # ── Máscara de horas críticas ─────────────────────────────────────────
        if sh_use_critical and bloques_crit:
            _crit_mask, _crit_labels = build_critical_mask(df_d, bloques_crit)
        else:
            _crit_mask  = np.zeros(n_rows, dtype=bool)
            _crit_labels = np.full(n_rows, '')

        # Inyectar etiquetas en key_arr no — las inyectamos en rows después
        _crit_soc_min = float(sh_crit_soc_min) if sh_use_critical else float(sh_bess_dod_min if sh_use_bess else 0.0)

        # ── Parámetros BESS ───────────────────────────────────────────────────
        _bess_kwh     = bess_kwh if sh_use_bess else 0.0
        _bess_kw      = (_bess_kwh / _ratio_map.get(sh_bess_ratio, 2)) if sh_use_bess else 0.0
        _soc_res      = (_bess_kwh * sh_bess_res_pct / 100) if sh_use_bess else 0.0
        _bess_dod_min = (_bess_kwh * (1 - sh_bess_dod / 100)) if sh_use_bess else 0.0

        # ── CHP ───────────────────────────────────────────────────────────────
        _chp_kw       = chp_kw if sh_use_chp else 0.0
        _chp_disp_f   = sh_chp_disp / 100 if sh_use_chp else 0.0
        _gas_mxn_kwh2 = gas_usdmmbtu_to_mxn_kwh(sh_gas_price, sh_tc)
        _chp_var      = (_gas_mxn_kwh2 / (sh_chp_eff_e / 100) + sh_chp_om) if sh_use_chp else 0.0

        # ── Baseline CFE ──────────────────────────────────────────────────────
        baseline_cfe = sum(
            demand_arr[i] * (
                sh_tarifas_dict.get(key_arr[i], {}).get(periodo_arr[i], 0.0) +
                sh_tarifas_dict.get(key_arr[i], {}).get('trans', 0.0)
            )
            for i in range(n_rows)
        ) + sum(
            max((demand_arr[i] for i in range(n_rows) if key_arr[i] == k), default=0) *
            (sh_tarifas_dict.get(k, {}).get('dist', 0.0) +
             sh_tarifas_dict.get(k, {}).get('cap', 0.0))
            for k in set(key_arr)
        )

        # ── Simulación ────────────────────────────────────────────────────────
        res = simulate_gdmth(
            demand_arr   = demand_arr,
            pv_arr       = pv_kw_arr,
            periodo_arr  = periodo_arr,
            mes_arr      = mes_arr,
            key_arr      = key_arr,
            tarifas_dict = sh_tarifas_dict,
            chp_kw       = _chp_kw,
            chp_disp     = _chp_disp_f,
            chp_var_cost = _chp_var,
            chp_fijo     = sh_chp_fijo,
            include_chp  = sh_use_chp,
            include_pv   = sh_use_pv,
            include_bess = sh_use_bess,
            bess_kwh     = _bess_kwh,
            bess_kw      = _bess_kw,
            soc_reserva  = _soc_res,
            bess_eff     = sh_bess_eff,
            bess_dod_min = _bess_dod_min,
            discharge_periods = set(sh_dis_per),
            charge_from_grid  = sh_grid_charge,
            lookahead_h       = sh_lookahead,
            chp_excess_charge = sh_chp_excess,
            critical_mask     = _crit_mask,
            critical_soc_min  = _crit_soc_min,
            sust_arr          = sust_arr,
            orig_arr          = demand_orig,
        )

        # ── Post-proceso: añadir etiquetas críticas y datos solares a rows ────
        for i, r in enumerate(res['rows']):
            r['Hora_Critica'] = _crit_labels[i] if _crit_labels[i] else (
                'SÍ' if _crit_mask[i] else '')
            r['GHI_Wm2'] = round(float(ghi_arr[i]), 1)
            r['Temp_C']  = round(float(temp_arr[i]), 1)
            r['PV_kW']   = round(float(pv_kw_arr[i]), 2)
            r['Fecha']   = str(df_d['Fecha'].iloc[i].date())
            r['Hora']    = int(df_d['Hora'].iloc[i])

        # ── Cálculo financiero ────────────────────────────────────────────────
        savings_cfe = baseline_cfe - res['total_cost']

        # Cobertura resiliencia
        _chp_desp  = _chp_kw * _chp_disp_f
        _diferencial = max(0.0, sh_carga_diseno - _chp_desp)
        if _diferencial > 0 and sh_use_bess:
            _auto_min = (_soc_res / _diferencial) * 60
        elif _diferencial == 0:
            _auto_min = 9999.0
        else:
            _auto_min = 0.0

        if _auto_min >= 60:   cov = 0.99
        elif _auto_min >= 30: cov = 0.97
        elif _auto_min >= 15: cov = 0.92
        else:                 cov = 0.85

        benef_res   = sh_total_res * cov
        total_benef = savings_cfe + benef_res

        # CAPEX
        capex_chp  = (_chp_kw * sh_chp_capex_kw)  if sh_use_chp  else 0.0
        capex_pv   = (pv_kwp  * sh_pv_capex)       if sh_use_pv   else 0.0
        capex_bess = (_bess_kwh * sh_bess_capex)    if sh_use_bess else 0.0
        capex_total = capex_chp + capex_pv + capex_bess

        # VPN / TIR / payback
        r_d = sh_tasa / 100
        inv = capex_total
        om_anual = inv * sh_om_pct / 100

        fn_y  = [total_benef - om_anual] * sh_horizonte
        fd_y  = [fn_y[i] / (1 + r_d) ** (i + 1) for i in range(sh_horizonte)]
        vpn   = -inv + sum(fd_y)
        tir   = _bisection_irr([-inv] + fn_y)
        pb    = None
        acum  = -inv
        for i, fn in enumerate(fn_y):
            prev = acum; acum += fn
            if acum >= 0 and pb is None:
                pb = round(i + (-prev) / (acum - prev), 2)

        return dict(
            label       = label,
            chp_kw      = _chp_kw, pv_kwp=pv_kwp, bess_kwh=_bess_kwh,
            # Energía
            demand_total= demand_arr.sum(),
            chp_kwh     = res['chp_kwh'],
            pv_kwh      = res['pv_kwh'],
            cfe_kwh     = res['cfe_kwh'],
            curtail     = res['curtail'],
            autogen_pct = res['autogen_pct'],
            fp_energia  = res['fp_energia'],
            # Costos
            cfe_e       = res['cfe_e'],
            cfe_t       = res['cfe_t'],
            demand_cost = res['demand_cost'],
            chp_f       = res['chp_f'],
            total_cost  = res['total_cost'],
            baseline_cfe= baseline_cfe,
            # BESS
            bess_ch_solar= res['bess_ch_solar'],
            bess_ch_chp  = res['bess_ch_chp'],
            bess_ch_red  = res['bess_ch_red'],
            bess_ch_tot  = res['bess_ch_tot'],
            bess_dis_P   = res['bess_dis_P'],
            bess_dis_I   = res['bess_dis_I'],
            bess_dis_B   = res['bess_dis_B'],
            bess_dis_crit= res['bess_dis_crit'],
            # Resiliencia
            auto_min    = _auto_min,
            cov         = cov,
            benef_res   = benef_res,
            # Financiero
            savings_cfe = savings_cfe,
            total_benef = total_benef,
            capex_chp   = capex_chp,
            capex_pv    = capex_pv,
            capex_bess  = capex_bess,
            capex_total = capex_total,
            om_anual    = om_anual,
            vpn         = vpn,
            tir         = tir,
            payback     = pb,
            horizonte   = sh_horizonte,
            fn_y        = fn_y,
            fd_y        = fd_y,
            # Datos para Excel
            rows        = res['rows'],
            demand_df   = df_d,
            pv_kw_arr   = pv_kw_arr,
            ghi_arr     = ghi_arr,
            temp_arr    = temp_arr,
            soc_reserva = _soc_res,
            has_sust    = bool(sust_arr.any()),
            has_critical= bool(_crit_mask.any()),
        )

    # ── Trigger simulación ────────────────────────────────────────────────────
    _last_result = None

    if (sh_btn_sim or sh_btn_add) and _can_run:
        _esc_label = sh_esc_nombre if sh_btn_add else '__preview__'

        if sh_modo == 'Diseño manual':
            with st.spinner('Simulando…'):
                _last_result = _run_simulation(
                    label    = _esc_label,
                    chp_kw   = float(sh_chp_kw) if sh_use_chp else 0.0,
                    pv_kwp   = float(sh_pv_kwp)  if sh_use_pv  else 0.0,
                    bess_kwh = float(sh_bess_kwh) if sh_use_bess else 0.0,
                )
        else:
            # Grid search
            _chp_range  = (list(range(int(sh_gs_chp_min),  int(sh_gs_chp_max)  + 1, int(sh_gs_chp_step)))
                           if sh_use_chp else [0])
            _pv_range   = (list(range(int(sh_gs_pv_min),   int(sh_gs_pv_max)   + 1, int(sh_gs_pv_step)))
                           if sh_use_pv else [0])
            _bess_range = (list(range(int(sh_gs_bess_min), int(sh_gs_bess_max) + 1, int(sh_gs_bess_step)))
                           if sh_use_bess else [0])

            _gs_results = []
            _prog = st.progress(0, text='Optimizando…')
            _total_gs = len(_chp_range) * len(_pv_range) * len(_bess_range)
            _count_gs = 0

            for _c in _chp_range:
                for _p in _pv_range:
                    for _b in _bess_range:
                        _r = _run_simulation(
                            label    = f'CHP{_c}_PV{_p}_BESS{_b}',
                            chp_kw   = float(_c),
                            pv_kwp   = float(_p),
                            bess_kwh = float(_b),
                        )
                        if _r:
                            _gs_results.append(_r)
                        _count_gs += 1
                        _prog.progress(_count_gs / _total_gs,
                                       text=f'Evaluando combinación {_count_gs}/{_total_gs}…')

            _prog.empty()
            _gs_results.sort(key=lambda x: x['vpn'], reverse=True)

            if _gs_results:
                _last_result = _gs_results[0]
                # Mostrar top 5 en col_out (se mostrará abajo)
                st.session_state['sh_gs_top5'] = _gs_results[:5]

        if _last_result:
            if sh_btn_add:
                st.session_state['sh_escenarios'][_esc_label] = _last_result
            st.session_state['sh_last_result'] = _last_result

    # ══════════════════════════════════════════════════════════════════════════
    # PANEL DERECHO — OUTPUTS
    # ══════════════════════════════════════════════════════════════════════════
    with col_out:

        _result = st.session_state.get('sh_last_result')

        if _result is None:
            st.markdown("""
<div style="text-align:center;padding:4rem 2rem;color:#475569;">
  <div style="font-size:48px;margin-bottom:1rem;">⚡</div>
  <div style="font-size:16px;font-weight:500;color:#94a3b8;">
    Configura el sistema y presiona <b>▶ Simular</b>
  </div>
  <div style="font-size:13px;margin-top:0.5rem;">
    Los resultados aparecerán aquí
  </div>
</div>""", unsafe_allow_html=True)
        else:
            R = _result  # alias corto

            # ── Comparativo de escenarios (si hay 2+) ─────────────────────────
            _escenarios = st.session_state.get('sh_escenarios', {})
            if len(_escenarios) >= 2:
                _sh('Comparativo de escenarios')
                _esc_list = list(_escenarios.values())
                _comp_data = {
                    'Métrica': [
                        'CHP (kW)', 'FV (kWp)', 'BESS (kWh)',
                        'Autogeneración (%)', 'FP CHP (%)',
                        'Ahorro CFE (MXN/año)', 'Benef. resiliencia',
                        'Beneficio total', 'CAPEX total',
                        'Payback (años)', f'VPN {_esc_list[0]["horizonte"]}a',
                        'Autonomía (min)', 'Cobertura eventos',
                    ]
                }
                for _e in _esc_list:
                    _comp_data[_e['label']] = [
                        f"{_e['chp_kw']:,.0f}",
                        f"{_e['pv_kwp']:,.0f}",
                        f"{_e['bess_kwh']:,.0f}",
                        f"{_e['autogen_pct']:.1f}%",
                        f"{_e['fp_energia']:.1%}",
                        f"${_e['savings_cfe']:,.0f}",
                        f"${_e['benef_res']:,.0f}",
                        f"${_e['total_benef']:,.0f}",
                        f"${_e['capex_total']:,.0f}",
                        f"{_e['payback']:.2f}" if _e['payback'] else '—',
                        f"${_e['vpn']:,.0f}",
                        f"{_e['auto_min']:.0f}" if _e['auto_min'] < 9999 else 'CHP cubre todo',
                        f"{_e['cov']:.0%}",
                    ]
                st.dataframe(
                    pd.DataFrame(_comp_data).set_index('Métrica'),
                    use_container_width=True,
                )

            # ── Top 5 grid search ─────────────────────────────────────────────
            if 'sh_gs_top5' in st.session_state:
                _sh('Top 5 combinaciones por VPN (optimización)')
                _top5 = st.session_state['sh_gs_top5']
                _top5_df = pd.DataFrame([{
                    'Config':    r['label'],
                    'CHP (kW)':  f"{r['chp_kw']:,.0f}",
                    'PV (kWp)':  f"{r['pv_kwp']:,.0f}",
                    'BESS (kWh)':f"{r['bess_kwh']:,.0f}",
                    'FP CHP':    f"{r['fp_energia']:.1%}",
                    'Ahorro CFE':f"${r['savings_cfe']/1e6:.2f}M",
                    'Benef. total':f"${r['total_benef']/1e6:.2f}M",
                    'CAPEX':     f"${r['capex_total']/1e6:.1f}M",
                    'Payback':   f"{r['payback']:.2f}a" if r['payback'] else '—',
                    'VPN':       f"${r['vpn']/1e6:.1f}M",
                } for r in _top5])
                st.dataframe(_top5_df.set_index('Config'), use_container_width=True)

            # ── KPIs ejecutivos ───────────────────────────────────────────────
            _sh(f'Resultados — {R["label"]}')

            k1, k2, k3, k4 = st.columns(4)
            k1.metric('Factura base CFE',  f"${R['baseline_cfe']/1e6:.1f}M", 'MXN/año')
            k2.metric('Costo sistema',     f"${R['total_cost']/1e6:.1f}M",   'MXN/año')
            k3.metric('Ahorro CFE',        f"${R['savings_cfe']/1e6:.1f}M",  f"{R['savings_cfe']/max(R['baseline_cfe'],1)*100:.1f}%")
            k4.metric('Beneficio total',   f"${R['total_benef']/1e6:.1f}M",  'incl. resiliencia')

            k5, k6, k7, k8 = st.columns(4)
            k5.metric('CAPEX total',       f"${R['capex_total']/1e6:.1f}M",  'MXN')
            k6.metric('Payback simple',    f"{R['payback']:.2f} años" if R['payback'] else '—', '')
            k7.metric(f"VPN {R['horizonte']}a",  f"${R['vpn']/1e6:.1f}M",   f"TIR {R['tir']:.1f}%" if R['tir'] else '')
            k8.metric('FP CHP',            f"{R['fp_energia']:.1%}",          f"Auto {R['autogen_pct']:.1f}%")

            # ── Balance energético ────────────────────────────────────────────
            _sh('Balance energético anual')
            lay_bal = copy.deepcopy(PLOT_LAYOUT)
            fig_bal = go.Figure(data=[go.Bar(
                x=['CHP', 'FV aprovechado', 'CFE importado', 'Curtailment'],
                y=[R['chp_kwh']/1e6,
                   (R['pv_kwh'] - R['curtail'])/1e6,
                   R['cfe_kwh']/1e6,
                   R['curtail']/1e6],
                marker_color=['#1A5C20', '#C44A00', '#9B1C1C', '#475569'],
                text=[f"{v:.2f} GWh" for v in [
                    R['chp_kwh']/1e6,
                    (R['pv_kwh'] - R['curtail'])/1e6,
                    R['cfe_kwh']/1e6,
                    R['curtail']/1e6,
                ]],
                textposition='outside',
            )])
            lay_bal.update({'height': 260,
                            'yaxis': dict(title='GWh/año', gridcolor='#1e2230'),
                            'margin': dict(l=10, r=10, t=20, b=40)})
            fig_bal.update_layout(**lay_bal)
            st.plotly_chart(fig_bal, use_container_width=True)

            # ── Desglose de costos ────────────────────────────────────────────
            _sh('Desglose de costos anuales')
            lay_cos = copy.deepcopy(PLOT_LAYOUT)
            fig_cos = go.Figure(data=[go.Bar(
                x=['CFE Energía', 'CFE Transporte', 'CFE Demanda', 'CHP Gas+O&M'],
                y=[R['cfe_e']/1e6, R['cfe_t']/1e6, R['demand_cost']/1e6, R['chp_f']/1e6],
                marker_color=['#E24B4A', '#888780', '#D85A30', '#1D9E75'],
                text=[f"${v:.1f}M" for v in [R['cfe_e']/1e6, R['cfe_t']/1e6,
                                               R['demand_cost']/1e6, R['chp_f']/1e6]],
                textposition='outside',
            )])
            lay_cos.update({'height': 240,
                            'yaxis': dict(title='MXN/año ($M)', gridcolor='#1e2230'),
                            'margin': dict(l=10, r=10, t=20, b=40)})
            fig_cos.update_layout(**lay_cos)
            st.plotly_chart(fig_cos, use_container_width=True)

            # ── VPN acumulado ─────────────────────────────────────────────────
            _sh(f'VPN acumulado a {R["horizonte"]} años')
            _vpn_cum = [-R['capex_total']]
            _acum = -R['capex_total']
            for fd in R['fd_y']:
                _acum += fd; _vpn_cum.append(round(_acum))
            lay_vpn = copy.deepcopy(PLOT_LAYOUT)
            fig_vpn = go.Figure(go.Scatter(
                x=list(range(R['horizonte'] + 1)),
                y=[v / 1e6 for v in _vpn_cum],
                mode='lines+markers',
                line=dict(color=AMBER, width=2.5),
                fill='tozeroy', fillcolor='rgba(245,158,11,0.07)',
                marker=dict(size=5),
            ))
            lay_vpn.update({'height': 230,
                            'xaxis': dict(title='Año', gridcolor='#1e2230'),
                            'yaxis': dict(title='VPN acum. ($M MXN)', gridcolor='#1e2230'),
                            'margin': dict(l=10, r=10, t=20, b=40)})
            fig_vpn.update_layout(**lay_vpn)
            st.plotly_chart(fig_vpn, use_container_width=True)

            # ── Perfil horario promedio ───────────────────────────────────────
            _sh('Perfil horario promedio (despacho)')
            _rows_df = pd.DataFrame(R['rows'])
            _h_avg = _rows_df.groupby('Hora')[
                ['Demanda_kW', 'CHP_kW', 'PV_kW', 'BESS_Descarga_kW', 'CFE_kW']
            ].mean().reset_index()

            lay_hr = copy.deepcopy(PLOT_LAYOUT)
            fig_hr = go.Figure()
            fig_hr.add_trace(go.Scatter(
                x=_h_avg['Hora'], y=_h_avg['Demanda_kW'],
                name='Demanda', mode='lines',
                line=dict(color=ROSE, width=2, dash='dot'),
            ))
            if R['chp_kw'] > 0:
                fig_hr.add_trace(go.Bar(
                    x=_h_avg['Hora'], y=_h_avg['CHP_kW'],
                    name='CHP', marker_color='#1A5C20', opacity=0.85,
                ))
            if R['pv_kwp'] > 0:
                fig_hr.add_trace(go.Bar(
                    x=_h_avg['Hora'], y=_h_avg['PV_kW'],
                    name='FV Solar', marker_color='#C44A00', opacity=0.85,
                ))
            if R['bess_kwh'] > 0:
                fig_hr.add_trace(go.Bar(
                    x=_h_avg['Hora'], y=_h_avg['BESS_Descarga_kW'],
                    name='BESS', marker_color='#4A148C', opacity=0.85,
                ))
            fig_hr.add_trace(go.Bar(
                x=_h_avg['Hora'], y=_h_avg['CFE_kW'],
                name='CFE', marker_color='#9B1C1C', opacity=0.85,
            ))
            lay_hr.update({'height': 280, 'barmode': 'stack',
                           'xaxis': dict(title='Hora del día', gridcolor='#1e2230',
                                         tickmode='linear', dtick=2),
                           'yaxis': dict(title='kW promedio', gridcolor='#1e2230'),
                           'legend': dict(orientation='h', y=1.12, bgcolor='rgba(0,0,0,0)'),
                           'margin': dict(l=10, r=10, t=40, b=40)})
            fig_hr.update_layout(**lay_hr)
            st.plotly_chart(fig_hr, use_container_width=True)

            # ── Curva de duración ─────────────────────────────────────────────
            _sh('Curva de duración — CFE importado')
            fig_dur = duration_curve_fig(
                _rows_df['CFE_kW'].tolist(), 'CFE importado', ROSE, 'kW',
            )
            st.plotly_chart(fig_dur, use_container_width=True)

            # ── Resiliencia ───────────────────────────────────────────────────
            _sh('Resiliencia')
            r1, r2, r3, r4 = st.columns(4)
            r1.metric('Carga esencial',   f"{sh_carga_diseno:,.0f} kW",  f'+{sh_margen_seg}% margen')
            r2.metric('Diferencial BESS', f"{max(0, sh_carga_diseno - R['chp_kw'] * sh_chp_disp/100):,.0f} kW")
            r3.metric('Autonomía',
                       f"{R['auto_min']:.0f} min" if R['auto_min'] < 9999 else 'CHP completo',
                       f"Reserva {R['soc_reserva']:,.0f} kWh" if R['bess_kwh'] > 0 else '')
            r4.metric('Cobertura eventos', f"{R['cov']:.0%}",
                       f"${R['benef_res']/1e6:.1f}M MXN/año")

            # ── BESS fuentes de carga ─────────────────────────────────────────
            if R['bess_kwh'] > 0:
                _sh('BESS — Fuentes de carga y descarga')
                b1, b2 = st.columns(2)
                with b1:
                    _tot_ch = max(R['bess_ch_tot'], 0.001)
                    st.markdown(f"""
<div class="panel-card">
  <div class="pc-title">Carga anual del BESS</div>
  <div class="pc-grid">
    <div class="pc-item"><span class="pc-label">☀️ Solar</span>
      <span class="pc-val">{R['bess_ch_solar']:,.0f} kWh ({R['bess_ch_solar']/_tot_ch*100:.0f}%)</span></div>
    <div class="pc-item"><span class="pc-label">⚡ CHP excedente</span>
      <span class="pc-val">{R['bess_ch_chp']:,.0f} kWh ({R['bess_ch_chp']/_tot_ch*100:.0f}%)</span></div>
    <div class="pc-item"><span class="pc-label">🔌 Red CFE-B</span>
      <span class="pc-val">{R['bess_ch_red']:,.0f} kWh ({R['bess_ch_red']/_tot_ch*100:.0f}%)</span></div>
    <div class="pc-item"><span class="pc-label">Total cargado</span>
      <span class="pc-val">{_tot_ch:,.0f} kWh/año</span></div>
  </div>
</div>""", unsafe_allow_html=True)
                with b2:
                    _tot_dis = max(R['bess_dis_P'] + R['bess_dis_I'] + R['bess_dis_B'], 0.001)
                    st.markdown(f"""
<div class="panel-card">
  <div class="pc-title">Descarga anual del BESS</div>
  <div class="pc-grid">
    <div class="pc-item"><span class="pc-label">Punta (P)</span>
      <span class="pc-val">{R['bess_dis_P']:,.0f} kWh</span></div>
    <div class="pc-item"><span class="pc-label">Intermedio (I)</span>
      <span class="pc-val">{R['bess_dis_I']:,.0f} kWh</span></div>
    <div class="pc-item"><span class="pc-label">Horas críticas</span>
      <span class="pc-val">{R['bess_dis_crit']:,.0f} kWh</span></div>
    <div class="pc-item"><span class="pc-label">Total descargado</span>
      <span class="pc-val">{_tot_dis:,.0f} kWh/año</span></div>
  </div>
</div>""", unsafe_allow_html=True)

            # ── Exportar Excel ────────────────────────────────────────────────
            _sh('Exportar despacho horario')
            if st.button('📊 Generar Excel de despacho', type='primary',
                          use_container_width=True, key='sh_btn_excel'):
                with st.spinner('Generando Excel…'):
                    _xlsx = build_dispatch_excel(
                        rows         = R['rows'],
                        demand_df    = R['demand_df'],
                        pv_kw_arr    = R['pv_kw_arr'],
                        ghi_arr      = R['ghi_arr'],
                        temp_arr     = R['temp_arr'],
                        include_chp  = sh_use_chp,
                        include_pv   = sh_use_pv,
                        include_bess = sh_use_bess,
                        has_critical = R['has_critical'],
                        has_sust     = R['has_sust'],
                        proj_nombre  = sh_nombre,
                        chp_kw       = R['chp_kw'],
                        pv_kwp       = R['pv_kwp'],
                        bess_kwh     = R['bess_kwh'],
                        soc_reserva  = R['soc_reserva'],
                    )
                _fname = (
                    f"Despacho_{sh_nombre.replace(' ', '_')}"
                    f"_CHP{int(R['chp_kw'])}"
                    f"_PV{int(R['pv_kwp'])}"
                    f"_BESS{int(R['bess_kwh'])}.xlsx"
                )
                st.download_button(
                    '⬇️ Descargar Excel de despacho',
                    data=_xlsx,
                    file_name=_fname,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                )

            # ── Nota metodológica ─────────────────────────────────────────────
            _info(
                f'<b>Metodología:</b> '
                f'Despacho hora a hora · CHP {"fijo" if sh_chp_fijo else "sigue carga"} · '
                f'PV con corrección térmica IEC 61724 (NOCT {sh_noct}°C) · '
                f'BESS descarga {", ".join(sh_dis_per)} · '
                f'Reserva contingencia {sh_bess_res_pct}% · '
                f'Lookahead {sh_lookahead}h · '
                f'Tarifas {sh_modalidad} · '
                f'VPN a tasa {sh_tasa}% a {sh_horizonte} años'
            )
